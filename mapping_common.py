"""
통합 매핑 공통 모듈

SNUH, SNOMED 등 모든 데이터 소스에 공통 적용되는:
- 데이터 소스 설정 (기본 경로, 전처리)
- 로깅 설정
- JSON/LOG/XLSX 출력
"""

import json
import logging
from dataclasses import asdict, is_dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

# -----------------------------------------------------------------------------
# 데이터 소스 설정 (확장 가능: 새 데이터 추가 시 여기에 등록)
# -----------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent

DATA_SOURCES = {
    "snuh": {
        "csv_path": str(PROJECT_ROOT / "data" / "mapping_test_snuh_top10k.csv"),
        "vocabulary_filter": ["SNOMED", "LOINC"],  # 기본 전처리
        "domains": ["Condition", "Procedure", "Drug", "Observation", "Measurement", "Device"],
        "id_col": "snuh_id",
        "loader": "load_snuh_data",
        "row_to_input": "snuh_row_to_input",
    },
    "snomed": {
        "csv_path": str(PROJECT_ROOT / "data" / "mapping_test_snomed_no_note.csv"),
        "filter_domains": ["Condition", "Measurement", "Drug", "Observation", "Procedure"],  # 기본 전처리
        "domains": ["Condition", "Measurement", "Drug", "Observation", "Procedure"],
        "id_col": "note_id",
        "loader": "load_snomed_data",
        "row_to_input": "snomed_row_to_input",
    },
}

# openpyxl for xlsx
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# XLSX 열 구성 (사용자 요구사항)
XLSX_HEADERS = [
    "Test Index",
    "ID",
    "Entity Name",
    "Input Domain",
    "Ground Truth Concept ID",
    "Success",
    "Mapping Correct",
    "Best Result Domain",
    "Best Concept ID",
    "Best Concept Name",
    "Best Score",
    "Stage1 Candidates",
    "Stage2 Candidates",
    "Stage3 Candidates",
]


def setup_logging(output_dir: Path, data_type: str, timestamp: str) -> tuple[logging.Logger, Path]:
    """통합 로깅 설정. 동일 timestamp 사용."""
    output_dir.mkdir(parents=True, exist_ok=True)
    log_file = output_dir / f"mapping_{data_type}_{timestamp}.log"

    logger = logging.getLogger(f"mapping_{data_type}")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(logging.INFO)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    # MapOMOP API 로거들도 동일 파일에 기록
    for name in [
        "MapOMOP.entity_mapping_api",
        "MapOMOP.mapping_stages.stage1_candidate_retrieval",
        "MapOMOP.mapping_stages.stage2_standard_collection",
        "MapOMOP.mapping_stages.stage3_hybrid_scoring",
    ]:
        api_log = logging.getLogger(name)
        api_log.setLevel(logging.INFO)
        api_log.addHandler(file_handler)

    logger.info(f"로그 파일: {log_file}")
    return logger, log_file


def save_json(results: List[Dict], output_dir: Path, data_type: str, timestamp: str) -> Path:
    """매핑 raw 데이터를 JSON으로 저장."""
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"mapping_{data_type}_{timestamp}.json"

    def _serialize(obj: Any) -> Any:
        if is_dataclass(obj) and not isinstance(obj, type):
            return {k: _serialize(v) for k, v in asdict(obj).items()}
        if isinstance(obj, dict):
            return {k: _serialize(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_serialize(x) for x in obj]
        if isinstance(obj, (str, bool, type(None))):
            return obj
        if isinstance(obj, (int, float)):
            return obj
        # numpy 등
        try:
            return float(obj)
        except (TypeError, ValueError):
            pass
        try:
            return int(obj)
        except (TypeError, ValueError):
            pass
        return str(obj)

    serializable = _serialize(results)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(serializable, f, ensure_ascii=False, indent=2)
    return out_path


def _sort_stage1_by_score(candidates: List[Dict]) -> List[Dict]:
    """Stage1 후보를 점수(elasticsearch_score) 높은 순 정렬. lexical/semantic/combined 구분 없이."""
    if not candidates:
        return []
    return sorted(
        candidates,
        key=lambda x: float(x.get("elasticsearch_score") or x.get("_score") or 0),
        reverse=True,
    )


def _sort_stage2_by_score(candidates: List[Dict]) -> List[Dict]:
    """Stage2 후보를 elasticsearch_score 순 정렬."""
    if not candidates:
        return []
    return sorted(
        candidates,
        key=lambda x: float(x.get("elasticsearch_score") or 0),
        reverse=True,
    )


def _sort_stage3_by_score(candidates: List[Dict]) -> List[Dict]:
    """Stage3 후보를 final_score 순 정렬."""
    if not candidates:
        return []
    return sorted(
        candidates,
        key=lambda x: float(x.get("final_score") or 0),
        reverse=True,
    )


def _format_candidates_for_cell(candidates: List[Dict], stage_type: str) -> str:
    """후보군을 XLSX 셀용 텍스트로 포맷. 점수 높은 순 표시."""
    if not candidates:
        return "후보 없음"

    if stage_type == "stage1":
        sorted_candidates = _sort_stage1_by_score(candidates)
    elif stage_type == "stage2":
        sorted_candidates = _sort_stage2_by_score(candidates)
    else:
        sorted_candidates = _sort_stage3_by_score(candidates)

    lines = []
    max_show = 15 if stage_type in ("stage1", "stage2") else 10

    for i, c in enumerate(sorted_candidates[:max_show], 1):
        st = c.get("search_type", "unknown")
        name = c.get("concept_name", "N/A")
        cid = c.get("concept_id", "N/A")

        if stage_type == "stage1":
            es_score = float(c.get("elasticsearch_score") or c.get("_score") or 0)
            line = f"{i}. [{st}] {name} (ID: {cid})\n"
            line += f"   ES점수: {es_score:.4f}, Standard: {c.get('standard_concept', 'N/A')}, Domain: {c.get('domain_id', 'N/A')}"
        elif stage_type == "stage2":
            is_std = "✓" if c.get("is_original_standard", True) else "→"
            line = f"{i}. [{st}] {is_std} {name} (ID: {cid})\n"
            line += f"   Standard: {c.get('standard_concept', 'N/A')}, Domain: {c.get('domain_id', 'N/A')}"
            if not c.get("is_original_standard", True):
                ons = c.get("original_non_standard", {})
                if ons:
                    line += f"\n   원본 Non-std: {ons.get('concept_name', 'N/A')} (ID: {ons.get('concept_id', 'N/A')})"
        else:
            fin = float(c.get("final_score") or 0)
            sem = c.get("semantic_similarity")
            line = f"{i}. [{st}] {name} (ID: {cid})\n"
            line += f"   최종: {fin:.4f}"
            if sem is not None:
                line += f", 의미적: {sem:.4f}"
            line += f", Standard: {c.get('standard_concept', 'N/A')}, Domain: {c.get('domain_id', 'N/A')}"

        lines.append(line)

    return "\n\n".join(lines)


def save_xlsx(results: List[Dict], output_dir: Path, data_type: str, timestamp: str) -> Path:
    """매핑 결과를 XLSX로 저장. 열: Test Index, ID, Entity Name, Input Domain, ..."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for XLSX output. pip install openpyxl")

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"mapping_{data_type}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed Results"

    # 헤더
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(XLSX_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 데이터
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r.get("test_index", ""))
        ws.cell(row=row_idx, column=2, value=r.get("id", r.get("snuh_id", r.get("note_id", "N/A"))))
        ws.cell(row=row_idx, column=3, value=r.get("entity_name", ""))
        ws.cell(row=row_idx, column=4, value=r.get("input_domain", "All"))
        ws.cell(row=row_idx, column=5, value=r.get("ground_truth_concept_id", ""))
        ws.cell(row=row_idx, column=6, value="성공" if r.get("success") else "실패")

        correct_cell = ws.cell(row=row_idx, column=7, value="정답" if r.get("mapping_correct") else "오답")
        if r.get("mapping_correct"):
            correct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            correct_cell.font = Font(color="006100")
        else:
            correct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            correct_cell.font = Font(color="9C0006")

        ws.cell(row=row_idx, column=8, value=r.get("best_result_domain", "N/A"))
        ws.cell(row=row_idx, column=9, value=r.get("best_concept_id", "N/A"))
        ws.cell(row=row_idx, column=10, value=r.get("best_concept_name", "N/A"))
        ws.cell(row=row_idx, column=11, value=r.get("best_score", 0.0))

        # Stage1: 점수 높은 순 (lexical/semantic/combined 무관)
        stage1_text = _format_candidates_for_cell(r.get("stage1_candidates", []), "stage1")
        ws.cell(row=row_idx, column=12, value=stage1_text)

        stage2_text = _format_candidates_for_cell(r.get("stage2_candidates", []), "stage2")
        ws.cell(row=row_idx, column=13, value=stage2_text)

        stage3_text = _format_candidates_for_cell(r.get("stage3_candidates", []), "stage3")
        ws.cell(row=row_idx, column=14, value=stage3_text)

        for c in range(12, 15):
            ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    # 열 너비
    widths = {
        "A": 10, "B": 18, "C": 40, "D": 15, "E": 20, "F": 10, "G": 12,
        "H": 18, "I": 15, "J": 45, "K": 12, "L": 70, "M": 70, "N": 85,
    }
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    for rn in range(2, len(results) + 2):
        ws.row_dimensions[rn].height = 150

    wb.save(out_path)
    return out_path


# --- SNUH 데이터 로더 ---
def load_snuh_data(
    csv_path: str,
    sample_size: int = 10000,
    use_random: bool = False,
    random_state: int = 42,
    sample_per_domain: Optional[Dict[str, int]] = None,
    vocabulary_filter: Optional[List[str]] = None,
    chunk_size: int = 100000,
) -> pd.DataFrame:
    """SNUH CSV 로드. entity=source_name, domain=domain, gt=omop_concept_id, id=snuh_id.
    기본 전처리: vocabulary IN ('SNOMED', 'LOINC')
    """
    from tqdm import tqdm

    if vocabulary_filter is None:
        vocabulary_filter = ["SNOMED", "LOINC"]

    if sample_per_domain or vocabulary_filter:
        chunks = []
        for chunk in tqdm(
            pd.read_csv(csv_path, chunksize=chunk_size),
            desc="데이터 로딩",
        ):
            chunks.append(chunk)
        df = pd.concat(chunks, ignore_index=True)

        if vocabulary_filter and "vocabulary" in df.columns:
            df = df[df["vocabulary"].isin(vocabulary_filter)]

        if sample_per_domain:
            sampled = []
            for domain, size in sample_per_domain.items():
                subset = df[df["domain"] == domain]
                n = min(size, len(subset))
                if n == 0:
                    continue
                if use_random:
                    sampled.append(subset.sample(n=n, random_state=random_state))
                else:
                    sampled.append(subset.head(n))
            df = pd.concat(sampled, ignore_index=True)
            if use_random:
                df = df.sample(frac=1, random_state=random_state).reset_index(drop=True)
        else:
            n = min(sample_size, len(df))
            if use_random:
                df = df.sample(n=n, random_state=random_state)
            else:
                df = df.head(n)
    else:
        df = pd.read_csv(csv_path, nrows=sample_size)

    return df.reset_index(drop=True)


def snuh_row_to_input(row: pd.Series, domain_map: Dict[str, Any]) -> tuple[str, Optional[Any], str, Optional[int]]:
    """(entity_name, domain_id, record_id, ground_truth)"""
    entity = str(row["source_name"]).strip()
    domain_str = str(row["domain"]).strip() if pd.notna(row.get("domain")) else None
    domain_id = domain_map.get(domain_str) if domain_str else None
    rid = str(row["snuh_id"]) if pd.notna(row.get("snuh_id")) else "N/A"
    gt = int(row["omop_concept_id"]) if pd.notna(row.get("omop_concept_id")) else None
    return entity, domain_id, rid, gt


# --- SNOMED 데이터 로더 ---
def load_snomed_data(
    csv_path: str,
    sample_size: int = 10000,
    use_random: bool = False,
    random_state: int = 42,
    sample_per_domain: Optional[Dict[str, int]] = None,
    filter_domains: Optional[List[str]] = None,
    chunk_size: int = 100000,
) -> pd.DataFrame:
    """SNOMED CSV 로드. entity=entity_name, domain=domain_id, gt=concept_id, id=note_id.
    기본 전처리: domain_id IN ('Condition','Measurement','Drug','Observation','Procedure')
    """
    from tqdm import tqdm

    if filter_domains is None:
        filter_domains = ["Condition", "Measurement", "Drug", "Observation", "Procedure"]

    chunks = []
    for chunk in tqdm(pd.read_csv(csv_path, chunksize=chunk_size), desc="데이터 로딩"):
        chunks.append(chunk)
    df = pd.concat(chunks, ignore_index=True)

    if filter_domains and "domain_id" in df.columns:
        df = df[df["domain_id"].isin(filter_domains)]

    if sample_per_domain:
        sampled = []
        for domain, size in sample_per_domain.items():
            subset = df[df["domain_id"] == domain]
            n = min(size, len(subset))
            if n == 0:
                continue
            if use_random:
                sampled.append(subset.sample(n=n, random_state=random_state))
            else:
                sampled.append(subset.head(n))
        df = pd.concat(sampled, ignore_index=True)
        if use_random:
            df = df.sample(frac=1, random_state=random_state).reset_index(drop=True)
    else:
        n = min(sample_size, len(df))
        if use_random:
            df = df.sample(n=n, random_state=random_state)
        else:
            df = df.head(n)

    return df.reset_index(drop=True)


def snomed_row_to_input(row: pd.Series, domain_map: Dict[str, Any]) -> tuple[str, Optional[Any], str, Optional[int]]:
    """(entity_name, domain_id, record_id, ground_truth)"""
    entity = str(row["entity_name"]).strip()
    domain_str = str(row["domain_id"]).strip() if pd.notna(row.get("domain_id")) else None
    domain_id = domain_map.get(domain_str) if domain_str else None
    rid = str(row["note_id"]) if pd.notna(row.get("note_id")) else "N/A"
    gt = int(row["concept_id"]) if pd.notna(row.get("concept_id")) else None
    return entity, domain_id, rid, gt
