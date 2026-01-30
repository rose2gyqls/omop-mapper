import pandas as pd
import logging
import os
from datetime import datetime
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from tqdm import tqdm
import time

sys.path.append('/home/work/skku/hyo/MapOMOP/src')

from MapOMOP.entity_mapping_api import EntityMappingAPI, EntityInput, DomainID
from MapOMOP.elasticsearch_client import ElasticsearchClient

class RealDataEntityMappingTester:
    def __init__(self, log_dir: str = "test_logs_real_data", scoring_mode: str = "llm"):
        """실제 데이터 테스터 초기화
        
        Args:
            log_dir: 로그 디렉토리
            scoring_mode: Stage 3 점수 계산 방식 ('llm' 또는 'hybrid')
        """
        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(exist_ok=True)
        
        self.setup_logging()
        
        self.es_client = ElasticsearchClient()
        self.es_client.concept_index = "concept-small"
        self.es_client.concept_synonym_index = "concept-synonym"
        
        self.api = EntityMappingAPI(
            es_client=self.es_client,
            scoring_mode=scoring_mode
        )
        
        self.logger.info(f"✅ Scoring Mode: {scoring_mode.upper()}")
        
        self.domain_mapping = {
            'Condition': DomainID.CONDITION,
            'Procedure': DomainID.PROCEDURE,
            'Drug': DomainID.DRUG,
            'Observation': DomainID.OBSERVATION,
            'Measurement': DomainID.MEASUREMENT,
            'Period': DomainID.PERIOD,
            'Provider': DomainID.PROVIDER,
            'Device': DomainID.DEVICE,
        }
    
    def setup_logging(self):
        """로깅 설정"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        log_file = self.log_dir / f"entity_mapping_real_data_{timestamp}.log"
        
        self.logger = logging.getLogger('entity_mapping_real_data')
        self.logger.setLevel(logging.INFO)
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        # API 로거들도 파일에 기록
        api_logger = logging.getLogger('MapOMOP.entity_mapping_api')
        api_logger.setLevel(logging.INFO)
        api_logger.addHandler(file_handler)
        
        stage1_logger = logging.getLogger('MapOMOP.mapping_stages.stage1_candidate_retrieval')
        stage1_logger.setLevel(logging.INFO)
        stage1_logger.addHandler(file_handler)
        
        stage2_logger = logging.getLogger('MapOMOP.mapping_stages.stage2_standard_collection')
        stage2_logger.setLevel(logging.INFO)
        stage2_logger.addHandler(file_handler)
        
        stage3_logger = logging.getLogger('MapOMOP.mapping_stages.stage3_hybrid_scoring')
        stage3_logger.setLevel(logging.INFO)
        stage3_logger.addHandler(file_handler)
        
        self.logger.info(f"로그 파일: {log_file}")
    
    def load_and_sample_data(self, csv_path: str, sample_size: int = 10000, use_random: bool = False, random_state: int = 42, filter_domains: list = None, sample_per_domain: dict = None) -> pd.DataFrame:
        """CSV 파일에서 데이터 로딩 및 샘플링
        
        Args:
            csv_path: CSV 파일 경로
            sample_size: 샘플 크기 (sample_per_domain이 None일 때만 사용)
            use_random: True면 랜덤 샘플링, False면 순서대로 (기본값: False)
            random_state: 랜덤 시드 (use_random=True일 때만 사용)
            filter_domains: 필터링할 도메인 리스트 (예: ['Condition', 'Measurement']) - sample_per_domain이 None일 때만 사용
            sample_per_domain: 도메인별 샘플 크기 딕셔너리 (예: {'Condition': 500, 'Procedure': 500})
        """
        self.logger.info(f"데이터 로딩 시작: {csv_path}")
        
        # 청크 단위로 데이터 로드
        chunk_size = 100000
        chunks = []
        
        self.logger.info("청크 단위로 데이터 읽는 중...")
        for chunk in tqdm(pd.read_csv(csv_path, chunksize=chunk_size), desc="데이터 로딩"):
            chunks.append(chunk)
        
        # 전체 데이터 병합
        df = pd.concat(chunks, ignore_index=True)
        self.logger.info(f"전체 데이터 크기: {len(df):,}개")
        
        # 도메인별 샘플링 모드
        if sample_per_domain:
            self.logger.info(f"도메인별 샘플링 모드")
            for domain, size in sample_per_domain.items():
                self.logger.info(f"  {domain}: {size}개")
            
            # 도메인별 샘플링
            sampled_dfs = []
            for domain, size in sample_per_domain.items():
                domain_df = df[df['domain_id'] == domain]
                domain_count = len(domain_df)
                
                if domain_count == 0:
                    self.logger.warning(f"  {domain}: 데이터 없음")
                    continue
                
                # 샘플 크기 조정 (있는 데이터보다 많이 요청하면 전체 사용)
                actual_sample_size = min(size, domain_count)
                
                if use_random:
                    domain_sample = domain_df.sample(n=actual_sample_size, random_state=random_state)
                else:
                    domain_sample = domain_df.head(actual_sample_size)
                
                sampled_dfs.append(domain_sample)
                self.logger.info(f"  {domain}: {actual_sample_size:,}개 샘플링 (전체: {domain_count:,}개)")
            
            # 모든 도메인 샘플 병합
            df_sample = pd.concat(sampled_dfs, ignore_index=True)
            
            # 랜덤 샘플링인 경우 전체를 다시 섞음
            if use_random:
                df_sample = df_sample.sample(frac=1, random_state=random_state).reset_index(drop=True)
            
            self.logger.info(f"총 샘플링 완료: {len(df_sample):,}개")
        
        # 기존 필터링 및 샘플링 모드
        else:
            self.logger.info(f"샘플 크기: {sample_size}개")
            self.logger.info(f"샘플링 방식: {'랜덤' if use_random else '순서대로'}")
            if filter_domains:
                self.logger.info(f"도메인 필터링: {filter_domains}")
                # 도메인 필터링
                if 'domain_id' in df.columns:
                    df = df[df['domain_id'].isin(filter_domains)]
                    self.logger.info(f"필터링 후 데이터 크기: {len(df):,}개")
            
            # 샘플링
            if not use_random:
                # 순서대로 샘플링
                df_sample = df.head(min(sample_size, len(df)))
                self.logger.info(f"순서대로 샘플링 완료: {len(df_sample):,}개")
            else:
                # 랜덤 샘플링
                df_sample = df.sample(n=min(sample_size, len(df)), random_state=random_state)
                df_sample = df_sample.reset_index(drop=True)
                self.logger.info(f"랜덤 샘플링 완료: {len(df_sample):,}개")
        
        self.logger.info(f"컬럼: {list(df_sample.columns)}")
        
        # 도메인 분포 출력
        if 'domain_id' in df_sample.columns:
            domain_dist = df_sample['domain_id'].value_counts()
            self.logger.info("\n도메인 분포:")
            for domain, count in domain_dist.items():
                self.logger.info(f"  {domain}: {count}개 ({count/len(df_sample)*100:.1f}%)")
        
        return df_sample
    
    def create_entity_input(self, row) -> EntityInput:
        """DataFrame 행에서 EntityInput 생성"""
        entity_name = str(row['entity_name']).strip()
        
        # 도메인 정보가 있으면 사용, 없으면 None (모든 도메인 검색)
        domain_id = None
        if 'domain_id' in row and pd.notna(row['domain_id']):
            domain_str = str(row['domain_id']).strip()
            if domain_str and domain_str in self.domain_mapping:
                domain_id = self.domain_mapping[domain_str]
        
        return EntityInput(
            entity_name=entity_name,
            domain_id=domain_id,
            vocabulary_id=None
        )
    
    def test_single_entity(self, entity_input: EntityInput, test_index: int, ground_truth_concept_id: int) -> dict:
        """단일 엔티티 테스트"""
        try:
            # 매핑 수행
            results = self.api.map_entity(entity_input)
            
            # 단계별 후보군 수집
            stage1_candidates = []
            stage2_candidates = []
            stage3_candidates = []
            
            if hasattr(self.api, '_last_stage1_candidates') and self.api._last_stage1_candidates:
                stage1_candidates = self.api._last_stage1_candidates
            
            if hasattr(self.api, '_last_stage2_candidates') and self.api._last_stage2_candidates:
                stage2_candidates = self.api._last_stage2_candidates
            
            if hasattr(self.api, '_last_rerank_candidates') and self.api._last_rerank_candidates:
                stage3_candidates = self.api._last_rerank_candidates
            
            # 도메인별 결과 정리
            domain_results = []
            if results:
                for result in results:
                    domain_info = {
                        'domain_id': result.domain_id,
                        'mapped_concept_id': result.mapped_concept_id,
                        'mapped_concept_name': result.mapped_concept_name,
                        'mapping_score': result.mapping_score,
                        'mapping_confidence': result.mapping_confidence,
                        'mapping_method': result.mapping_method,
                        'vocabulary_id': result.vocabulary_id
                    }
                    domain_results.append(domain_info)
            
            # 최고 점수 결과 선택
            best_result = max(results, key=lambda x: x.mapping_score) if results else None
            
            # 도메인별 Stage 경로 정보
            domain_stage_paths = {}
            best_search_domain = None
            if hasattr(self.api, '_all_domain_stage_results') and self.api._all_domain_stage_results:
                domain_stage_paths = self.api._all_domain_stage_results
                
                if best_result:
                    for search_domain, stage_info in domain_stage_paths.items():
                        if stage_info.get('result_domain') == best_result.domain_id:
                            best_search_domain = search_domain
                            break
            
            # 매핑 성공 여부 판단 (concept_id 일치)
            mapping_correct = False
            if best_result and ground_truth_concept_id:
                # 타입을 int로 통일하여 비교
                try:
                    best_concept_id_int = int(best_result.mapped_concept_id)
                    ground_truth_int = int(ground_truth_concept_id)
                    mapping_correct = (best_concept_id_int == ground_truth_int)
                    
                    # 디버깅 로그 (처음 10개만)
                    if test_index <= 10:
                        self.logger.info(f"   🔍 정답 판정: Ground Truth={ground_truth_int}, Best={best_concept_id_int}, Correct={mapping_correct}")
                except (ValueError, TypeError) as e:
                    self.logger.warning(f"   ⚠️ Concept ID 비교 오류: {e}")
                    mapping_correct = False
            
            test_result = {
                'test_index': test_index,
                'entity_name': entity_input.entity_name,
                'ground_truth_concept_id': ground_truth_concept_id,
                'success': results is not None and len(results) > 0,
                'mapping_correct': mapping_correct,
                'domain_count': len(results) if results else 0,
                'domain_results': domain_results,
                'domain_stage_paths': domain_stage_paths,
                'best_search_domain': best_search_domain,
                'best_result_domain': best_result.domain_id if best_result else None,
                'best_concept_id': best_result.mapped_concept_id if best_result else None,
                'best_concept_name': best_result.mapped_concept_name if best_result else None,
                'best_score': best_result.mapping_score if best_result else 0.0,
                'best_confidence': best_result.mapping_confidence if best_result else None,
                'stage1_candidates': stage1_candidates,
                'stage2_candidates': stage2_candidates,
                'stage3_candidates': stage3_candidates
            }
            
            return test_result
            
        except Exception as e:
            self.logger.error(f"테스트 #{test_index} 오류: {str(e)}")
            return {
                'test_index': test_index,
                'entity_name': entity_input.entity_name,
                'ground_truth_concept_id': ground_truth_concept_id,
                'success': False,
                'mapping_correct': False,
                'domain_count': 0,
                'domain_results': [],
                'domain_stage_paths': {},
                'best_search_domain': None,
                'best_result_domain': None,
                'best_concept_id': None,
                'best_concept_name': None,
                'best_score': 0.0,
                'best_confidence': None,
                'error': str(e),
                'stage1_candidates': [],
                'stage2_candidates': [],
                'stage3_candidates': []
            }
    
    def run_test_with_real_data(self, csv_path: str, sample_size: int = 10000, use_random: bool = False, filter_domains: list = None, sample_per_domain: dict = None):
        """실제 데이터로 테스트 실행
        
        Args:
            csv_path: CSV 파일 경로
            sample_size: 샘플 크기 (sample_per_domain이 None일 때만 사용)
            use_random: True면 랜덤 샘플링, False면 순서대로 (기본값: False)
            filter_domains: 필터링할 도메인 리스트 (예: ['Condition', 'Measurement']) - sample_per_domain이 None일 때만 사용
            sample_per_domain: 도메인별 샘플 크기 딕셔너리 (예: {'Condition': 500, 'Procedure': 500})
        """
        self.logger.info("=" * 100)
        self.logger.info("🚀 실제 데이터 Entity Mapping 테스트 시작")
        self.logger.info("=" * 100)
        
        start_time = time.time()
        
        # 데이터 로딩 및 샘플링
        test_data = self.load_and_sample_data(csv_path, sample_size, use_random=use_random, filter_domains=filter_domains, sample_per_domain=sample_per_domain)
        
        # 테스트 결과 저장
        test_results = []
        successful_mappings = 0
        correct_mappings = 0
        
        # tqdm으로 진행 상황 표시
        for idx, row in tqdm(test_data.iterrows(), total=len(test_data), desc="엔티티 매핑 테스트"):
            try:
                entity_input = self.create_entity_input(row)
                ground_truth = int(row['concept_id']) if pd.notna(row['concept_id']) else None
                
                result = self.test_single_entity(entity_input, idx + 1, ground_truth)
                test_results.append(result)
                
                if result['success']:
                    successful_mappings += 1
                    if result['mapping_correct']:
                        correct_mappings += 1
                        # 정답인 경우 로그 (처음 10개만)
                        if idx < 10:
                            self.logger.info(f"✅ #{idx + 1} 정답! {entity_input.entity_name}: GT={ground_truth} → Best={result.get('best_concept_id')}")
                    else:
                        # 오답인 경우 로그 (처음 10개만)
                        if idx < 10:
                            self.logger.info(f"❌ #{idx + 1} 오답: {entity_input.entity_name}: GT={ground_truth} → Best={result.get('best_concept_id')}")
                else:
                    # 매핑 실패 (처음 10개만)
                    if idx < 10:
                        self.logger.info(f"⚠️ #{idx + 1} 매핑 실패: {entity_input.entity_name}")
                        
            except Exception as e:
                self.logger.error(f"테스트 #{idx + 1} 처리 오류: {str(e)}")
                continue
        
        # 테스트 완료 시간
        end_time = time.time()
        elapsed_time = end_time - start_time
        
        # 결과 요약
        total_tests = len(test_results)
        success_rate = (successful_mappings / total_tests * 100) if total_tests > 0 else 0
        accuracy = (correct_mappings / total_tests * 100) if total_tests > 0 else 0
        
        self.logger.info("\n" + "=" * 100)
        self.logger.info("📊 테스트 결과 요약")
        self.logger.info("=" * 100)
        self.logger.info(f"총 테스트: {total_tests:,}개")
        self.logger.info(f"매핑 성공: {successful_mappings:,}개 ({success_rate:.2f}%)")
        self.logger.info(f"정답 매칭: {correct_mappings:,}개 ({accuracy:.2f}%)")
        self.logger.info(f"오답 매칭: {successful_mappings - correct_mappings:,}개")
        self.logger.info(f"매핑 실패: {total_tests - successful_mappings:,}개")
        self.logger.info(f"소요 시간: {elapsed_time:.2f}초 ({elapsed_time/60:.2f}분)")
        self.logger.info(f"평균 처리 시간: {elapsed_time/total_tests:.3f}초/엔티티")
        
        # 정답/오답 예시 출력
        correct_examples = [r for r in test_results if r['mapping_correct']]
        incorrect_examples = [r for r in test_results if r['success'] and not r['mapping_correct']]
        
        if correct_examples:
            self.logger.info("\n✅ 정답 예시 (최대 5개):")
            for i, result in enumerate(correct_examples[:5], 1):
                self.logger.info(f"  {i}. {result['entity_name']}")
                self.logger.info(f"     GT: {result['ground_truth_concept_id']} → Best: {result['best_concept_id']} ({result['best_concept_name']})")
        
        if incorrect_examples:
            self.logger.info("\n❌ 오답 예시 (최대 5개):")
            for i, result in enumerate(incorrect_examples[:5], 1):
                self.logger.info(f"  {i}. {result['entity_name']}")
                self.logger.info(f"     GT: {result['ground_truth_concept_id']} → Best: {result['best_concept_id']} ({result['best_concept_name']})")
        
        self.logger.info("=" * 100)
        
        # 결과를 XLSX로 저장
        self.save_results_to_xlsx(test_results)
        
        return test_results
    
    def save_results_to_xlsx(self, test_results: list):
        """테스트 결과를 XLSX 파일로 저장 (stage 후보군 포함)"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_file = self.log_dir / f"real_data_results_detailed_{timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detailed Results"
        
        self._create_detailed_sheet(ws, test_results)
        
        wb.save(xlsx_file)
        self.logger.info(f"📊 테스트 결과 XLSX 저장: {xlsx_file}")
    
    def _create_detailed_sheet(self, ws, test_results):
        """상세 시트 생성"""
        
        # 헤더 설정
        headers = [
            "Test Index", "Entity Name", "Ground Truth Concept ID", 
            "Success", "Mapping Correct", "Domain Count",
            "Best Search Domain", "Best Result Domain", 
            "Best Concept ID", "Best Concept Name", 
            "Best Score", "Best Confidence",
            "All Domains", "Domain Stage Paths", 
            "Stage1 Candidates", "Stage2 Candidates", "Stage3 Candidates"
        ]
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row, result in enumerate(test_results, 2):
            ws.cell(row=row, column=1, value=result['test_index'])
            ws.cell(row=row, column=2, value=result['entity_name'])
            ws.cell(row=row, column=3, value=result['ground_truth_concept_id'])
            ws.cell(row=row, column=4, value="성공" if result['success'] else "실패")
            
            # 매핑 정확도 표시 (색상 적용)
            correct_cell = ws.cell(row=row, column=5, value="정답" if result['mapping_correct'] else "오답")
            if result['mapping_correct']:
                correct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                correct_cell.font = Font(color="006100")
            else:
                correct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                correct_cell.font = Font(color="9C0006")
            
            ws.cell(row=row, column=6, value=result.get('domain_count', 0))
            ws.cell(row=row, column=7, value=result.get('best_search_domain', 'N/A'))
            ws.cell(row=row, column=8, value=result.get('best_result_domain', 'N/A'))
            ws.cell(row=row, column=9, value=result.get('best_concept_id', 'N/A'))
            ws.cell(row=row, column=10, value=result.get('best_concept_name', 'N/A'))
            ws.cell(row=row, column=11, value=result.get('best_score', 0.0))
            ws.cell(row=row, column=12, value=result.get('best_confidence', 'N/A'))
            
            # 도메인 결과
            domain_results_text = self._format_domain_results(result.get('domain_results', []))
            ws.cell(row=row, column=13, value=domain_results_text)
            
            # Stage 경로
            stage_paths_text = self._format_stage_paths(result.get('domain_stage_paths', {}))
            ws.cell(row=row, column=14, value=stage_paths_text)
            
            # Stage 후보군
            stage1_text = self._format_candidates_for_cell(result.get('stage1_candidates', []), 'stage1')
            ws.cell(row=row, column=15, value=stage1_text)
            
            stage2_text = self._format_candidates_for_cell(result.get('stage2_candidates', []), 'stage2')
            ws.cell(row=row, column=16, value=stage2_text)
            
            stage3_text = self._format_candidates_for_cell(result.get('stage3_candidates', []), 'stage3')
            ws.cell(row=row, column=17, value=stage3_text)
            
            # 셀 스타일 설정
            for col in range(13, 18):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 열 너비 설정
        column_widths = {
            'A': 10,  # Test Index
            'B': 40,  # Entity Name
            'C': 20,  # Ground Truth Concept ID
            'D': 10,  # Success
            'E': 12,  # Mapping Correct
            'F': 12,  # Domain Count
            'G': 18,  # Best Search Domain
            'H': 18,  # Best Result Domain
            'I': 15,  # Best Concept ID
            'J': 45,  # Best Concept Name
            'K': 12,  # Best Score
            'L': 15,  # Best Confidence
            'M': 50,  # All Domains
            'N': 45,  # Domain Stage Paths
            'O': 70,  # Stage1 Candidates
            'P': 70,  # Stage2 Candidates
            'Q': 85   # Stage3 Candidates
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 행 높이 설정
        for row_num in range(2, len(test_results) + 2):
            ws.row_dimensions[row_num].height = 150
    
    def _format_domain_results(self, domain_results):
        """도메인별 결과 포맷팅"""
        if not domain_results:
            return "도메인 결과 없음"
        
        lines = []
        for i, domain in enumerate(domain_results, 1):
            line = f"{i}. [{domain.get('domain_id', 'N/A')}] {domain.get('mapped_concept_name', 'N/A')}\n"
            line += f"   ID: {domain.get('mapped_concept_id', 'N/A')}, "
            line += f"Score: {domain.get('mapping_score', 0):.4f}, "
            line += f"Conf: {domain.get('mapping_confidence', 'N/A')}\n"
            line += f"   Vocab: {domain.get('vocabulary_id', 'N/A')}"
            lines.append(line)
        
        return "\n\n".join(lines)
    
    def _format_stage_paths(self, stage_paths):
        """Stage 경로 포맷팅"""
        if not stage_paths:
            return "경로 정보 없음"
        
        lines = []
        for domain_name, stage_info in sorted(stage_paths.items()):
            search_domain = stage_info.get('search_domain', domain_name)
            result_domain = stage_info.get('result_domain', 'N/A')
            
            if search_domain != result_domain:
                line = f"[{search_domain} → {result_domain}]\n"
            else:
                line = f"[{search_domain}]\n"
            
            line += f"  Stage1: {stage_info.get('stage1_count', 0)}개\n"
            line += f"  Stage2: {stage_info.get('stage2_count', 0)}개\n"
            line += f"  Stage3: {stage_info.get('stage3_count', 0)}개"
            lines.append(line)
        
        return "\n\n".join(lines)
    
    def _format_candidates_for_cell(self, candidates, stage_type):
        """후보군 포맷팅"""
        if not candidates:
            return "후보 없음"
        
        lines = []
        max_candidates = 15 if stage_type in ['stage1', 'stage2'] else 10
        
        for i, candidate in enumerate(candidates[:max_candidates], 1):
            if stage_type == 'stage1':
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   ES점수: {candidate.get('elasticsearch_score', 0):.4f}, "
                line += f"Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            elif stage_type == 'stage2':
                search_type = candidate.get('search_type', 'unknown')
                is_std = "✓" if candidate.get('is_original_standard', True) else "→"
                line = f"{i}. [{search_type}] {is_std} {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
                if not candidate.get('is_original_standard', True):
                    original_non_std = candidate.get('original_non_standard', {})
                    if original_non_std:
                        line += f"\n   원본 Non-std: {original_non_std.get('concept_name', 'N/A')} (ID: {original_non_std.get('concept_id', 'N/A')})"
            else:  # stage3
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   텍스트: {candidate.get('text_similarity', 0):.4f}, "
                line += f"의미적: {candidate.get('semantic_similarity', 0):.4f}, "
                line += f"최종: {candidate.get('final_score', 0):.4f}\n"
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            
            lines.append(line)
        
        return "\n\n".join(lines)

def main():
    """메인 함수"""
    # scoring_mode 설정: 'llm' 또는 'hybrid'
    SCORING_MODE = "llm"  # 'llm' 또는 'hybrid' 선택
    
    tester = RealDataEntityMappingTester(scoring_mode=SCORING_MODE)
    
    # 실제 데이터 경로
    csv_path = "/home/work/skku/hyo/MapOMOP/data/mapping_test_snomed_no_note.csv"
    
    # 도메인별 샘플링 설정 (각 도메인당 500개씩 랜덤 샘플)
    SAMPLE_PER_DOMAIN = {
        'Condition': 500,
        'Procedure': 500,
        'Measurement': 500,
        'Observation': 500
    }
    USE_RANDOM = True  # 랜덤 샘플링 활성화
    
    results = tester.run_test_with_real_data(
        csv_path, 
        use_random=USE_RANDOM,
        sample_per_domain=SAMPLE_PER_DOMAIN
    )
    
    print(f"\n✅ 테스트 완료! 결과는 {tester.log_dir} 디렉토리에 저장되었습니다.")

if __name__ == "__main__":
    main()

