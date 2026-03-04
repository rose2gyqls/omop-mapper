#!/usr/bin/env python3
"""
MapOMOP 평가용 엑셀 파일 생성 스크립트

Blind test용 결과/평가 엑셀 2개 파일 생성:
  - 결과 파일: index, entity_name, domain_id, concept1, concept2, ... (중복 제거된 매핑 결과)
  - 평가 파일: index, entity_name, domain_id, score1, score2, ... (concept1→score1, concept2→score2)
  - unique concept 수보다 많은 score 열은 N/A 처리

입력: run_mapping.py --repeat N 실행 후 생성된 JSON 파일
  - 단일 run: [result, result, ...]
  - 다중 run: {"num_runs": N, "runs": [[run1], [run2], ...]}

Usage:
    python test/generate_evaluation_excel.py test_logs/mapping_snuh_20250101_120000.json
    python test/generate_evaluation_excel.py test_logs/mapping_snuh_20250101_120000.json -o evaluation_output
"""

import argparse
import json
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def load_mapping_results(json_path: str) -> list[list[dict]]:
    """
    JSON 파일에서 매핑 결과 로드.
    Returns: all_results = [run1_results, run2_results, ...]
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict) and "num_runs" in data and "runs" in data:
        return data["runs"]
    if isinstance(data, list):
        return [data]
    raise ValueError(f"지원하지 않는 JSON 형식: {type(data)}")


def get_unique_concepts_for_entity(run_results: list[dict], test_index: int) -> list[tuple[str, str]]:
    """
    특정 test_index에 대한 N회 run 결과에서 concept_id 기준 중복 제거.
    Returns: [(concept_name, concept_id), ...] 순서 유지 (첫 등장 순)
    """
    seen_ids = set()
    unique = []
    for r in run_results:
        if r.get("test_index") != test_index:
            continue
        cid = r.get("best_concept_id")
        cname = r.get("best_concept_name") or "N/A"
        if cid is None or str(cid).strip() == "":
            continue
        cid_str = str(int(cid)) if isinstance(cid, (int, float)) else str(cid)
        if cid_str not in seen_ids:
            seen_ids.add(cid_str)
            unique.append((cname, cid_str))
    return unique


def build_results_data(all_results: list[list[dict]]) -> tuple[list[dict], int]:
    """
    all_results를 기반으로 결과 시트/평가 시트용 데이터 구성.
    Returns: (rows, max_concepts)
    """
    # test_index별로 모든 run 결과 수집
    by_index: dict[int, list[dict]] = {}
    for run_results in all_results:
        for r in run_results:
            idx = r.get("test_index")
            if idx is None:
                continue
            if idx not in by_index:
                by_index[idx] = []
            by_index[idx].append(r)

    rows = []
    max_concepts = 0

    for test_index in sorted(by_index.keys()):
        run_results = by_index[test_index]
        r0 = run_results[0]

        unique_concepts = get_unique_concepts_for_entity(run_results, test_index)
        max_concepts = max(max_concepts, len(unique_concepts))

        row = {
            "index": test_index,
            "entity_name": r0.get("entity_name", ""),
            "domain_id": r0.get("input_domain", "All"),
            "unique_concepts": unique_concepts,
        }
        rows.append(row)

    return rows, max_concepts


def create_evaluation_excel(
    json_path: str,
    output_path: str | None = None,
) -> tuple[Path, Path]:
    """
    평가용 엑셀 파일 2개 생성 (결과/평가 각각 별도 파일).
    - 결과: index, entity_name, domain_id, concept1, concept2, ...
    - 평가: index, entity_name, domain_id, score1, score2, ... (평가 불필요 칸은 N/A)
    Returns: (results_path, evaluation_path)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl 필요. pip install openpyxl")

    all_results = load_mapping_results(json_path)
    rows, max_concepts = build_results_data(all_results)

    if output_path is None:
        base = Path(json_path).parent / "evaluation_blind_test"
    else:
        base = Path(output_path)
        if base.suffix == ".xlsx":
            base = base.with_suffix("")

    results_path = base.parent / f"{base.name}_결과.xlsx"
    eval_path = base.parent / f"{base.name}_평가.xlsx"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # ----- 결과 파일 -----
    wb_results = openpyxl.Workbook()
    ws_results = wb_results.active
    ws_results.title = "결과"

    results_headers = ["index", "entity_name", "domain_id"] + [
        f"concept{i + 1}" for i in range(max_concepts)
    ]
    for col, h in enumerate(results_headers, 1):
        cell = ws_results.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, 2):
        ws_results.cell(row=row_idx, column=1, value=row["index"])
        ws_results.cell(row=row_idx, column=2, value=row["entity_name"])
        ws_results.cell(row=row_idx, column=3, value=row["domain_id"])
        for i, (cname, cid) in enumerate(row["unique_concepts"]):
            val = f"{cname}({cid})"
            ws_results.cell(row=row_idx, column=4 + i, value=val)

    ws_results.column_dimensions["A"].width = 8
    ws_results.column_dimensions["B"].width = 45
    ws_results.column_dimensions["C"].width = 15
    for i in range(max_concepts):
        ws_results.column_dimensions[openpyxl.utils.get_column_letter(4 + i)].width = 50

    wb_results.save(results_path)

    # ----- 평가 파일 -----
    wb_eval = openpyxl.Workbook()
    ws_eval = wb_eval.active
    ws_eval.title = "평가"

    eval_headers = ["index", "entity_name", "domain_id"] + [
        f"score{i + 1}" for i in range(max_concepts)
    ]
    for col, h in enumerate(eval_headers, 1):
        cell = ws_eval.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, 2):
        ws_eval.cell(row=row_idx, column=1, value=row["index"])
        ws_eval.cell(row=row_idx, column=2, value=row["entity_name"])
        ws_eval.cell(row=row_idx, column=3, value=row["domain_id"])
        num_concepts = len(row["unique_concepts"])
        for i in range(max_concepts):
            # 평가 필요한 칸: 빈칸, 평가 불필요 칸: N/A
            val = "" if i < num_concepts else "N/A"
            ws_eval.cell(row=row_idx, column=4 + i, value=val)

    ws_eval.column_dimensions["A"].width = 8
    ws_eval.column_dimensions["B"].width = 45
    ws_eval.column_dimensions["C"].width = 15
    for i in range(max_concepts):
        ws_eval.column_dimensions[openpyxl.utils.get_column_letter(4 + i)].width = 12

    wb_eval.save(eval_path)

    return results_path, eval_path


def main():
    parser = argparse.ArgumentParser(
        description="MapOMOP 평가용 엑셀 생성 (Blind test: 결과 시트 + 평가 시트)"
    )
    parser.add_argument(
        "json_path",
        help="매핑 결과 JSON 경로 (run_mapping --repeat N 출력)",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="출력 파일 경로 prefix (기본: JSON과 동일 폴더의 evaluation_blind_test). 결과: {prefix}_결과.xlsx, 평가: {prefix}_평가.xlsx",
    )
    args = parser.parse_args()

    results_path, eval_path = create_evaluation_excel(
        args.json_path,
        output_path=args.output,
    )
    print(f"결과: {results_path}")
    print(f"평가: {eval_path}")


if __name__ == "__main__":
    main()
