import pandas as pd
import logging
import os
from datetime import datetime
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm
import time

# 상대 경로로 src 디렉토리 추가
sys.path.append(str(Path(__file__).parent.parent / 'src'))

from MapOMOP.entity_mapping_api import EntityMappingAPI, EntityInput, DomainID
from MapOMOP.elasticsearch_client import ElasticsearchClient

class EntityMappingTester:
    def __init__(self, log_dir: str = "test_logs", scoring_mode: str = "llm"):
        """테스터 초기화
        
        Args:
            log_dir: 로그 디렉토리
            scoring_mode: Stage 3 점수 계산 방식
                - 'llm': LLM without score (디폴트)
                - 'llm_with_score': LLM with semantic score in prompt
                - 'semantic': Semantic similarity only
        """
        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(exist_ok=True)
        self.scoring_mode = scoring_mode
        
        self.setup_logging()
        
        self.es_client = ElasticsearchClient()
        self.es_client.concept_index = "concept-small"
        self.es_client.concept_synonym_index = "concept-synonym"
        
        self.api = EntityMappingAPI(
            es_client=self.es_client,
            scoring_mode=scoring_mode
        )
        
        self.logger.info(f"✅ Scoring Mode: {scoring_mode}")
        
        self.domain_mapping = {
            'Condition': DomainID.CONDITION,
            'Procedure': DomainID.PROCEDURE,
            'Drug': DomainID.DRUG,
            'Observation': DomainID.OBSERVATION,
            'Measurement': DomainID.MEASUREMENT,
            'Period': DomainID.PERIOD,
            'Provider': DomainID.PROVIDER,
            'Device': DomainID.DEVICE,
        }
    
    def setup_logging(self):
        """로깅 설정"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        log_file = self.log_dir / f"entity_mapping_test_{timestamp}.log"
        
        self.logger = logging.getLogger('entity_mapping_test')
        self.logger.setLevel(logging.INFO)
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        api_logger = logging.getLogger('MapOMOP.entity_mapping_api')
        api_logger.setLevel(logging.INFO)
        api_logger.addHandler(file_handler)
        
        stage1_logger = logging.getLogger('MapOMOP.mapping_stages.stage1_candidate_retrieval')
        stage1_logger.setLevel(logging.INFO)
        stage1_logger.addHandler(file_handler)
        
        stage2_logger = logging.getLogger('MapOMOP.mapping_stages.stage2_standard_collection')
        stage2_logger.setLevel(logging.INFO)
        stage2_logger.addHandler(file_handler)
        
        stage3_logger = logging.getLogger('MapOMOP.mapping_stages.stage3_hybrid_scoring')
        stage3_logger.setLevel(logging.INFO)
        stage3_logger.addHandler(file_handler)
        
        self.logger.info(f"로그 파일: {log_file}")
    
    def load_test_data_from_list(self, entity_list: list) -> pd.DataFrame:
        """리스트에서 테스트 데이터 생성
        
        Args:
            entity_list: 엔티티 리스트. 다음 형식 지원:
                - 문자열 리스트: ['entity1', 'entity2', ...]
                - (entity, domain) 튜플 리스트: [('entity1', 'Condition'), ('entity2', 'Drug'), ...]
                - domain이 None이면 모든 도메인 검색
        """
        self.logger.info(f"테스트 데이터 생성: {len(entity_list)}개 엔티티")
        
        test_data = []
        for i, item in enumerate(entity_list):
            # (entity, domain) 튜플인 경우
            if isinstance(item, tuple) and len(item) == 2:
                entity_name, domain = item
                test_data.append({
                    'entity_plain_name': entity_name,
                    'entity_domain': domain,
                    'sheet': 'manual'
                })
                domain_str = domain if domain else 'All'
                self.logger.info(f"  {i+1}. {entity_name} [{domain_str}]")
            # 문자열인 경우
            else:
                test_data.append({
                    'entity_plain_name': item,
                    'entity_domain': None,
                    'sheet': 'manual'
                })
                self.logger.info(f"  {i+1}. {item} [All]")
        
        df = pd.DataFrame(test_data)
        self.logger.info(f"전체 테스트 데이터: {len(df)}개 엔티티")
        
        # 도메인 분포 출력
        if 'entity_domain' in df.columns:
            domain_dist = df['entity_domain'].fillna('All').value_counts()
            self.logger.info("\n도메인 분포:")
            for domain, count in domain_dist.items():
                self.logger.info(f"  {domain}: {count}개")
        
        return df
    
    def create_entity_input(self, row) -> EntityInput:
        """DataFrame 행에서 EntityInput 생성"""
        entity_name = str(row['entity_plain_name']).strip()
        
        # 도메인은 기본적으로 None (모든 도메인 검색)
        # 사용자가 entity_domain을 지정한 경우에만 사용
        domain_id = None
        if 'entity_domain' in row and pd.notna(row['entity_domain']):
            domain_str = str(row['entity_domain']).strip()
            if domain_str and domain_str in self.domain_mapping:
                domain_id = self.domain_mapping[domain_str]
        
        return EntityInput(
            entity_name=entity_name,
            domain_id=domain_id,
            vocabulary_id=None
        )
    
    def test_single_entity(self, entity_input: EntityInput, test_index: int, sheet: str) -> dict:
        """단일 엔티티 테스트 (도메인별 결과)"""
        input_domain = entity_input.domain_id.value if entity_input.domain_id else 'All'
        
        self.logger.info("=" * 100)
        self.logger.info(f"🧪 테스트 #{test_index} [{input_domain}]: {entity_input.entity_name}")
        self.logger.info("=" * 100)
        
        try:
            # 매핑 수행 (도메인별 결과 반환)
            results = self.api.map_entity(entity_input)
            
            # 단계별 상세 정보 로깅 (마지막 도메인)
            stage1_candidates = []
            stage2_candidates = []
            stage3_candidates = []
            
            if hasattr(self.api, '_last_stage1_candidates') and self.api._last_stage1_candidates:
                stage1_candidates = self.api._last_stage1_candidates
            
            if hasattr(self.api, '_last_stage2_candidates') and self.api._last_stage2_candidates:
                stage2_candidates = self.api._last_stage2_candidates
                self.logger.info("📊 Stage 2 후보군 상세 정보:")
                for i, candidate in enumerate(stage2_candidates, 1):
                    search_type = candidate.get('search_type', 'unknown')
                    is_std = "✓" if candidate['is_original_standard'] else "→"
                    self.logger.info(f"   {i}. [{search_type}] {is_std} {candidate['concept_name']} (ID: {candidate['concept_id']})")
                    self.logger.info(f"      - Domain: {candidate.get('domain_id', 'N/A')}, Vocabulary: {candidate['vocabulary_id']}")
            
            if hasattr(self.api, '_last_rerank_candidates') and self.api._last_rerank_candidates:
                stage3_candidates = self.api._last_rerank_candidates
                # LLM 모드인 경우 Stage 3 결과 상세 로깅
                if self.scoring_mode in ['llm', 'llm_with_score'] and stage3_candidates:
                    header = "📊 Stage 3 LLM 평가 결과"
                    if self.scoring_mode == 'llm_with_score':
                        header += " (SapBERT 의미유사도 포함)"
                    self.logger.info(f"\n{header}:")
                    for i, candidate in enumerate(stage3_candidates[:10], 1):
                        llm_score = candidate.get('llm_score', candidate.get('final_score', 0))
                        llm_rank = candidate.get('llm_rank', i)
                        llm_reasoning = candidate.get('llm_reasoning', 'N/A')
                        semantic_sim = candidate.get('semantic_similarity')
                        
                        self.logger.info(f"   {i}. {candidate['concept_name']} (ID: {candidate['concept_id']})")
                        if semantic_sim is not None:
                            self.logger.info(f"      - LLM 점수: {llm_score}, 순위: {llm_rank}, 의미유사도: {semantic_sim:.4f}")
                        else:
                            self.logger.info(f"      - LLM 점수: {llm_score}, 순위: {llm_rank}")
                        if llm_reasoning and llm_reasoning != 'N/A':
                            reasoning_short = llm_reasoning[:80] + '...' if len(llm_reasoning) > 80 else llm_reasoning
                            self.logger.info(f"      - 이유: {reasoning_short}")
            
            # 도메인별 결과 정리
            domain_results = []
            if results:
                self.logger.info("\n" + "=" * 100)
                self.logger.info("📊 도메인별 매핑 결과 요약")
                self.logger.info("=" * 100)
                
                for idx, result in enumerate(results, 1):
                    domain_info = {
                        'domain_id': result.domain_id,
                        'mapped_concept_id': result.mapped_concept_id,
                        'mapped_concept_name': result.mapped_concept_name,
                        'mapping_score': result.mapping_score,
                        'mapping_confidence': result.mapping_confidence,
                        'mapping_method': result.mapping_method,
                        'vocabulary_id': result.vocabulary_id
                    }
                    domain_results.append(domain_info)
                    
                    self.logger.info(f"\n{idx}. [{result.domain_id}] 매핑 성공!")
                    self.logger.info(f"   개념: {result.mapped_concept_name} (ID: {result.mapped_concept_id})")
                    self.logger.info(f"   점수: {result.mapping_score:.4f} | 신뢰도: {result.mapping_confidence}")
                    self.logger.info(f"   방법: {result.mapping_method} | Vocabulary: {result.vocabulary_id}")
            
            # 결과 정리 (최고 점수 도메인 선택)
            best_result = max(results, key=lambda x: x.mapping_score) if results else None
            
            # 도메인별 Stage 경로 정보 추출
            domain_stage_paths = {}
            best_search_domain = None
            if hasattr(self.api, '_all_domain_stage_results') and self.api._all_domain_stage_results:
                domain_stage_paths = self.api._all_domain_stage_results
                
                # Best result의 검색 도메인 찾기
                if best_result:
                    for search_domain, stage_info in domain_stage_paths.items():
                        if stage_info.get('result_domain') == best_result.domain_id:
                            best_search_domain = search_domain
                            break
            
            test_result = {
                'test_index': test_index,
                'sheet': sheet,
                'entity_name': entity_input.entity_name,
                'input_domain': input_domain,
                'success': results is not None and len(results) > 0,
                'domain_count': len(results) if results else 0,
                'domain_results': domain_results,
                'domain_stage_paths': domain_stage_paths,
                'best_search_domain': best_search_domain,
                'best_result_domain': best_result.domain_id if best_result else None,
                'best_concept_id': best_result.mapped_concept_id if best_result else None,
                'best_concept_name': best_result.mapped_concept_name if best_result else None,
                'best_score': best_result.mapping_score if best_result else 0.0,
                'best_confidence': best_result.mapping_confidence if best_result else None,
                'stage1_candidates': stage1_candidates,
                'stage2_candidates': stage2_candidates,
                'stage3_candidates': stage3_candidates
            }
            
            if not results:
                self.logger.info(f"❌ 모든 도메인에서 매핑 실패")
            else:
                self.logger.info(f"\n" + "=" * 100)
                self.logger.info(f"📊 최종 요약: {len(results)}개 도메인에서 매핑 성공")
                self.logger.info("=" * 100)
                self.logger.info(f"🏆 최고 점수: [{best_result.domain_id}] {best_result.mapped_concept_name} ({best_result.mapping_score:.4f})")
                
                # 도메인별 Stage 경로 출력
                if hasattr(self.api, '_all_domain_stage_results') and self.api._all_domain_stage_results:
                    self.logger.info(f"\n📈 도메인별 Stage 경로:")
                    for domain_name, stage_info in self.api._all_domain_stage_results.items():
                        self.logger.info(f"  [{domain_name}] Stage1: {stage_info.get('stage1_count', 0)}개 → "
                                       f"Stage2: {stage_info.get('stage2_count', 0)}개 → "
                                       f"Stage3: {stage_info.get('stage3_count', 0)}개")
                self.logger.info("=" * 100)
                
            return test_result
            
        except Exception as e:
            self.logger.error(f"❌ 테스트 오류: {str(e)}")
            return {
                'test_index': test_index,
                'sheet': sheet,
                'entity_name': entity_input.entity_name,
                'input_domain': input_domain,
                'success': False,
                'domain_count': 0,
                'domain_results': [],
                'best_search_domain': None,
                'best_result_domain': None,
                'best_concept_id': None,
                'best_concept_name': None,
                'best_score': 0.0,
                'best_confidence': None,
                'error': str(e),
                'stage1_candidates': [],
                'stage2_candidates': [],
                'stage3_candidates': []
            }
    
    def run_test_with_entities(self, entity_list: list, max_entities: int = None):
        """엔티티 리스트로 테스트 실행
        
        Args:
            entity_list: 엔티티 리스트 (문자열 또는 (entity, domain) 튜플)
            max_entities: 테스트할 최대 엔티티 수
        """
        self.logger.info("=" * 100)
        self.logger.info("🚀 Entity Mapping API 테스트 시작")
        self.logger.info("=" * 100)
        self.logger.info(f"테스트 엔티티 리스트: {len(entity_list)}개")
        self.logger.info(f"Scoring Mode: {self.scoring_mode}")
        
        start_time = time.time()
        
        # 데이터 생성
        test_data = self.load_test_data_from_list(entity_list)
        
        if max_entities:
            test_data = test_data.head(max_entities)
            self.logger.info(f"테스트 제한: 최대 {max_entities}개 엔티티")
        
        # 테스트 결과 저장
        test_results = []
        successful_tests = 0
        
        # tqdm으로 진행 상황 표시
        for idx, row in tqdm(test_data.iterrows(), total=len(test_data), desc="엔티티 매핑 테스트"):
            try:
                entity_input = self.create_entity_input(row)
                result = self.test_single_entity(entity_input, idx + 1, row['sheet'])
                test_results.append(result)
                
                if result['success']:
                    successful_tests += 1
                    
            except Exception as e:
                self.logger.error(f"테스트 #{idx + 1} 처리 오류: {str(e)}")
                continue
        
        # 테스트 완료 시간
        end_time = time.time()
        elapsed_time = end_time - start_time
        
        # 결과 요약
        total_tests = len(test_results)
        success_rate = (successful_tests / total_tests * 100) if total_tests > 0 else 0
        
        self.logger.info("\n" + "=" * 100)
        self.logger.info("📊 테스트 결과 요약")
        self.logger.info("=" * 100)
        self.logger.info(f"총 테스트: {total_tests}개")
        self.logger.info(f"매핑 성공: {successful_tests}개 ({success_rate:.2f}%)")
        self.logger.info(f"매핑 실패: {total_tests - successful_tests}개")
        self.logger.info(f"소요 시간: {elapsed_time:.2f}초 ({elapsed_time/60:.2f}분)")
        if total_tests > 0:
            self.logger.info(f"평균 처리 시간: {elapsed_time/total_tests:.3f}초/엔티티")
        
        # 엔티티별 요약
        self.logger.info("\n📋 엔티티별 결과:")
        for i, result in enumerate(test_results, 1):
            status = "✅ 성공" if result['success'] else "❌ 실패"
            input_domain = result.get('input_domain', 'All')
            self.logger.info(f"  {i}. [{input_domain}] {result['entity_name']}: {status}")
            if result['success']:
                search_domain = result.get('best_search_domain', 'N/A')
                result_domain = result.get('best_result_domain', 'N/A')
                if search_domain == result_domain:
                    domain_info = f"[{result_domain}]"
                else:
                    domain_info = f"[{search_domain} → {result_domain}]"
                self.logger.info(f"     -> {domain_info} {result.get('best_concept_name', 'N/A')} (점수: {result.get('best_score', 0.0):.4f})")
        
        # 결과를 CSV와 XLSX로 저장
        self.save_results_to_csv(test_results)
        self.save_results_to_xlsx(test_results)
        
        return test_results
    
    def save_results_to_csv(self, test_results: list):
        """테스트 결과를 CSV 파일로 저장 (도메인별 결과 포함)"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_file = self.log_dir / f"test_results_{self.scoring_mode}_{timestamp}.csv"
        
        # CSV용 데이터 정리 (도메인별 결과 평탄화)
        csv_results = []
        for result in test_results:
            # 기본 정보
            base_info = {
                'test_index': result['test_index'],
                'entity_name': result['entity_name'],
                'input_domain': result.get('input_domain', 'All'),
                'success': result['success'],
                'domain_count': result.get('domain_count', 0),
                'best_search_domain': result.get('best_search_domain', 'N/A'),
                'best_result_domain': result.get('best_result_domain', 'N/A'),
                'best_concept_id': result.get('best_concept_id', 'N/A'),
                'best_concept_name': result.get('best_concept_name', 'N/A'),
                'best_score': result.get('best_score', 0.0),
                'best_confidence': result.get('best_confidence', 'N/A')
            }
            csv_results.append(base_info)
        
        df_results = pd.DataFrame(csv_results)
        df_results.to_csv(csv_file, index=False, encoding='utf-8')
        
        self.logger.info(f"📄 테스트 결과 CSV 저장: {csv_file}")
    
    def save_results_to_xlsx(self, test_results: list):
        """테스트 결과를 XLSX 파일로 저장 (stage1, stage3 후보군을 열로 분리)"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_file = self.log_dir / f"test_results_detailed_{self.scoring_mode}_{timestamp}.xlsx"
        
        # 엑셀 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detailed Results"
        
        # 통합 상세 시트 생성
        self._create_integrated_detail_sheet(ws, test_results)
        
        # 파일 저장
        wb.save(xlsx_file)
        self.logger.info(f"📊 테스트 결과 XLSX 저장: {xlsx_file}")
    
    def _create_integrated_detail_sheet(self, ws, test_results):
        """통합 상세 시트 생성 (모든 엔티티를 하나의 시트에, 도메인별 결과 포함)"""
        
        # 헤더 설정
        headers = [
            "Test Index", "Entity Name", "Input Domain", "Success", "Domain Count",
            "Best Search Domain", "Best Result Domain", "Best Concept ID", "Best Concept Name", 
            "Best Score", "Best Confidence",
            "All Domains", "Domain Stage Paths", "Stage1 Candidates", "Stage2 Candidates", "Stage3 Candidates"
        ]
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row, result in enumerate(test_results, 2):
            ws.cell(row=row, column=1, value=result['test_index'])
            ws.cell(row=row, column=2, value=result['entity_name'])
            ws.cell(row=row, column=3, value=result.get('input_domain', 'All'))
            ws.cell(row=row, column=4, value="성공" if result['success'] else "실패")
            ws.cell(row=row, column=5, value=result.get('domain_count', 0))
            ws.cell(row=row, column=6, value=result.get('best_search_domain', 'N/A'))
            ws.cell(row=row, column=7, value=result.get('best_result_domain', 'N/A'))
            ws.cell(row=row, column=8, value=result.get('best_concept_id', 'N/A'))
            ws.cell(row=row, column=9, value=result.get('best_concept_name', 'N/A'))
            ws.cell(row=row, column=10, value=result.get('best_score', 0.0))
            ws.cell(row=row, column=11, value=result.get('best_confidence', 'N/A'))
            
            # 모든 도메인 결과를 문자열로 변환
            domain_results_text = self._format_domain_results(result.get('domain_results', []))
            ws.cell(row=row, column=12, value=domain_results_text)
            
            # 도메인별 Stage 경로 정보
            stage_paths_text = self._format_stage_paths(result.get('domain_stage_paths', {}))
            ws.cell(row=row, column=13, value=stage_paths_text)
            
            # Stage1 후보군 정보를 문자열로 변환
            stage1_text = self._format_candidates_for_cell(result.get('stage1_candidates', []), 'stage1')
            ws.cell(row=row, column=14, value=stage1_text)
            
            # Stage2 후보군 정보를 문자열로 변환
            stage2_text = self._format_candidates_for_cell(result.get('stage2_candidates', []), 'stage2')
            ws.cell(row=row, column=15, value=stage2_text)
            
            # Stage3 후보군 정보를 문자열로 변환
            stage3_text = self._format_candidates_for_cell(result.get('stage3_candidates', []), 'stage3')
            ws.cell(row=row, column=16, value=stage3_text)
            
            # 셀 스타일 설정 (텍스트 줄바꿈 허용)
            for col in range(12, 17):  # All Domains, Stage Paths, Stage1, Stage2, Stage3 열
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 열 너비 설정
        column_widths = {
            'A': 10,  # Test Index
            'B': 35,  # Entity Name
            'C': 15,  # Input Domain
            'D': 10,  # Success
            'E': 12,  # Domain Count
            'F': 15,  # Best Search Domain
            'G': 15,  # Best Result Domain
            'H': 15,  # Best Concept ID
            'I': 45,  # Best Concept Name
            'J': 12,  # Best Score
            'K': 15,  # Best Confidence
            'L': 50,  # All Domains
            'M': 45,  # Domain Stage Paths
            'N': 70,  # Stage1 Candidates
            'O': 70,  # Stage2 Candidates
            'P': 85   # Stage3 Candidates
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 행 높이 자동 조정 (후보군 정보가 많은 경우)
        for row_num in range(2, len(test_results) + 2):
            ws.row_dimensions[row_num].height = 150  # 충분한 높이 설정
    
    def _format_domain_results(self, domain_results):
        """도메인별 결과를 엑셀 셀용 텍스트로 포맷팅"""
        if not domain_results:
            return "도메인 결과 없음"
        
        lines = []
        for i, domain in enumerate(domain_results, 1):
            line = f"{i}. [{domain.get('domain_id', 'N/A')}] {domain.get('mapped_concept_name', 'N/A')}\n"
            line += f"   ID: {domain.get('mapped_concept_id', 'N/A')}, "
            line += f"Score: {domain.get('mapping_score', 0):.4f}, "
            line += f"Conf: {domain.get('mapping_confidence', 'N/A')}\n"
            line += f"   Vocab: {domain.get('vocabulary_id', 'N/A')}"
            lines.append(line)
        
        return "\n\n".join(lines)
    
    def _format_stage_paths(self, stage_paths):
        """도메인별 Stage 경로를 엑셀 셀용 텍스트로 포맷팅"""
        if not stage_paths:
            return "경로 정보 없음"
        
        lines = []
        for domain_name, stage_info in sorted(stage_paths.items()):
            search_domain = stage_info.get('search_domain', domain_name)
            result_domain = stage_info.get('result_domain', 'N/A')
            
            # 검색 도메인과 결과 도메인이 다른 경우 표시
            if search_domain != result_domain:
                line = f"[{search_domain} → {result_domain}]\n"
            else:
                line = f"[{search_domain}]\n"
            
            line += f"  Stage1: {stage_info.get('stage1_count', 0)}개\n"
            line += f"  Stage2: {stage_info.get('stage2_count', 0)}개\n"
            line += f"  Stage3: {stage_info.get('stage3_count', 0)}개"
            lines.append(line)
        
        return "\n\n".join(lines)
    
    def _format_candidates_for_cell(self, candidates, stage_type):
        """후보군 정보를 엑셀 셀용 텍스트로 포맷팅"""
        if not candidates:
            return "후보 없음"
        
        lines = []
        max_candidates = 15 if stage_type == 'stage1' else (15 if stage_type == 'stage2' else 10)  # Stage1, Stage2는 15개, Stage3는 10개 표시
        
        for i, candidate in enumerate(candidates[:max_candidates], 1):
            if stage_type == 'stage1':
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   ES점수: {candidate.get('elasticsearch_score', 0):.4f}, "
                line += f"Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            elif stage_type == 'stage2':
                search_type = candidate.get('search_type', 'unknown')
                is_std = "✓" if candidate.get('is_original_standard', True) else "→"
                line = f"{i}. [{search_type}] {is_std} {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
                if not candidate.get('is_original_standard', True):
                    original_non_std = candidate.get('original_non_standard', {})
                    if original_non_std:
                        line += f"\n   원본 Non-std: {original_non_std.get('concept_name', 'N/A')} (ID: {original_non_std.get('concept_id', 'N/A')})"
            else:  # stage3
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                
                # LLM 모드인 경우 LLM 점수/순위/이유 표시
                llm_score = candidate.get('llm_score')
                llm_rank = candidate.get('llm_rank')
                llm_reasoning = candidate.get('llm_reasoning')
                semantic_sim = candidate.get('semantic_similarity')
                
                if llm_score is not None:
                    if semantic_sim is not None:
                        line += f"   LLM점수: {llm_score}, 순위: {llm_rank}, 의미유사도: {semantic_sim:.4f}\n"
                    else:
                        line += f"   LLM점수: {llm_score}, 순위: {llm_rank}\n"
                    if llm_reasoning:
                        reasoning_short = llm_reasoning[:60] + '...' if len(llm_reasoning) > 60 else llm_reasoning
                        line += f"   이유: {reasoning_short}\n"
                else:
                    # Hybrid 모드인 경우
                    line += f"   텍스트: {candidate.get('text_similarity', 0):.4f}, "
                    line += f"의미적: {candidate.get('semantic_similarity', 0):.4f}, "
                    line += f"최종: {candidate.get('final_score', 0):.4f}\n"
                
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            
            lines.append(line)
        
        return "\n\n".join(lines)

def main():
    """메인 함수"""
    # ============================================================
    # 설정 옵션
    # ============================================================
    # scoring_mode 선택: 'llm' (디폴트), 'llm_with_score', 'semantic'
    SCORING_MODE = "llm"
    
    # 테스터 초기화
    tester = EntityMappingTester(
        log_dir="test_logs",
        scoring_mode=SCORING_MODE
    )
    
    # ============================================================
    # 테스트할 엔티티 리스트
    # 형식: 문자열 또는 (entity, domain) 튜플
    # domain: 'Condition', 'Procedure', 'Drug', 'Observation', 
    #         'Measurement', 'Device' 또는 None (모든 도메인)
    # ============================================================
    test_entities = [
        # (entity, domain) 튜플 형식 - 특정 도메인 지정
        ('decompression', 'Procedure'),
        ('procedure', 'Procedure'),
        ('acute coronary syndrome', 'Condition'),
        ('flexible bronchoscopic removal of trachea or bronchial foreign body', 'Procedure')
        # ('mass removal', 'Procedure'),
        # ('congenital ring syndrome', 'Condition'),
        # ('monophasic synovial sarcoma', 'Observation'),
        # ('anaplastic astrocytoma', 'Observation'),
        # ('sacroiliac joint block', 'Procedure'),
        # ('endometrial polypectomy', 'Procedure'),
        # ('mandibular nerve block', 'Procedure'),
    ]
    
    # 테스트 실행
    results = tester.run_test_with_entities(test_entities)
    
    print(f"\n✅ 테스트 완료! 로그는 {tester.log_dir} 디렉토리에 저장되었습니다.")
    print(f"   - Scoring Mode: {SCORING_MODE}")

if __name__ == "__main__":
    main()