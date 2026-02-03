"""
6가지 Ablation 테스트 실행 스크립트

테스트 조건:
1. Stage1 (Semantic + Combined) + Stage3 LLM (점수 포함)
2. Stage1 (Semantic + Combined) + Stage3 LLM (점수 미포함)
3. Stage1 (Semantic + Combined) + Stage3 Semantic Only
4. Stage1 (Lexical + Semantic + Combined) + Stage3 LLM (점수 포함)
5. Stage1 (Lexical + Semantic + Combined) + Stage3 LLM (점수 미포함)
6. Stage1 (Lexical + Semantic + Combined) + Stage3 Semantic Only
"""
import pandas as pd
import logging
import os
from datetime import datetime
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from tqdm import tqdm
import time
import json

# 상대 경로로 src 디렉토리 추가
sys.path.append(str(Path(__file__).parent.parent / 'src'))

from MapOMOP.entity_mapping_api import EntityMappingAPI, EntityInput, DomainID
from MapOMOP.elasticsearch_client import ElasticsearchClient


# 6가지 테스트 조건 정의
TEST_CONDITIONS = [
    {
        'name': 'semantic_combined_llm_with_scores',
        'description': 'Stage1 (Semantic + Combined) + Stage3 LLM (점수 포함)',
        'use_lexical': False,
        'scoring_mode': 'llm',
        'include_stage1_scores': True
    },
    {
        'name': 'semantic_combined_llm_no_scores',
        'description': 'Stage1 (Semantic + Combined) + Stage3 LLM (점수 미포함)',
        'use_lexical': False,
        'scoring_mode': 'llm',
        'include_stage1_scores': False
    },
    {
        'name': 'semantic_combined_semantic_only',
        'description': 'Stage1 (Semantic + Combined) + Stage3 Semantic Only',
        'use_lexical': False,
        'scoring_mode': 'semantic_only',
        'include_stage1_scores': False
    },
    {
        'name': 'full_search_llm_with_scores',
        'description': 'Stage1 (Lexical + Semantic + Combined) + Stage3 LLM (점수 포함)',
        'use_lexical': True,
        'scoring_mode': 'llm',
        'include_stage1_scores': True
    },
    {
        'name': 'full_search_llm_no_scores',
        'description': 'Stage1 (Lexical + Semantic + Combined) + Stage3 LLM (점수 미포함)',
        'use_lexical': True,
        'scoring_mode': 'llm',
        'include_stage1_scores': False
    },
    {
        'name': 'full_search_semantic_only',
        'description': 'Stage1 (Lexical + Semantic + Combined) + Stage3 Semantic Only',
        'use_lexical': True,
        'scoring_mode': 'semantic_only',
        'include_stage1_scores': False
    }
]


class AblationTester:
    """6가지 Ablation 테스트 클래스"""
    
    def __init__(self, log_dir: str = "test_logs_ablation"):
        """테스터 초기화"""
        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(exist_ok=True)
        
        self.setup_logging()
        
        self.es_client = ElasticsearchClient()
        self.es_client.concept_index = "concept-small"
        self.es_client.concept_synonym_index = "concept-synonym"
        
        self.domain_mapping = {
            'Condition': DomainID.CONDITION,
            'Procedure': DomainID.PROCEDURE,
            'Drug': DomainID.DRUG,
            'Observation': DomainID.OBSERVATION,
            'Measurement': DomainID.MEASUREMENT,
            'Device': DomainID.DEVICE,
        }
    
    def setup_logging(self):
        """로깅 설정"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        log_file = self.log_dir / f"ablation_test_{timestamp}.log"
        
        self.logger = logging.getLogger('ablation_test')
        self.logger.setLevel(logging.INFO)
        
        # 기존 핸들러 제거
        self.logger.handlers.clear()
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        self.logger.info(f"로그 파일: {log_file}")
    
    def load_and_sample_data(
        self, 
        csv_path: str, 
        sample_size: int = 1000, 
        random_state: int = 42
    ) -> pd.DataFrame:
        """
        CSV 파일에서 데이터 로딩 및 랜덤 샘플링
        
        Args:
            csv_path: CSV 파일 경로
            sample_size: 샘플 크기 (기본값: 1000)
            random_state: 랜덤 시드 (기본값: 42) - 모든 테스트에서 동일한 데이터 사용
        """
        self.logger.info(f"데이터 로딩 시작: {csv_path}")
        self.logger.info(f"샘플 크기: {sample_size}개")
        self.logger.info(f"랜덤 시드: {random_state} (모든 테스트에서 동일한 데이터 사용)")
        
        # 전체 데이터 로드
        chunk_size = 100000
        chunks = []
        
        self.logger.info("청크 단위로 데이터 읽는 중...")
        for chunk in tqdm(pd.read_csv(csv_path, chunksize=chunk_size), desc="데이터 로딩"):
            chunks.append(chunk)
        
        # 전체 데이터 병합
        df = pd.concat(chunks, ignore_index=True)
        self.logger.info(f"전체 데이터 크기: {len(df):,}개")
        
        # 랜덤 샘플링 (고정된 시드 사용)
        df_sample = df.sample(n=min(sample_size, len(df)), random_state=random_state)
        df_sample = df_sample.reset_index(drop=True)
        self.logger.info(f"랜덤 샘플링 완료: {len(df_sample):,}개 (seed={random_state})")
        
        # 도메인 분포 출력
        if 'domain_id' in df_sample.columns:
            domain_dist = df_sample['domain_id'].value_counts()
            self.logger.info("\n도메인 분포:")
            for domain, count in domain_dist.items():
                self.logger.info(f"  {domain}: {count}개 ({count/len(df_sample)*100:.1f}%)")
        
        return df_sample
    
    def create_entity_input(self, row) -> EntityInput:
        """DataFrame 행에서 EntityInput 생성"""
        entity_name = str(row['source_value']).strip()
        
        # 도메인 정보 사용 (CSV 컬럼명: domain_id)
        domain_id = None
        if 'domain_id' in row and pd.notna(row['domain_id']):
            domain_str = str(row['domain_id']).strip()
            if domain_str and domain_str in self.domain_mapping:
                domain_id = self.domain_mapping[domain_str]
        
        return EntityInput(
            entity_name=entity_name,
            domain_id=domain_id,
            vocabulary_id=None
        )
    
    def test_single_entity(
        self, 
        api: EntityMappingAPI,
        entity_input: EntityInput, 
        ground_truth_concept_id: int
    ) -> dict:
        """단일 엔티티 테스트"""
        try:
            # 매핑 수행
            results = api.map_entity(entity_input)
            
            # 단계별 후보군 수집
            stage1_candidates = []
            stage2_candidates = []
            stage3_candidates = []
            
            if hasattr(api, '_last_stage1_candidates') and api._last_stage1_candidates:
                stage1_candidates = api._last_stage1_candidates
            
            if hasattr(api, '_last_stage2_candidates') and api._last_stage2_candidates:
                stage2_candidates = api._last_stage2_candidates
            
            if hasattr(api, '_last_rerank_candidates') and api._last_rerank_candidates:
                stage3_candidates = api._last_rerank_candidates
            
            # 최고 점수 결과 선택
            best_result = max(results, key=lambda x: x.mapping_score) if results else None
            
            # 매핑 성공 여부 판단 (concept_id 일치)
            mapping_correct = False
            if best_result and ground_truth_concept_id:
                try:
                    best_concept_id_int = int(best_result.mapped_concept_id)
                    ground_truth_int = int(ground_truth_concept_id)
                    mapping_correct = (best_concept_id_int == ground_truth_int)
                except (ValueError, TypeError):
                    mapping_correct = False
            
            return {
                'entity_name': entity_input.entity_name,
                'input_domain': entity_input.domain_id.value if entity_input.domain_id else 'All',
                'ground_truth_concept_id': ground_truth_concept_id,
                'success': results is not None and len(results) > 0,
                'mapping_correct': mapping_correct,
                'best_concept_id': best_result.mapped_concept_id if best_result else None,
                'best_concept_name': best_result.mapped_concept_name if best_result else None,
                'best_score': best_result.mapping_score if best_result else 0.0,
                'best_confidence': best_result.mapping_confidence if best_result else None,
                'stage1_candidates': stage1_candidates,
                'stage2_candidates': stage2_candidates,
                'stage3_candidates': stage3_candidates
            }
            
        except Exception as e:
            return {
                'entity_name': entity_input.entity_name,
                'input_domain': entity_input.domain_id.value if entity_input.domain_id else 'All',
                'ground_truth_concept_id': ground_truth_concept_id,
                'success': False,
                'mapping_correct': False,
                'best_concept_id': None,
                'best_concept_name': None,
                'best_score': 0.0,
                'best_confidence': None,
                'error': str(e),
                'stage1_candidates': [],
                'stage2_candidates': [],
                'stage3_candidates': []
            }
    
    def run_single_condition_test(
        self, 
        condition: dict, 
        test_data: pd.DataFrame
    ) -> dict:
        """단일 조건으로 테스트 실행"""
        condition_name = condition['name']
        
        self.logger.info("\n" + "=" * 100)
        self.logger.info(f"🧪 테스트 조건: {condition['description']}")
        self.logger.info(f"   - use_lexical: {condition['use_lexical']}")
        self.logger.info(f"   - scoring_mode: {condition['scoring_mode']}")
        self.logger.info(f"   - include_stage1_scores: {condition['include_stage1_scores']}")
        self.logger.info("=" * 100)
        
        start_time = time.time()
        
        # API 초기화 (조건에 맞게)
        api = EntityMappingAPI(
            es_client=self.es_client,
            scoring_mode=condition['scoring_mode'],
            include_stage1_scores=condition['include_stage1_scores'],
            use_lexical=condition['use_lexical']
        )
        
        # 테스트 실행
        test_results = []
        successful_mappings = 0
        correct_mappings = 0
        
        for idx, row in tqdm(test_data.iterrows(), total=len(test_data), 
                           desc=f"[{condition_name}]"):
            try:
                entity_input = self.create_entity_input(row)
                ground_truth = int(row['concept_id']) if pd.notna(row['concept_id']) else None
                
                result = self.test_single_entity(api, entity_input, ground_truth)
                test_results.append(result)
                
                if result['success']:
                    successful_mappings += 1
                    if result['mapping_correct']:
                        correct_mappings += 1
                        
            except Exception as e:
                self.logger.error(f"테스트 #{idx + 1} 처리 오류: {str(e)}")
                continue
        
        # 테스트 완료 시간
        end_time = time.time()
        elapsed_time = end_time - start_time
        
        # 결과 요약
        total_tests = len(test_results)
        success_rate = (successful_mappings / total_tests * 100) if total_tests > 0 else 0
        accuracy = (correct_mappings / total_tests * 100) if total_tests > 0 else 0
        
        summary = {
            'condition_name': condition_name,
            'description': condition['description'],
            'use_lexical': condition['use_lexical'],
            'scoring_mode': condition['scoring_mode'],
            'include_stage1_scores': condition['include_stage1_scores'],
            'total_tests': total_tests,
            'successful_mappings': successful_mappings,
            'correct_mappings': correct_mappings,
            'success_rate': success_rate,
            'accuracy': accuracy,
            'elapsed_time': elapsed_time,
            'avg_time_per_entity': elapsed_time / total_tests if total_tests > 0 else 0,
            'results': test_results
        }
        
        self.logger.info(f"\n📊 [{condition_name}] 결과 요약:")
        self.logger.info(f"   총 테스트: {total_tests:,}개")
        self.logger.info(f"   매핑 성공: {successful_mappings:,}개 ({success_rate:.2f}%)")
        self.logger.info(f"   정답 매칭: {correct_mappings:,}개 ({accuracy:.2f}%)")
        self.logger.info(f"   소요 시간: {elapsed_time:.2f}초 ({elapsed_time/60:.2f}분)")
        
        return summary
    
    def run_all_conditions(
        self, 
        csv_path: str, 
        sample_size: int = 1000, 
        random_state: int = 42,
        conditions: list = None
    ):
        """모든 조건으로 테스트 실행"""
        self.logger.info("=" * 100)
        self.logger.info("🚀 Ablation Study 테스트 시작")
        self.logger.info("=" * 100)
        
        # 데이터 로딩 (한 번만 로딩하여 모든 테스트에서 동일한 데이터 사용)
        test_data = self.load_and_sample_data(csv_path, sample_size, random_state)
        
        # 테스트할 조건 선택
        if conditions is None:
            conditions = TEST_CONDITIONS
        
        self.logger.info(f"\n총 {len(conditions)}개 조건 테스트 예정:")
        for i, cond in enumerate(conditions, 1):
            self.logger.info(f"  {i}. {cond['description']}")
        
        # 각 조건별로 테스트 실행
        all_summaries = []
        total_start_time = time.time()
        
        for i, condition in enumerate(conditions, 1):
            self.logger.info(f"\n{'='*50}")
            self.logger.info(f"[{i}/{len(conditions)}] 테스트 시작: {condition['name']}")
            self.logger.info(f"{'='*50}")
            
            summary = self.run_single_condition_test(condition, test_data)
            all_summaries.append(summary)
        
        total_elapsed_time = time.time() - total_start_time
        
        # 전체 결과 요약 출력
        self.logger.info("\n" + "=" * 100)
        self.logger.info("📊 전체 Ablation Study 결과 요약")
        self.logger.info("=" * 100)
        self.logger.info(f"{'조건명':<45} {'Accuracy':>10} {'Success Rate':>12} {'시간(초)':>10}")
        self.logger.info("-" * 80)
        
        for summary in all_summaries:
            self.logger.info(
                f"{summary['description']:<45} "
                f"{summary['accuracy']:>9.2f}% "
                f"{summary['success_rate']:>11.2f}% "
                f"{summary['elapsed_time']:>9.1f}s"
            )
        
        self.logger.info("-" * 80)
        self.logger.info(f"총 소요 시간: {total_elapsed_time:.2f}초 ({total_elapsed_time/60:.2f}분)")
        self.logger.info("=" * 100)
        
        # 결과 저장
        self.save_results(all_summaries, random_state)
        
        return all_summaries
    
    def save_results(self, all_summaries: list, random_state: int):
        """테스트 결과를 파일로 저장"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 1. 요약 CSV 저장
        summary_data = []
        for summary in all_summaries:
            summary_data.append({
                'condition_name': summary['condition_name'],
                'description': summary['description'],
                'use_lexical': summary['use_lexical'],
                'scoring_mode': summary['scoring_mode'],
                'include_stage1_scores': summary['include_stage1_scores'],
                'total_tests': summary['total_tests'],
                'successful_mappings': summary['successful_mappings'],
                'correct_mappings': summary['correct_mappings'],
                'success_rate': summary['success_rate'],
                'accuracy': summary['accuracy'],
                'elapsed_time': summary['elapsed_time'],
                'avg_time_per_entity': summary['avg_time_per_entity']
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_csv = self.log_dir / f"ablation_summary_{timestamp}.csv"
        summary_df.to_csv(summary_csv, index=False, encoding='utf-8')
        self.logger.info(f"📄 요약 CSV 저장: {summary_csv}")
        
        # 2. 상세 결과 XLSX 저장
        xlsx_file = self.log_dir / f"ablation_detailed_{timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # 요약 시트
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, all_summaries)
        
        # 각 조건별 상세 시트
        for summary in all_summaries:
            ws = wb.create_sheet(title=summary['condition_name'][:31])  # 시트명 길이 제한
            self._create_detail_sheet(ws, summary)
        
        wb.save(xlsx_file)
        self.logger.info(f"📊 상세 XLSX 저장: {xlsx_file}")
        
        # 3. JSON으로 전체 결과 저장
        json_file = self.log_dir / f"ablation_results_{timestamp}.json"
        json_data = {
            'timestamp': timestamp,
            'random_state': random_state,
            'summaries': [
                {k: v for k, v in s.items() if k != 'results'}
                for s in all_summaries
            ]
        }
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        self.logger.info(f"📁 JSON 저장: {json_file}")
    
    def _create_summary_sheet(self, ws, all_summaries):
        """요약 시트 생성"""
        headers = [
            "조건명", "설명", "use_lexical", "scoring_mode", "include_scores",
            "총 테스트", "매핑 성공", "정답 매칭", "Success Rate (%)", "Accuracy (%)",
            "소요시간(초)", "평균시간(초/엔티티)"
        ]
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터 작성
        for row, summary in enumerate(all_summaries, 2):
            ws.cell(row=row, column=1, value=summary['condition_name'])
            ws.cell(row=row, column=2, value=summary['description'])
            ws.cell(row=row, column=3, value=str(summary['use_lexical']))
            ws.cell(row=row, column=4, value=summary['scoring_mode'])
            ws.cell(row=row, column=5, value=str(summary['include_stage1_scores']))
            ws.cell(row=row, column=6, value=summary['total_tests'])
            ws.cell(row=row, column=7, value=summary['successful_mappings'])
            ws.cell(row=row, column=8, value=summary['correct_mappings'])
            ws.cell(row=row, column=9, value=round(summary['success_rate'], 2))
            ws.cell(row=row, column=10, value=round(summary['accuracy'], 2))
            ws.cell(row=row, column=11, value=round(summary['elapsed_time'], 2))
            ws.cell(row=row, column=12, value=round(summary['avg_time_per_entity'], 4))
            
            # Accuracy 컬러링
            accuracy_cell = ws.cell(row=row, column=10)
            if summary['accuracy'] >= 80:
                accuracy_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif summary['accuracy'] >= 60:
                accuracy_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                accuracy_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 열 너비 설정
        column_widths = [35, 55, 12, 15, 15, 10, 12, 12, 15, 12, 12, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    def _create_detail_sheet(self, ws, summary):
        """상세 결과 시트 생성"""
        headers = [
            "Entity Name", "Domain", "Ground Truth ID", "Success", "Correct",
            "Best Concept ID", "Best Concept Name", "Score", "Confidence",
            "Stage1 Candidates", "Stage2 Candidates", "Stage3 Candidates"
        ]
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 데이터 작성
        for row, result in enumerate(summary['results'], 2):
            ws.cell(row=row, column=1, value=result.get('entity_name', 'N/A'))
            ws.cell(row=row, column=2, value=result.get('input_domain', 'N/A'))
            ws.cell(row=row, column=3, value=result.get('ground_truth_concept_id', 'N/A'))
            ws.cell(row=row, column=4, value="성공" if result.get('success') else "실패")
            
            # 정답 여부 컬러링
            correct_cell = ws.cell(row=row, column=5, value="정답" if result.get('mapping_correct') else "오답")
            if result.get('mapping_correct'):
                correct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                correct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws.cell(row=row, column=6, value=result.get('best_concept_id', 'N/A'))
            ws.cell(row=row, column=7, value=result.get('best_concept_name', 'N/A'))
            ws.cell(row=row, column=8, value=result.get('best_score', 0.0))
            ws.cell(row=row, column=9, value=result.get('best_confidence', 'N/A'))
            
            # Stage별 후보군 추가
            stage1_text = self._format_candidates_for_cell(result.get('stage1_candidates', []), 'stage1')
            ws.cell(row=row, column=10, value=stage1_text)
            
            stage2_text = self._format_candidates_for_cell(result.get('stage2_candidates', []), 'stage2')
            ws.cell(row=row, column=11, value=stage2_text)
            
            stage3_text = self._format_candidates_for_cell(result.get('stage3_candidates', []), 'stage3', summary.get('scoring_mode', 'llm'))
            ws.cell(row=row, column=12, value=stage3_text)
            
            # Stage 컬럼 스타일 설정
            for col in range(10, 13):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 열 너비 설정
        column_widths = [45, 15, 18, 10, 10, 18, 50, 10, 12, 70, 70, 85]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 행 높이 설정 (Stage 후보군 표시를 위해)
        for row_num in range(2, len(summary['results']) + 2):
            ws.row_dimensions[row_num].height = 150
    
    def _format_candidates_for_cell(self, candidates, stage_type, scoring_mode='llm'):
        """후보군 포맷팅 (엑셀 셀용)"""
        if not candidates:
            return "후보 없음"
        
        lines = []
        max_candidates = 15 if stage_type in ['stage1', 'stage2'] else 10
        
        for i, candidate in enumerate(candidates[:max_candidates], 1):
            if stage_type == 'stage1':
                # Stage 1: Elasticsearch 결과
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   ES점수: {candidate.get('elasticsearch_score', 0):.4f}, "
                line += f"Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            
            elif stage_type == 'stage2':
                # Stage 2: Standard 변환 결과
                search_type = candidate.get('search_type', 'unknown')
                is_std = "✓" if candidate.get('is_original_standard', True) else "→"
                line = f"{i}. [{search_type}] {is_std} {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
                if not candidate.get('is_original_standard', True):
                    original_non_std = candidate.get('original_non_standard', {})
                    if original_non_std:
                        line += f"\n   원본 Non-std: {original_non_std.get('concept_name', 'N/A')} (ID: {original_non_std.get('concept_id', 'N/A')})"
            
            else:  # stage3
                # Stage 3: LLM 또는 Semantic Only 결과
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                
                # LLM 모드인 경우
                llm_score = candidate.get('llm_score')
                llm_rank = candidate.get('llm_rank')
                llm_reasoning = candidate.get('llm_reasoning')
                semantic_similarity = candidate.get('semantic_similarity')
                
                if llm_score is not None:
                    line += f"   LLM점수: {llm_score}, 순위: {llm_rank}"
                    if semantic_similarity is not None:
                        line += f" | 의미유사도: {semantic_similarity:.4f}"
                    line += "\n"
                    if llm_reasoning:
                        reasoning_short = llm_reasoning[:60] + '...' if len(llm_reasoning) > 60 else llm_reasoning
                        line += f"   이유: {reasoning_short}\n"
                else:
                    # Semantic Only 또는 Hybrid 모드인 경우
                    text_sim = candidate.get('text_similarity', 0)
                    sem_sim = candidate.get('semantic_similarity', 0)
                    final_score = candidate.get('final_score', 0)
                    line += f"   텍스트: {text_sim:.4f}, 의미적: {sem_sim:.4f}, 최종: {final_score:.4f}\n"
                
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            
            lines.append(line)
        
        return "\n\n".join(lines)


def main():
    """메인 함수"""
    # ============================================================
    # 설정
    # ============================================================
    CSV_PATH = "/home/work/skku/hyo/omop-mapper/data/mapomop_test_data.csv"
    SAMPLE_SIZE = 4129  # 샘플 크기
    RANDOM_STATE = 42   # 랜덤 시드 (모든 테스트에서 동일한 데이터 사용)
    
    # 테스트할 조건 선택 (None이면 모든 조건 테스트)
    # 특정 조건만 테스트하려면 인덱스 지정
    # 예: [TEST_CONDITIONS[0], TEST_CONDITIONS[3]]  # 1번, 4번 조건만
    CONDITIONS_TO_TEST = [TEST_CONDITIONS[4]]  # 모든 6가지 조건 테스트
    
    # ============================================================
    # 테스트 실행
    # ============================================================
    tester = AblationTester()
    
    results = tester.run_all_conditions(
        csv_path=CSV_PATH,
        sample_size=SAMPLE_SIZE,
        random_state=RANDOM_STATE,
        conditions=CONDITIONS_TO_TEST
    )
    
    print(f"\n✅ Ablation Study 완료!")
    print(f"   결과는 {tester.log_dir} 디렉토리에 저장되었습니다.")


if __name__ == "__main__":
    main()