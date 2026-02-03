"""
매핑 실패 데이터 재시도 테스트 스크립트

데이터: data/fail-retry.xlsx
도메인: Entity별 Domain 컬럼에 따라 지정
Stage 1: Lexical + Semantic + Combined
Stage 3: LLM (점수 미포함) + Non-std 정보 포함 프롬프트
"""
import pandas as pd
import logging
import os
from datetime import datetime
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from tqdm import tqdm
import time
import json

# 상대 경로로 src 디렉토리 추가
sys.path.append(str(Path(__file__).parent.parent / 'src'))

from MapOMOP.entity_mapping_api import EntityMappingAPI, EntityInput, DomainID
from MapOMOP.elasticsearch_client import ElasticsearchClient


class FailRetryMappingTester:
    """매핑 실패 데이터 재시도 테스트 클래스"""
    
    # Domain 문자열 -> DomainID 매핑
    DOMAIN_MAP = {
        'Condition': DomainID.CONDITION,
        'Procedure': DomainID.PROCEDURE,
        'Drug': DomainID.DRUG,
        'Measurement': DomainID.MEASUREMENT,
        'Observation': DomainID.OBSERVATION,
        'Device': DomainID.DEVICE,
    }
    
    def __init__(self, log_dir: str = "test_logs_fail_retry"):
        """테스터 초기화"""
        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(exist_ok=True)
        
        self.setup_logging()
        
        self.es_client = ElasticsearchClient()
        self.es_client.concept_index = "concept-small"
        self.es_client.concept_synonym_index = "concept-synonym"
    
    def setup_logging(self):
        """로깅 설정"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        log_file = self.log_dir / f"fail_retry_mapping_test_{timestamp}.log"
        
        self.logger = logging.getLogger('fail_retry_mapping_test')
        self.logger.setLevel(logging.INFO)
        
        # 기존 핸들러 제거
        self.logger.handlers.clear()
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        self.logger.info(f"로그 파일: {log_file}")
    
    def load_excel_data(self, excel_path: str) -> pd.DataFrame:
        """
        엑셀 파일에서 데이터 로딩
        
        Args:
            excel_path: 엑셀 파일 경로
        """
        self.logger.info(f"데이터 로딩 시작: {excel_path}")
        
        df = pd.read_excel(excel_path)
        self.logger.info(f"전체 데이터 크기: {len(df):,}개")
        self.logger.info(f"컬럼: {df.columns.tolist()}")
        
        # Entity Name 열 확인
        if 'Entity Name' not in df.columns:
            raise ValueError("'Entity Name' 컬럼이 존재하지 않습니다.")
        
        # Domain 열 확인
        if 'Domain' not in df.columns:
            raise ValueError("'Domain' 컬럼이 존재하지 않습니다.")
        
        # 빈 값 제외
        df_clean = df[df['Entity Name'].notna()].copy()
        df_clean = df_clean.reset_index(drop=True)
        
        self.logger.info(f"유효한 엔티티 수: {len(df_clean):,}개")
        
        # 도메인별 분포
        domain_counts = df_clean['Domain'].value_counts()
        self.logger.info("도메인별 분포:")
        for domain, count in domain_counts.items():
            self.logger.info(f"  - {domain}: {count:,}개")
        
        return df_clean
    
    def get_domain_id(self, domain_str: str) -> DomainID:
        """Domain 문자열을 DomainID로 변환"""
        if pd.isna(domain_str) or domain_str is None:
            return None
        
        domain_str = str(domain_str).strip()
        return self.DOMAIN_MAP.get(domain_str, None)
    
    def create_entity_input(self, entity_name: str, domain_str: str) -> EntityInput:
        """엔티티명과 도메인으로부터 EntityInput 생성"""
        domain_id = self.get_domain_id(domain_str)
        return EntityInput(
            entity_name=entity_name.strip(),
            domain_id=domain_id,
            vocabulary_id=None
        )
    
    def test_single_entity(
        self, 
        api: EntityMappingAPI,
        entity_input: EntityInput,
        ground_truth_id: str = None,
        ground_truth_name: str = None
    ) -> dict:
        """단일 엔티티 테스트"""
        # 초기화
        results = None
        error_msg = None
        
        # 매핑 수행
        try:
            results = api.map_entity(entity_input)
        except Exception as e:
            error_msg = str(e)
        
        # 매핑 성공/실패와 관계없이 항상 stage candidates 수집
        # API에서 매핑 시작 시 초기화되므로, 현재 매핑의 candidates만 포함됨
        stage1_candidates = getattr(api, '_last_stage1_candidates', []) or []
        stage2_candidates = getattr(api, '_last_stage2_candidates', []) or []
        stage3_candidates = getattr(api, '_last_rerank_candidates', []) or []
        
        # LLM이 1위로 선택한 것 확인
        llm_top_pick = None
        if stage3_candidates:
            llm_top = stage3_candidates[0]
            llm_top_pick = {
                'concept_id': llm_top.get('concept_id'),
                'concept_name': llm_top.get('concept_name'),
                'llm_score': llm_top.get('llm_score'),
                'llm_rank': llm_top.get('llm_rank')
            }
        
        # 최고 점수 결과 선택
        best_result = None
        if results and len(results) > 0:
            best_result = max(results, key=lambda x: x.mapping_score)
        
        # Validation으로 인해 결과가 변경되었는지 확인
        validation_changed = False
        if best_result and llm_top_pick:
            if str(best_result.mapped_concept_id) != str(llm_top_pick.get('concept_id')):
                validation_changed = True
        
        # 실패 원인 판단
        if not best_result and not error_msg:
            if not stage1_candidates:
                error_msg = "Stage 1: No candidates found"
            elif not stage2_candidates:
                error_msg = "Stage 2: No standard candidates found"
            elif not stage3_candidates:
                error_msg = "Stage 3: Scoring failed"
            else:
                error_msg = "Validation: All candidates failed"
        
        # Ground Truth와 비교
        gt_matched = None
        if best_result and ground_truth_id:
            try:
                gt_matched = str(best_result.mapped_concept_id) == str(int(float(ground_truth_id)))
            except:
                gt_matched = False
        
        return {
            'entity_name': entity_input.entity_name,
            'input_domain': entity_input.domain_id.value if entity_input.domain_id else 'All',
            'ground_truth_id': ground_truth_id,
            'ground_truth_name': ground_truth_name,
            'success': best_result is not None,
            'gt_matched': gt_matched,
            'best_concept_id': best_result.mapped_concept_id if best_result else None,
            'best_concept_name': best_result.mapped_concept_name if best_result else None,
            'best_score': best_result.mapping_score if best_result else 0.0,
            'best_confidence': best_result.mapping_confidence if best_result else None,
            'mapping_method': best_result.mapping_method if best_result else None,
            'vocabulary_id': best_result.vocabulary_id if best_result else None,
            'llm_top_pick': llm_top_pick,
            'validation_changed': validation_changed,
            'error': error_msg,
            'stage1_candidates': stage1_candidates,
            'stage2_candidates': stage2_candidates,
            'stage3_candidates': stage3_candidates
        }
    
    def run_mapping_test(self, excel_path: str) -> dict:
        """매핑 테스트 실행"""
        self.logger.info("\n" + "=" * 100)
        self.logger.info("🔄 매핑 실패 데이터 재시도 테스트 시작")
        self.logger.info("=" * 100)
        self.logger.info("테스트 조건:")
        self.logger.info("  - Domain: 각 엔티티별 지정된 도메인")
        self.logger.info("  - Stage 1: Lexical + Semantic + Combined (use_lexical: True)")
        self.logger.info("  - Stage 3: LLM (점수 미포함, include_stage1_scores: False)")
        self.logger.info("  - Non-std 정보 포함 프롬프트 (include_non_std_info: True)")
        self.logger.info("=" * 100)
        
        # 데이터 로딩
        test_data = self.load_excel_data(excel_path)
        
        start_time = time.time()
        
        # API 초기화
        api = EntityMappingAPI(
            es_client=self.es_client,
            scoring_mode='llm',
            include_stage1_scores=False,
            use_lexical=True,
            include_non_std_info=True
        )
        
        # 테스트 실행
        test_results = []
        successful_mappings = 0
        gt_matched_count = 0
        
        for idx, row in tqdm(test_data.iterrows(), total=len(test_data), desc="매핑 테스트"):
            try:
                entity_name = str(row['Entity Name']).strip()
                domain_str = str(row['Domain']).strip() if pd.notna(row['Domain']) else None
                ground_truth_id = row.get('Ground Truth ID', None)
                ground_truth_name = row.get('concept_name', None)
                
                entity_input = self.create_entity_input(entity_name, domain_str)
                
                result = self.test_single_entity(
                    api, entity_input, 
                    ground_truth_id=ground_truth_id,
                    ground_truth_name=ground_truth_name
                )
                test_results.append(result)
                
                if result['success']:
                    successful_mappings += 1
                    if result['gt_matched']:
                        gt_matched_count += 1
                    
            except Exception as e:
                self.logger.error(f"테스트 #{idx + 1} 처리 오류: {str(e)}")
                continue
        
        # 테스트 완료 시간
        end_time = time.time()
        elapsed_time = end_time - start_time
        
        # 결과 요약
        total_tests = len(test_results)
        success_rate = (successful_mappings / total_tests * 100) if total_tests > 0 else 0
        gt_match_rate = (gt_matched_count / total_tests * 100) if total_tests > 0 else 0
        
        summary = {
            'test_name': 'fail_retry_mapping',
            'description': 'Stage1 (Lexical + Semantic + Combined) + Stage3 LLM (점수 미포함, Non-std 정보 포함)',
            'use_lexical': True,
            'scoring_mode': 'llm',
            'include_stage1_scores': False,
            'include_non_std_info': True,
            'total_tests': total_tests,
            'successful_mappings': successful_mappings,
            'success_rate': success_rate,
            'gt_matched_count': gt_matched_count,
            'gt_match_rate': gt_match_rate,
            'elapsed_time': elapsed_time,
            'avg_time_per_entity': elapsed_time / total_tests if total_tests > 0 else 0,
            'results': test_results
        }
        
        self.logger.info(f"\n📊 결과 요약:")
        self.logger.info(f"   총 테스트: {total_tests:,}개")
        self.logger.info(f"   매핑 성공: {successful_mappings:,}개 ({success_rate:.2f}%)")
        self.logger.info(f"   GT 일치: {gt_matched_count:,}개 ({gt_match_rate:.2f}%)")
        self.logger.info(f"   소요 시간: {elapsed_time:.2f}초 ({elapsed_time/60:.2f}분)")
        
        # 결과 저장
        self.save_results(summary)
        
        return summary
    
    def save_results(self, summary: dict):
        """테스트 결과를 파일로 저장"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 1. 요약 CSV 저장
        summary_data = [{
            'test_name': summary['test_name'],
            'description': summary['description'],
            'use_lexical': summary['use_lexical'],
            'scoring_mode': summary['scoring_mode'],
            'include_stage1_scores': summary['include_stage1_scores'],
            'include_non_std_info': summary['include_non_std_info'],
            'total_tests': summary['total_tests'],
            'successful_mappings': summary['successful_mappings'],
            'success_rate': summary['success_rate'],
            'gt_matched_count': summary['gt_matched_count'],
            'gt_match_rate': summary['gt_match_rate'],
            'elapsed_time': summary['elapsed_time'],
            'avg_time_per_entity': summary['avg_time_per_entity']
        }]
        
        summary_df = pd.DataFrame(summary_data)
        summary_csv = self.log_dir / f"fail_retry_mapping_summary_{timestamp}.csv"
        summary_df.to_csv(summary_csv, index=False, encoding='utf-8')
        self.logger.info(f"📄 요약 CSV 저장: {summary_csv}")
        
        # 2. 상세 결과 XLSX 저장
        xlsx_file = self.log_dir / f"fail_retry_mapping_detailed_{timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # 요약 시트
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, summary)
        
        # 상세 결과 시트
        ws_detail = wb.create_sheet(title="Details")
        self._create_detail_sheet(ws_detail, summary)
        
        wb.save(xlsx_file)
        self.logger.info(f"📊 상세 XLSX 저장: {xlsx_file}")
        
        # 3. JSON으로 전체 결과 저장
        json_file = self.log_dir / f"fail_retry_mapping_results_{timestamp}.json"
        json_data = {
            'timestamp': timestamp,
            'summary': {k: v for k, v in summary.items() if k != 'results'}
        }
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        self.logger.info(f"📁 JSON 저장: {json_file}")
    
    def _create_summary_sheet(self, ws, summary):
        """요약 시트 생성"""
        headers = [
            "항목", "값"
        ]
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터 작성
        summary_items = [
            ("테스트명", summary['test_name']),
            ("설명", summary['description']),
            ("use_lexical", str(summary['use_lexical'])),
            ("scoring_mode", summary['scoring_mode']),
            ("include_stage1_scores", str(summary['include_stage1_scores'])),
            ("include_non_std_info", str(summary['include_non_std_info'])),
            ("총 테스트", summary['total_tests']),
            ("매핑 성공", summary['successful_mappings']),
            ("성공률 (%)", round(summary['success_rate'], 2)),
            ("GT 일치 수", summary['gt_matched_count']),
            ("GT 일치율 (%)", round(summary['gt_match_rate'], 2)),
            ("소요시간 (초)", round(summary['elapsed_time'], 2)),
            ("평균시간 (초/엔티티)", round(summary['avg_time_per_entity'], 4))
        ]
        
        for row, (key, value) in enumerate(summary_items, 2):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 80
    
    def _create_detail_sheet(self, ws, summary):
        """상세 결과 시트 생성"""
        headers = [
            "Entity Name", "Domain", "Ground Truth ID", "GT Name", "Success", "GT Matched",
            "Best Concept ID", "Best Concept Name", "Vocabulary ID",
            "Score", "Confidence", "Mapping Method",
            "Validation Changed", "LLM Top Pick", "Error",
            "Stage1 Candidates", "Stage2 Candidates", "Stage3 Candidates"
        ]
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 데이터 작성
        for row, result in enumerate(summary['results'], 2):
            ws.cell(row=row, column=1, value=result.get('entity_name', 'N/A'))
            ws.cell(row=row, column=2, value=result.get('input_domain', 'N/A'))
            ws.cell(row=row, column=3, value=result.get('ground_truth_id', 'N/A'))
            ws.cell(row=row, column=4, value=result.get('ground_truth_name', 'N/A'))
            
            # 성공 여부 컬러링
            success_cell = ws.cell(row=row, column=5, value="성공" if result.get('success') else "실패")
            if result.get('success'):
                success_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                success_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # GT 일치 여부 컬러링
            gt_matched = result.get('gt_matched')
            if gt_matched is not None:
                gt_cell = ws.cell(row=row, column=6, value="일치" if gt_matched else "불일치")
                if gt_matched:
                    gt_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    gt_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                ws.cell(row=row, column=6, value="N/A")
            
            ws.cell(row=row, column=7, value=result.get('best_concept_id', 'N/A'))
            ws.cell(row=row, column=8, value=result.get('best_concept_name', 'N/A'))
            ws.cell(row=row, column=9, value=result.get('vocabulary_id', 'N/A'))
            ws.cell(row=row, column=10, value=result.get('best_score', 0.0))
            ws.cell(row=row, column=11, value=result.get('best_confidence', 'N/A'))
            ws.cell(row=row, column=12, value=result.get('mapping_method', 'N/A'))
            
            # Validation changed 여부 컬러링
            validation_changed = result.get('validation_changed', False)
            val_cell = ws.cell(row=row, column=13, value="변경됨" if validation_changed else "유지")
            if validation_changed:
                val_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            # LLM Top Pick 정보
            llm_top = result.get('llm_top_pick')
            if llm_top:
                llm_top_text = f"ID: {llm_top.get('concept_id', 'N/A')}\n이름: {llm_top.get('concept_name', 'N/A')}\nLLM점수: {llm_top.get('llm_score', 'N/A')}"
            else:
                llm_top_text = "N/A"
            ws.cell(row=row, column=14, value=llm_top_text)
            ws.cell(row=row, column=14).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Error 정보
            error_text = result.get('error', '')
            ws.cell(row=row, column=15, value=error_text if error_text else "")
            
            # Stage별 후보군 추가
            stage1_text = self._format_candidates_for_cell(result.get('stage1_candidates', []), 'stage1')
            ws.cell(row=row, column=16, value=stage1_text)
            
            stage2_text = self._format_candidates_for_cell(result.get('stage2_candidates', []), 'stage2')
            ws.cell(row=row, column=17, value=stage2_text)
            
            stage3_text = self._format_candidates_for_cell(result.get('stage3_candidates', []), 'stage3', 'llm')
            ws.cell(row=row, column=18, value=stage3_text)
            
            # Stage 컬럼 스타일 설정
            for col in range(16, 19):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 열 너비 설정
        column_widths = [45, 15, 18, 40, 10, 12, 18, 50, 15, 10, 12, 20, 12, 40, 30, 70, 70, 85]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 행 높이 설정 (Stage 후보군 표시를 위해)
        for row_num in range(2, len(summary['results']) + 2):
            ws.row_dimensions[row_num].height = 150
    
    def _format_candidates_for_cell(self, candidates, stage_type, scoring_mode='llm'):
        """후보군 포맷팅 (엑셀 셀용)"""
        if not candidates:
            return "후보 없음"
        
        lines = []
        max_candidates = 15 if stage_type in ['stage1', 'stage2'] else 10
        
        for i, candidate in enumerate(candidates[:max_candidates], 1):
            if stage_type == 'stage1':
                # Stage 1: Elasticsearch 결과
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   ES점수: {candidate.get('elasticsearch_score', 0):.4f}, "
                line += f"Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            
            elif stage_type == 'stage2':
                # Stage 2: Standard 변환 결과
                search_type = candidate.get('search_type', 'unknown')
                is_std = "✓" if candidate.get('is_original_standard', True) else "→"
                line = f"{i}. [{search_type}] {is_std} {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
                if not candidate.get('is_original_standard', True):
                    original_non_std = candidate.get('original_non_standard', {})
                    if original_non_std:
                        line += f"\n   원본 Non-std: {original_non_std.get('concept_name', 'N/A')} (ID: {original_non_std.get('concept_id', 'N/A')})"
            
            else:  # stage3
                # Stage 3: LLM 또는 Semantic Only 결과
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                
                # LLM 모드인 경우
                llm_score = candidate.get('llm_score')
                llm_rank = candidate.get('llm_rank')
                llm_reasoning = candidate.get('llm_reasoning')
                semantic_similarity = candidate.get('semantic_similarity')
                
                if llm_score is not None:
                    line += f"   LLM점수: {llm_score}, 순위: {llm_rank}"
                    if semantic_similarity is not None:
                        line += f" | 의미유사도: {semantic_similarity:.4f}"
                    line += "\n"
                    if llm_reasoning:
                        reasoning_short = llm_reasoning[:60] + '...' if len(llm_reasoning) > 60 else llm_reasoning
                        line += f"   이유: {reasoning_short}\n"
                else:
                    # Semantic Only 또는 Hybrid 모드인 경우
                    text_sim = candidate.get('text_similarity', 0)
                    sem_sim = candidate.get('semantic_similarity', 0)
                    final_score = candidate.get('final_score', 0)
                    line += f"   텍스트: {text_sim:.4f}, 의미적: {sem_sim:.4f}, 최종: {final_score:.4f}\n"
                
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            
            lines.append(line)
        
        return "\n\n".join(lines)


def main():
    """메인 함수"""
    # ============================================================
    # 설정
    # ============================================================
    EXCEL_PATH = "/home/work/skku/hyo/omop-mapper/data/fail-retry.xlsx"
    
    # ============================================================
    # 테스트 실행
    # ============================================================
    tester = FailRetryMappingTester()
    
    results = tester.run_mapping_test(
        excel_path=EXCEL_PATH
    )
    
    print(f"\n✅ 매핑 실패 데이터 재시도 테스트 완료!")
    print(f"   결과는 {tester.log_dir} 디렉토리에 저장되었습니다.")


if __name__ == "__main__":
    main()
