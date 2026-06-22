"""
Shared utilities for the mapping pipeline.

Common to all data sources (e.g. SNUH, SNOMED):
- Data source configuration (default paths, preprocessing)
- Logging setup
- JSON/LOG/XLSX output
"""

import json
import logging
from dataclasses import asdict, is_dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

# -----------------------------------------------------------------------------
# Data source registry (add a new entry here to support another dataset)
# -----------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent

DATA_SOURCES = {
    "snuh": {
        "csv_path": str(PROJECT_ROOT / "data" / "snuh-baseline-mapping-data.csv"),
        "vocabulary_filter": ["SNOMED", "LOINC"],  # default preprocessing
        "filter_domains": ["Procedure"],  # load only these domains (keep aligned with "domains")
        "domains": ["Procedure"],
        "id_col": "row_id",
        "loader": "load_snuh_data",
        "row_to_input": "snuh_row_to_input",
    },
    "snomed": {
        "csv_path": str(PROJECT_ROOT / "data" / "snomed-mapping-data-1000.csv"),
        "filter_domains": ["Condition", "Measurement", "Observation", "Procedure"],  # default preprocessing
        "domains": ["Condition", "Measurement", "Observation", "Procedure"],
        "id_col": "test_index",  # if note_id is missing, load_snomed_data generates row_id
        "loader": "load_snomed_data",
        "row_to_input": "snomed_row_to_input",
    },
}

# openpyxl for xlsx
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# XLSX column layout
XLSX_HEADERS = [
    "Test Index",
    "ID",
    "Entity Name",
    "Input Domain",
    "Ground Truth Concept ID",
    "Ground Truth Concept Name",
    "Success",
    "Mapping Correct",
    "Best Result Domain",
    "Best Concept ID",
    "Best Concept Name",
    "Best Score",
    "Stage1 Candidates",
    "Stage2 Candidates",
    "Stage3 Candidates",
]

# Base headers for the repeated-mapping summary sheet (--repeat option).
# "Mapped Concept 1..N" columns are added dynamically by save_xlsx_repeat.
SUMMARY_BASE_HEADERS = [
    "Test Index",
    "ID",
    "Entity Name",
    "Input Domain",
    "Ground Truth Concept ID",
    "Ground Truth Concept Name",
    "All Same",        # whether all N runs produced the same result
    "Correct",         # number of correct results across the N runs
]


def setup_logging(
    output_dir: Path, data_type: str, timestamp: str, console: bool = True
) -> tuple[logging.Logger, Path]:
    """Configure logging with a shared timestamp.
    console=False: disable terminal output (file only); used in parallel runs where
    only the progress bar is shown in the terminal.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    log_file = output_dir / f"mapping_{data_type}_{timestamp}.log"

    logger = logging.getLogger(f"mapping_{data_type}")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(logging.INFO)

    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    file_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    if console:
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)

    # MapOMOP API loggers: always log to file; console depends on the console flag
    # (in parallel runs the detailed logs go to the file only).
    for name in [
        "MapOMOP.entity_mapping_api",
        "MapOMOP.mapping_stages.stage1_candidate_retrieval",
        "MapOMOP.mapping_stages.stage2_standard_collection",
        "MapOMOP.mapping_stages.stage3_hybrid_scoring",
        "MapOMOP.mapping_validation",
    ]:
        api_log = logging.getLogger(name)
        api_log.setLevel(logging.INFO)
        api_log.handlers.clear()
        api_log.addHandler(file_handler)
        if console:
            api_log.addHandler(console_handler)

    logger.info(f"Log file: {log_file}")
    return logger, log_file


def setup_worker_logging(log_file_path: str, capture_only: bool = False) -> None:
    """Configure logging for a worker process.

    capture_only=True: do not write logs to file/console; _map_single_task captures
    and returns them instead.
    capture_only=False: write to file and console in real time (not used when workers=1).
    """
    _API_LOGGER_NAMES = [
        "MapOMOP.entity_mapping_api",
        "MapOMOP.mapping_stages.stage1_candidate_retrieval",
        "MapOMOP.mapping_stages.stage2_standard_collection",
        "MapOMOP.mapping_stages.stage3_hybrid_scoring",
        "MapOMOP.mapping_validation",
    ]
    if capture_only:
        for name in _API_LOGGER_NAMES:
            api_log = logging.getLogger(name)
            api_log.setLevel(logging.INFO)
            api_log.handlers.clear()
            api_log.propagate = False
        return

    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    file_handler = logging.FileHandler(log_file_path, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    for name in _API_LOGGER_NAMES:
        api_log = logging.getLogger(name)
        api_log.setLevel(logging.INFO)
        api_log.handlers.clear()
        api_log.addHandler(file_handler)
        api_log.addHandler(console_handler)


class LogCaptureHandler(logging.Handler):
    """Collect log records into a list as formatted strings."""

    def __init__(self, log_list: list):
        super().__init__()
        self.log_list = log_list

    def emit(self, record: logging.LogRecord) -> None:
        try:
            msg = self.formatter.format(record) if self.formatter else record.getMessage()
            self.log_list.append(msg)
        except Exception:
            self.handleError(record)


API_LOGGER_NAMES = [
    "MapOMOP.entity_mapping_api",
    "MapOMOP.mapping_stages.stage1_candidate_retrieval",
    "MapOMOP.mapping_stages.stage2_standard_collection",
    "MapOMOP.mapping_stages.stage3_hybrid_scoring",
    "MapOMOP.mapping_validation",
]


def capture_entity_logs(logger_names: Optional[list[str]] = None, formatter: Optional[logging.Formatter] = None) -> tuple[list, list]:
    """Attach a LogCaptureHandler to the given loggers and return (handlers_added, log_list).
    The caller removes the handlers afterward and includes log_list in its return value.
    """
    if logger_names is None:
        logger_names = API_LOGGER_NAMES
    if formatter is None:
        formatter = logging.Formatter(
            "%(asctime)s - %(name)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    log_list: list[str] = []
    handler = LogCaptureHandler(log_list)
    handler.setLevel(logging.INFO)
    handler.setFormatter(formatter)

    handlers_added = []
    for name in logger_names:
        logger = logging.getLogger(name)
        logger.addHandler(handler)
        handlers_added.append((logger, handler))

    return handlers_added, log_list


def save_json(results: List[Dict], output_dir: Path, data_type: str, timestamp: str) -> Path:
    """Save raw mapping results to JSON."""
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"mapping_{data_type}_{timestamp}.json"

    def _serialize(obj: Any) -> Any:
        if is_dataclass(obj) and not isinstance(obj, type):
            return {k: _serialize(v) for k, v in asdict(obj).items()}
        if isinstance(obj, dict):
            return {k: _serialize(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_serialize(x) for x in obj]
        if isinstance(obj, (str, bool, type(None))):
            return obj
        if isinstance(obj, (int, float)):
            return obj
        # numpy scalars and similar
        try:
            return float(obj)
        except (TypeError, ValueError):
            pass
        try:
            return int(obj)
        except (TypeError, ValueError):
            pass
        return str(obj)

    serializable = _serialize(results)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(serializable, f, ensure_ascii=False, indent=2)
    return out_path


def _sort_stage1_by_score(candidates: List[Dict]) -> List[Dict]:
    """Sort stage1 candidates by descending elasticsearch_score (regardless of lexical/semantic/combined)."""
    if not candidates:
        return []
    return sorted(
        candidates,
        key=lambda x: float(x.get("elasticsearch_score") or x.get("_score") or 0),
        reverse=True,
    )


def _sort_stage2_by_score(candidates: List[Dict]) -> List[Dict]:
    """Sort stage2 candidates by elasticsearch_score."""
    if not candidates:
        return []
    return sorted(
        candidates,
        key=lambda x: float(x.get("elasticsearch_score") or 0),
        reverse=True,
    )


def _sort_stage3_by_score(candidates: List[Dict]) -> List[Dict]:
    """Sort stage3 candidates by final_score."""
    if not candidates:
        return []
    return sorted(
        candidates,
        key=lambda x: float(x.get("final_score") or 0),
        reverse=True,
    )


def _format_candidates_for_cell(candidates: List[Dict], stage_type: str) -> str:
    """Format the candidate list as text for an XLSX cell, ordered by descending score."""
    if not candidates:
        return "No candidates"

    if stage_type == "stage1":
        sorted_candidates = _sort_stage1_by_score(candidates)
    elif stage_type == "stage2":
        sorted_candidates = _sort_stage2_by_score(candidates)
    else:
        sorted_candidates = _sort_stage3_by_score(candidates)

    lines = []
    max_show = 15 if stage_type in ("stage1", "stage2") else 10

    for i, c in enumerate(sorted_candidates[:max_show], 1):
        st = c.get("search_type", "unknown")
        name = c.get("concept_name", "N/A")
        cid = c.get("concept_id", "N/A")

        if stage_type == "stage1":
            es_score = float(c.get("elasticsearch_score") or c.get("_score") or 0)
            line = f"{i}. [{st}] {name} (ID: {cid})\n"
            line += f"   ES score: {es_score:.4f}, Standard: {c.get('standard_concept', 'N/A')}, Domain: {c.get('domain_id', 'N/A')}"
        elif stage_type == "stage2":
            is_std = "✓" if c.get("is_original_standard", True) else "→"
            line = f"{i}. [{st}] {is_std} {name} (ID: {cid})\n"
            line += f"   Standard: {c.get('standard_concept', 'N/A')}, Domain: {c.get('domain_id', 'N/A')}"
            if not c.get("is_original_standard", True):
                ons = c.get("original_non_standard", {})
                if ons:
                    line += f"\n   Original Non-std: {ons.get('concept_name', 'N/A')} (ID: {ons.get('concept_id', 'N/A')})"
        else:
            fin = float(c.get("final_score") or 0)
            sem = c.get("semantic_similarity")
            line = f"{i}. [{st}] {name} (ID: {cid})\n"
            line += f"   Final: {fin:.1f}"
            if sem is not None:
                line += f", Semantic: {sem:.4f}"
            line += f", Standard: {c.get('standard_concept', 'N/A')}, Domain: {c.get('domain_id', 'N/A')}"

        lines.append(line)

    return "\n\n".join(lines)


def save_xlsx(results: List[Dict], output_dir: Path, data_type: str, timestamp: str) -> Path:
    """Save mapping results to XLSX. Columns: Test Index, ID, Entity Name, Input Domain, ..."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for XLSX output. pip install openpyxl")

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"mapping_{data_type}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed Results"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(XLSX_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r.get("test_index", ""))
        ws.cell(row=row_idx, column=2, value=r.get("id", r.get("snuh_id", r.get("note_id", "N/A"))))
        ws.cell(row=row_idx, column=3, value=r.get("entity_name", ""))
        ws.cell(row=row_idx, column=4, value=r.get("input_domain", "All"))
        ws.cell(row=row_idx, column=5, value=r.get("ground_truth_concept_id", ""))
        ws.cell(row=row_idx, column=6, value=r.get("ground_truth_concept_name", ""))
        ws.cell(row=row_idx, column=7, value="Success" if r.get("success") else "Fail")

        correct_cell = ws.cell(row=row_idx, column=8, value="Correct" if r.get("mapping_correct") else "Incorrect")
        if r.get("mapping_correct"):
            correct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            correct_cell.font = Font(color="006100")
        else:
            correct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            correct_cell.font = Font(color="9C0006")

        ws.cell(row=row_idx, column=9, value=r.get("best_result_domain", "N/A"))
        ws.cell(row=row_idx, column=10, value=r.get("best_concept_id", "N/A"))
        ws.cell(row=row_idx, column=11, value=r.get("best_concept_name", "N/A"))
        ws.cell(row=row_idx, column=12, value=r.get("best_score", 0.0))

        # Stage1: ordered by descending score (lexical/semantic/combined alike)
        stage1_text = _format_candidates_for_cell(r.get("stage1_candidates", []), "stage1")
        ws.cell(row=row_idx, column=13, value=stage1_text)

        stage2_text = _format_candidates_for_cell(r.get("stage2_candidates", []), "stage2")
        ws.cell(row=row_idx, column=14, value=stage2_text)

        stage3_text = _format_candidates_for_cell(r.get("stage3_candidates", []), "stage3")
        ws.cell(row=row_idx, column=15, value=stage3_text)

        for c in range(13, 16):
            ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    widths = {
        "A": 10, "B": 18, "C": 40, "D": 15, "E": 20, "F": 45, "G": 10, "H": 12,
        "I": 18, "J": 15, "K": 45, "L": 12, "M": 70, "N": 70, "O": 85,
    }
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    for rn in range(2, len(results) + 2):
        ws.row_dimensions[rn].height = 150

    wb.save(out_path)
    return out_path


def save_xlsx_repeat(
    all_results: List[List[Dict]],
    output_dir: Path,
    data_type: str,
    timestamp: str,
) -> Path:
    """Save repeated-mapping results: a summary sheet plus one detail sheet per run."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for XLSX output. pip install openpyxl")

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"mapping_{data_type}_{timestamp}.xlsx"

    num_runs = len(all_results)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()

    # Sheet 1: Summary (Mapped Concept 1..N created dynamically per num_runs)
    summary_headers = SUMMARY_BASE_HEADERS + [
        f"Mapped Concept {i}" for i in range(1, num_runs + 1)
    ]
    ws_summary = wb.active
    ws_summary.title = "Summary"
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Aggregate the N runs per test_index
    by_index = {}  # test_index -> [r1, r2, ..., rN]
    for run_results in all_results:
        for r in run_results:
            idx = r.get("test_index")
            if idx not in by_index:
                by_index[idx] = []
            by_index[idx].append(r)

    for row_idx, test_index in enumerate(sorted(by_index.keys()), 2):
        rows = by_index[test_index]
        r0 = rows[0]
        ws_summary.cell(row=row_idx, column=1, value=test_index)
        ws_summary.cell(row=row_idx, column=2, value=r0.get("id", "N/A"))
        ws_summary.cell(row=row_idx, column=3, value=r0.get("entity_name", ""))
        ws_summary.cell(row=row_idx, column=4, value=r0.get("input_domain", "All"))
        ws_summary.cell(row=row_idx, column=5, value=r0.get("ground_truth_concept_id", ""))
        ws_summary.cell(row=row_idx, column=6, value=r0.get("ground_truth_concept_name", ""))

        # All Same: whether best_concept_id is identical across all runs
        concept_ids = [
            str(r.get("best_concept_id") or "") for r in rows
        ]
        all_same = len(set(concept_ids)) == 1 if concept_ids else False
        ws_summary.cell(row=row_idx, column=7, value="Y" if all_same else "N")

        # Correct: number of correct results across runs
        correct_count = sum(1 for r in rows if r.get("mapping_correct"))
        ws_summary.cell(row=row_idx, column=8, value=correct_count)

        # Mapped Concept 1..N: concept_name(concept_id) for every run
        for i, r in enumerate(rows[:num_runs]):
            cid = r.get("best_concept_id") or "N/A"
            cname = r.get("best_concept_name") or "N/A"
            ws_summary.cell(row=row_idx, column=9 + i, value=f"{cname}({cid})")

    # Column widths (summary) - Mapped Concept columns are set per num_runs
    for col_letter, w in {"A": 10, "B": 18, "C": 40, "D": 15, "E": 20, "F": 45, "G": 10, "H": 8}.items():
        ws_summary.column_dimensions[col_letter].width = w
    for i in range(num_runs):
        ws_summary.column_dimensions[openpyxl.utils.get_column_letter(9 + i)].width = 50

    # Sheets 2..(1+num_runs): per-run detail (same format as save_xlsx)
    for run_idx, run_results in enumerate(all_results, 1):
        ws = wb.create_sheet(title=f"Mapping {run_idx}", index=run_idx)

        for col, h in enumerate(XLSX_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, r in enumerate(run_results, 2):
            ws.cell(row=row_idx, column=1, value=r.get("test_index", ""))
            ws.cell(row=row_idx, column=2, value=r.get("id", r.get("snuh_id", r.get("note_id", "N/A"))))
            ws.cell(row=row_idx, column=3, value=r.get("entity_name", ""))
            ws.cell(row=row_idx, column=4, value=r.get("input_domain", "All"))
            ws.cell(row=row_idx, column=5, value=r.get("ground_truth_concept_id", ""))
            ws.cell(row=row_idx, column=6, value=r.get("ground_truth_concept_name", ""))
            ws.cell(row=row_idx, column=7, value="Success" if r.get("success") else "Fail")

            correct_cell = ws.cell(row=row_idx, column=8, value="Correct" if r.get("mapping_correct") else "Incorrect")
            if r.get("mapping_correct"):
                correct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                correct_cell.font = Font(color="006100")
            else:
                correct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                correct_cell.font = Font(color="9C0006")

            ws.cell(row=row_idx, column=9, value=r.get("best_result_domain", "N/A"))
            ws.cell(row=row_idx, column=10, value=r.get("best_concept_id", "N/A"))
            ws.cell(row=row_idx, column=11, value=r.get("best_concept_name", "N/A"))
            ws.cell(row=row_idx, column=12, value=r.get("best_score", 0.0))

            stage1_text = _format_candidates_for_cell(r.get("stage1_candidates", []), "stage1")
            ws.cell(row=row_idx, column=13, value=stage1_text)
            stage2_text = _format_candidates_for_cell(r.get("stage2_candidates", []), "stage2")
            ws.cell(row=row_idx, column=14, value=stage2_text)
            stage3_text = _format_candidates_for_cell(r.get("stage3_candidates", []), "stage3")
            ws.cell(row=row_idx, column=15, value=stage3_text)

            for c in range(13, 16):
                ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")

        widths = {
            "A": 10, "B": 18, "C": 40, "D": 15, "E": 20, "F": 45, "G": 10, "H": 12,
            "I": 18, "J": 15, "K": 45, "L": 12, "M": 70, "N": 70, "O": 85,
        }
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w
        for rn in range(2, len(run_results) + 2):
            ws.row_dimensions[rn].height = 150

    wb.save(out_path)
    return out_path


# --- SNUH data loader ---
def load_snuh_data(
    csv_path: str,
    sample_size: Optional[int] = None,
    use_random: bool = False,
    random_state: int = 42,
    sample_per_domain: Optional[Dict[str, int]] = None,
    vocabulary_filter: Optional[List[str]] = None,
    filter_domains: Optional[List[str]] = None,
    chunk_size: int = 100000,
) -> pd.DataFrame:
    """Load the SNUH CSV. entity=source_name, domain=domain, gt=omop_concept_id, id=snuh_id.
    Default preprocessing: vocabulary IN ('SNOMED', 'LOINC')
    """
    from tqdm import tqdm

    if vocabulary_filter is None:
        vocabulary_filter = ["SNOMED", "LOINC"]

    if sample_per_domain or vocabulary_filter:
        chunks = []
        for chunk in tqdm(
            pd.read_csv(csv_path, chunksize=chunk_size),
            desc="Loading data",
        ):
            chunks.append(chunk)
        df = pd.concat(chunks, ignore_index=True)

        if vocabulary_filter and "vocabulary" in df.columns:
            df = df[df["vocabulary"].isin(vocabulary_filter)]

        if filter_domains:
            domain_col = "domain_id" if "domain_id" in df.columns else "domain"
            if domain_col in df.columns:
                df = df[df[domain_col].isin(filter_domains)]

        if sample_per_domain:
            domain_col = "domain_id" if "domain_id" in df.columns else "domain"
            sampled = []
            for domain, size in sample_per_domain.items():
                subset = df[df[domain_col] == domain]
                n = min(size, len(subset))
                if n == 0:
                    continue
                if use_random:
                    sampled.append(subset.sample(n=n, random_state=random_state))
                else:
                    sampled.append(subset.head(n))
            df = pd.concat(sampled, ignore_index=True)
            if use_random:
                df = df.sample(frac=1, random_state=random_state).reset_index(drop=True)
        else:
            n = len(df) if sample_size is None else min(sample_size, len(df))
            if use_random:
                df = df.sample(n=n, random_state=random_state)
            else:
                df = df.head(n)
    else:
        if sample_size is None:
            df = pd.read_csv(csv_path)
        else:
            df = pd.read_csv(csv_path, nrows=sample_size)

    return df.reset_index(drop=True)


def snuh_row_to_input(row: pd.Series, domain_map: Dict[str, Any]) -> tuple[str, Optional[Any], str, Optional[int], Optional[str]]:
    """(entity_name, domain_id, record_id, ground_truth, ground_truth_concept_name)"""
    entity = str(row.get("source_name", row.get("source_value", ""))).strip()
    domain_str = str(row.get("domain", row.get("domain_id", ""))).strip()
    domain_str = None if not domain_str or domain_str == "nan" else domain_str
    domain_id = domain_map.get(domain_str) if domain_str else None
    rid = str(row.get("snuh_id", row.get("row_id", row.name if row.name is not None else "N/A")))
    rid = "N/A" if not rid or rid == "nan" else rid
    gt_raw = row.get("omop_concept_id", row.get("concept_id"))
    gt = int(gt_raw) if pd.notna(gt_raw) else None
    gt_name = str(row.get("concept_name", "")).strip() if pd.notna(row.get("concept_name")) else None
    return entity, domain_id, rid, gt, gt_name


# --- SNOMED data loader ---
def load_snomed_data(
    csv_path: str,
    sample_size: Optional[int] = None,
    use_random: bool = False,
    random_state: int = 42,
    sample_per_domain: Optional[Dict[str, int]] = None,
    filter_domains: Optional[List[str]] = None,
    chunk_size: int = 100000,
) -> pd.DataFrame:
    """Load the SNOMED CSV. entity=entity_name, domain=domain_id, gt=concept_id, id=note_id.
    Default preprocessing: domain_id IN ('Condition','Measurement','Drug','Observation','Procedure')
    """
    from tqdm import tqdm

    if filter_domains is None:
        filter_domains = ["Condition", "Measurement", "Drug", "Observation", "Procedure"]

    chunks = []
    for chunk in tqdm(pd.read_csv(csv_path, chunksize=chunk_size), desc="Loading data"):
        chunks.append(chunk)
    df = pd.concat(chunks, ignore_index=True)

    if filter_domains and "domain_id" in df.columns:
        df = df[df["domain_id"].isin(filter_domains)]

    if sample_per_domain:
        sampled = []
        for domain, size in sample_per_domain.items():
            subset = df[df["domain_id"] == domain]
            n = min(size, len(subset))
            if n == 0:
                continue
            if use_random:
                sampled.append(subset.sample(n=n, random_state=random_state))
            else:
                sampled.append(subset.head(n))
        df = pd.concat(sampled, ignore_index=True)
        if use_random:
            df = df.sample(frac=1, random_state=random_state).reset_index(drop=True)
    else:
        n = len(df) if sample_size is None else min(sample_size, len(df))
        if use_random:
            df = df.sample(n=n, random_state=random_state)
        else:
            df = df.head(n)

    df = df.reset_index(drop=True)
    # If note_id is missing, generate row_id from the row index
    if "note_id" not in df.columns:
        df["row_id"] = df.index.astype(str)
    elif "row_id" not in df.columns:
        df["row_id"] = df["note_id"].astype(str)
    return df


def snomed_row_to_input(row: pd.Series, domain_map: Dict[str, Any]) -> tuple[str, Optional[Any], str, Optional[int], Optional[str]]:
    """(entity_name, domain_id, record_id, ground_truth, ground_truth_concept_name)
    record_id: prefers row_id (auto-generated) or note_id
    """
    entity = str(row["entity_name"]).strip()
    domain_str = str(row["domain_id"]).strip() if pd.notna(row.get("domain_id")) else None
    domain_id = domain_map.get(domain_str) if domain_str else None
    rid = str(row.get("row_id", row.get("note_id", row.name if row.name is not None else "N/A")))
    gt = int(row["concept_id"]) if pd.notna(row.get("concept_id")) else None
    gt_name = str(row.get("concept_name", "")).strip() if pd.notna(row.get("concept_name")) else None
    return entity, domain_id, rid, gt, gt_name
