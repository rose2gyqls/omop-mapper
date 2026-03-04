#!/usr/bin/env python3
"""
매핑 결과를 Excel로 변환. JSON 또는 로그 파일 입력 지원.

- JSON 입력: mapping_common.save_xlsx/save_xlsx_repeat와 동일한 형식으로 저장
- 로그 입력: 로그 파싱 후 Excel 변환 (CSV 병합으로 ground_truth 보완)

중간 결과(JSON)를 실시간으로 Excel로 확인 가능.

Usage:
    # JSON → Excel (기존 저장 형식과 동일)
    python scripts/log_to_xlsx.py test_logs/mapping_snuh_20260304_123456.json
    python scripts/log_to_xlsx.py test_logs/mapping_snuh_20260304_123456.json --watch  # 10초마다 갱신

    # 로그 → Excel
    python scripts/log_to_xlsx.py test_logs/mapping_snuh_20260304_123456.log
    python scripts/log_to_xlsx.py test_logs/mapping_snuh_20260304_123456.log --csv data/snuh-baseline-mapping-data.csv
"""

import argparse
import json
import re
import sys
from pathlib import Path

_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_root))

import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from mapping_common import (
    DATA_SOURCES,
    XLSX_HEADERS,
    SUMMARY_HEADERS,
    save_xlsx,
    save_xlsx_repeat,
)


# -----------------------------------------------------------------------------
# 로그 파싱 정규식
# -----------------------------------------------------------------------------
RE_START = re.compile(
    r"Starting single-domain mapping: (.+)$",
    re.MULTILINE,
)
RE_TARGET_DOMAIN = re.compile(
    r"Target domain: (\w+)$",
    re.MULTILINE,
)
RE_STAGE3_SCORE = re.compile(
    r"^\s*(.+?)\s+\((\d+)\)\s+([\d.]+)\s*$",
)
RE_STATUS = re.compile(
    r"#(\d+)\s+(.+?):\s+(정답|오답|실패)$",
)
RE_RUN_BOUNDARY = re.compile(
    r"매핑 Run (\d+)/(\d+)",
)


def _extract_log_message(line: str) -> str | None:
    """로그 라인에서 메시지 부분 추출."""
    if " - INFO - " in line:
        return line.split(" - INFO - ", 1)[1].strip()
    return None


def infer_data_type_and_timestamp(path: Path) -> tuple[str, str]:
    """파일명에서 data_type, timestamp 추출. mapping_snuh_20260304_123456 -> (snuh, 20260304_123456)"""
    stem = path.stem
    if "manual" in stem:
        data_type = "manual_withval" if "withval" in stem else "manual"
    elif "snuh" in stem:
        data_type = "snuh_withval" if "withval" in stem else "snuh"
    elif "snomed" in stem:
        data_type = "snomed_withval" if "withval" in stem else "snomed"
    else:
        data_type = "snuh"
    parts = stem.split("_")
    timestamp = "_".join(parts[-2:]) if len(parts) >= 2 else "unknown"
    return data_type, timestamp


# -----------------------------------------------------------------------------
# JSON → Excel
# -----------------------------------------------------------------------------
def load_json_and_convert(input_path: Path, output_dir: Path, data_type: str, timestamp: str) -> Path | None:
    """
    JSON 파일을 읽어 mapping_common과 동일 형식의 Excel로 저장.
    - 단일 run: save_xlsx
    - 다중 run (num_runs, runs): save_xlsx_repeat
    """
    if not input_path.exists():
        return None
    try:
        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        print(f"  JSON 로드 실패: {e}")
        return None

    if isinstance(data, dict) and "num_runs" in data and "runs" in data:
        all_results = data["runs"]
        num_runs = data["num_runs"]
        if not all_results:
            print("  runs가 비어 있음. Excel 저장 건너뜀.")
            return None
        out_path = save_xlsx_repeat(all_results, output_dir, data_type, timestamp)
        print(f"  JSON → Excel: {num_runs}회 Run, {sum(len(r) for r in all_results)}건")
    elif isinstance(data, list):
        if not data:
            print("  JSON 리스트가 비어 있음. Excel 저장 건너뜀.")
            return None
        out_path = save_xlsx(data, output_dir, data_type, timestamp)
        print(f"  JSON → Excel: 단일 Run, {len(data)}건")
    else:
        print("  지원하지 않는 JSON 형식 (dict with num_runs/runs 또는 list)")
        return None
    return out_path


# -----------------------------------------------------------------------------
# 로그 파싱
# -----------------------------------------------------------------------------
def parse_log_blocks(log_path: Path) -> tuple[list[dict], list[tuple[int, int, int]], int, set[int]]:
    """
    로그 파일을 파싱하여 엔티티별 블록 리스트 반환.
    Returns: (blocks, run_boundaries, num_runs, completed_run_nums)
    """
    text = log_path.read_text(encoding="utf-8")
    lines = text.split("\n")

    run_boundaries: list[tuple[int, int, int]] = []
    for i, line in enumerate(lines):
        msg = _extract_log_message(line)
        if msg:
            m = RE_RUN_BOUNDARY.search(msg)
            if m:
                run_num, total = int(m.group(1)), int(m.group(2))
                run_boundaries.append((i, run_num, total))
    num_runs = run_boundaries[-1][2] if run_boundaries else 1
    if len(run_boundaries) > 1:
        completed_run_nums = {rn for (_, rn, _) in run_boundaries[:-1]}
    elif run_boundaries:
        completed_run_nums = {run_boundaries[0][1]}
    else:
        completed_run_nums = {1}

    block_starts: list[tuple[int, str]] = []
    for i, line in enumerate(lines):
        msg = _extract_log_message(line)
        if not msg:
            continue
        m = RE_START.search(msg)
        if m:
            entity_name = m.group(1).strip()
            entity_key = " ".join(entity_name.split())
            block_starts.append((i, entity_key))

    def _run_at_line(line_idx: int) -> int:
        run = 1
        for (bound_idx, rn, _) in run_boundaries:
            if line_idx >= bound_idx:
                run = rn
        return run

    status_by_entity: dict[str, list[tuple[int, str, bool]]] = {}
    for line in lines:
        msg = _extract_log_message(line)
        if msg:
            m = RE_STATUS.search(msg)
            if m:
                test_index, entity_display, mapping_correct = int(m.group(1)), m.group(2).strip(), m.group(3) == "정답"
                entity_key = " ".join(entity_display.split())
                if entity_key not in status_by_entity:
                    status_by_entity[entity_key] = []
                status_by_entity[entity_key].append((test_index, entity_display, mapping_correct))

    blocks: list[dict] = []
    for block_idx, (start_idx, entity_key) in enumerate(block_starts):
        end_idx = block_starts[block_idx + 1][0] if block_idx + 1 < len(block_starts) else len(lines)
        block_lines = [lines[i] for i in range(start_idx, end_idx)]

        entity_name = entity_key
        domain = "All"
        stage1_raw: list[str] = []
        stage2_raw: list[str] = []
        stage3_raw: list[str] = []
        best_domain = None
        best_concept_id = None
        best_concept_name = None
        best_score = 0.0
        in_stage1 = False
        in_stage2 = False
        in_stage3 = False

        for line in block_lines:
            msg = _extract_log_message(line)
            if not msg:
                continue
            if "Target domain:" in msg:
                dm = RE_TARGET_DOMAIN.search(msg)
                if dm:
                    domain = dm.group(1)
            if "Stage 1: Candidate Retrieval" in msg:
                in_stage1, in_stage2, in_stage3 = True, False, False
                continue
            if "Stage 2: Standard Concept Collection" in msg:
                in_stage1, in_stage2, in_stage3 = False, True, False
                continue
            if "Stage 3: Scoring Results" in msg:
                in_stage1, in_stage2, in_stage3 = False, False, True
                continue
            if in_stage1:
                stage1_raw.append(msg)
            elif in_stage2:
                stage2_raw.append(msg)
            elif in_stage3:
                stage3_raw.append(msg)

        stage3_scores: list[tuple[str, int, float]] = []
        for msg in stage3_raw:
            if msg.strip().startswith("→"):
                continue
            m = RE_STAGE3_SCORE.search(msg)
            if m:
                cname, cid, score = m.groups()
                stage3_scores.append((cname.strip(), int(cid), float(score)))
        if stage3_scores:
            best_c = max(stage3_scores, key=lambda x: x[2])
            best_concept_name, best_concept_id, best_score = best_c

        if entity_key in status_by_entity and status_by_entity[entity_key]:
            test_index, entity_display, mapping_correct = status_by_entity[entity_key].pop(0)
        else:
            test_index = len(blocks) + 1
            entity_display = entity_name
            mapping_correct = False

        run_idx = _run_at_line(start_idx)
        success = best_concept_id is not None
        blocks.append({
            "test_index": test_index,
            "entity_name": entity_display,
            "entity_key": entity_key,
            "input_domain": domain,
            "ground_truth_concept_id": None,
            "ground_truth_concept_name": None,
            "id": None,
            "success": success,
            "mapping_correct": mapping_correct,
            "best_result_domain": best_domain or domain,
            "best_concept_id": best_concept_id,
            "best_concept_name": best_concept_name,
            "best_score": best_score,
            "stage1_raw": "\n".join(stage1_raw),
            "stage2_raw": "\n".join(stage2_raw),
            "stage3_raw": "\n".join(stage3_raw),
            "run_idx": run_idx,
        })

    return blocks, run_boundaries, num_runs, completed_run_nums


def merge_with_csv(
    blocks: list[dict],
    csv_path: Path,
    entity_col: str = "source_value",
    id_col: str = "index",
    domain_col: str = "domain_id",
    gt_id_col: str = "concept_id",
    gt_name_col: str = "concept_name",
) -> list[dict]:
    """CSV와 병합하여 ground_truth, id 보완."""
    df = pd.read_csv(csv_path)
    if entity_col not in df.columns and "source_name" in df.columns:
        entity_col = "source_name"

    csv_by_entity: dict[str, dict] = {}
    for _, row in df.iterrows():
        entity_val = str(row.get(entity_col, "")).strip()
        entity_key = " ".join(entity_val.split())
        idx = row.get(id_col, row.get("index", row.name))
        if pd.isna(idx):
            idx = len(csv_by_entity) + 1
        test_idx = int(idx) if pd.notna(idx) else None
        csv_by_entity[entity_key] = {
            "id": str(int(idx)) if pd.notna(idx) else "N/A",
            "test_index": test_idx,
            "ground_truth_concept_id": int(row[gt_id_col]) if pd.notna(row.get(gt_id_col)) else None,
            "ground_truth_concept_name": str(row.get(gt_name_col, "")).strip() if pd.notna(row.get(gt_name_col)) else None,
            "domain": str(row.get(domain_col, "")).strip() if pd.notna(row.get(domain_col)) else None,
        }

    for b in blocks:
        ek = b.get("entity_key", b["entity_name"])
        ek_norm = " ".join(ek.split())
        if ek_norm in csv_by_entity:
            c = csv_by_entity[ek_norm]
            b["id"] = c["id"]
            b["ground_truth_concept_id"] = c["ground_truth_concept_id"]
            b["ground_truth_concept_name"] = c["ground_truth_concept_name"]
            if c.get("domain") and b.get("input_domain") == "All":
                b["input_domain"] = c["domain"]
            if c.get("test_index") is not None:
                b["test_index"] = c["test_index"]
        else:
            for _, row in df.iterrows():
                if int(row.get(id_col, row.get("index", 0))) == b["test_index"]:
                    b["id"] = str(int(row.get(id_col, row.get("index", 0))))
                    b["ground_truth_concept_id"] = int(row[gt_id_col]) if pd.notna(row.get(gt_id_col)) else None
                    b["ground_truth_concept_name"] = str(row.get(gt_name_col, "")).strip() if pd.notna(row.get(gt_name_col)) else None
                    break
        if b.get("id") is None:
            b["id"] = str(b["test_index"])
        gt_id = b.get("ground_truth_concept_id")
        best_id = b.get("best_concept_id")
        mapping_correct = False
        if gt_id is not None and best_id is not None:
            try:
                mapping_correct = int(gt_id) == int(best_id)
            except (ValueError, TypeError):
                pass
        b["mapping_correct"] = mapping_correct

    return blocks


def save_xlsx_from_log(
    all_results: list[list[dict]],
    output_dir: Path,
    data_type: str,
    timestamp: str,
) -> Path:
    """로그 파싱 결과를 Excel로 저장. Stage 열에는 로그 원문 사용."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required. pip install openpyxl")

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"mapping_{data_type}_{timestamp}_from_log.xlsx"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()
    num_runs = len(all_results)
    summary_headers = SUMMARY_HEADERS[: 8 + min(5, num_runs)]
    ws_summary = wb.active
    ws_summary.title = "현황"
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    by_index: dict[int, list[dict]] = {}
    for run_results in all_results:
        for r in run_results:
            idx = r.get("test_index")
            if idx not in by_index:
                by_index[idx] = []
            by_index[idx].append(r)

    for row_idx, test_index in enumerate(sorted(by_index.keys()), 2):
        rows = by_index[test_index]
        r0 = rows[0]
        ws_summary.cell(row=row_idx, column=1, value=test_index)
        ws_summary.cell(row=row_idx, column=2, value=r0.get("id", "N/A"))
        ws_summary.cell(row=row_idx, column=3, value=r0.get("entity_name", ""))
        ws_summary.cell(row=row_idx, column=4, value=r0.get("input_domain", "All"))
        ws_summary.cell(row=row_idx, column=5, value=r0.get("ground_truth_concept_id", ""))
        ws_summary.cell(row=row_idx, column=6, value=r0.get("ground_truth_concept_name", ""))

        concept_ids = [str(r.get("best_concept_id") or "") for r in rows]
        all_same = len(set(concept_ids)) == 1 if concept_ids else False
        ws_summary.cell(row=row_idx, column=7, value="Y" if all_same else "N")
        correct_count = sum(1 for r in rows if r.get("mapping_correct"))
        correct_display = "Y" if (len(rows) == 1 and correct_count == 1) else ("N" if (len(rows) == 1 and correct_count == 0) else correct_count)
        ws_summary.cell(row=row_idx, column=8, value=correct_display)
        for i, r in enumerate(rows[:5]):
            cid = r.get("best_concept_id") or "N/A"
            cname = r.get("best_concept_name") or "N/A"
            ws_summary.cell(row=row_idx, column=9 + i, value=f"{cname}({cid})")

    for col_letter, w in {"A": 10, "B": 18, "C": 40, "D": 15, "E": 20, "F": 45, "G": 10, "H": 8}.items():
        ws_summary.column_dimensions[col_letter].width = w
    for i in range(5):
        ws_summary.column_dimensions[openpyxl.utils.get_column_letter(9 + i)].width = 50

    for run_idx, run_results in enumerate(all_results, 1):
        ws = wb.create_sheet(title=f"{run_idx}번째 매핑", index=run_idx)
        for col, h in enumerate(XLSX_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, r in enumerate(run_results, 2):
            ws.cell(row=row_idx, column=1, value=r.get("test_index", ""))
            ws.cell(row=row_idx, column=2, value=r.get("id", r.get("snuh_id", r.get("note_id", "N/A"))))
            ws.cell(row=row_idx, column=3, value=r.get("entity_name", ""))
            ws.cell(row=row_idx, column=4, value=r.get("input_domain", "All"))
            ws.cell(row=row_idx, column=5, value=r.get("ground_truth_concept_id", ""))
            ws.cell(row=row_idx, column=6, value=r.get("ground_truth_concept_name", ""))
            ws.cell(row=row_idx, column=7, value="성공" if r.get("success") else "실패")

            correct_cell = ws.cell(row=row_idx, column=8, value="정답" if r.get("mapping_correct") else "오답")
            if r.get("mapping_correct"):
                correct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                correct_cell.font = Font(color="006100")
            else:
                correct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                correct_cell.font = Font(color="9C0006")

            ws.cell(row=row_idx, column=9, value=r.get("best_result_domain", "N/A"))
            ws.cell(row=row_idx, column=10, value=r.get("best_concept_id", "N/A"))
            ws.cell(row=row_idx, column=11, value=r.get("best_concept_name", "N/A"))
            ws.cell(row=row_idx, column=12, value=r.get("best_score", 0.0))
            stage1_text = r.get("stage1_raw", "후보 없음") or "후보 없음"
            stage2_text = r.get("stage2_raw", "후보 없음") or "후보 없음"
            stage3_text = r.get("stage3_raw", "후보 없음") or "후보 없음"
            ws.cell(row=row_idx, column=13, value=stage1_text)
            ws.cell(row=row_idx, column=14, value=stage2_text)
            ws.cell(row=row_idx, column=15, value=stage3_text)
            for c in range(13, 16):
                ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")

        widths = {"A": 10, "B": 18, "C": 40, "D": 15, "E": 20, "F": 45, "G": 10, "H": 12, "I": 18, "J": 15, "K": 45, "L": 12, "M": 70, "N": 70, "O": 85}
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w
        for rn in range(2, len(run_results) + 2):
            ws.row_dimensions[rn].height = 150

    wb.save(out_path)
    return out_path


def infer_csv_path(input_path: Path) -> Path | None:
    """파일명에서 data_type 추론 후 DATA_SOURCES의 csv_path 반환."""
    name = input_path.stem
    if "snuh" in name:
        return Path(DATA_SOURCES["snuh"]["csv_path"])
    if "snomed" in name:
        return Path(DATA_SOURCES["snomed"]["csv_path"])
    return None


def main():
    parser = argparse.ArgumentParser(
        description="매핑 결과(JSON/로그) → Excel 변환. JSON 입력 시 기존 저장 형식과 동일."
    )
    parser.add_argument(
        "input_path",
        type=Path,
        help="JSON 또는 로그 파일 경로 (mapping_snuh_20260304_123456.json / .log)",
    )
    parser.add_argument("--csv", type=Path, default=None, help="CSV 경로 (로그 입력 시, 미지정 시 data_type으로 추론)")
    parser.add_argument("--output", "-o", type=Path, default=None, help="출력 Excel 경로 (미지정 시 입력과 동일 디렉터리)")
    parser.add_argument("--watch", action="store_true", help="10초마다 재변환 후 Excel 갱신 (Ctrl+C로 종료)")
    args = parser.parse_args()

    input_path = args.input_path
    if not input_path.exists():
        print(f"오류: 파일 없음: {input_path}")
        sys.exit(1)

    data_type, timestamp = infer_data_type_and_timestamp(input_path)
    output_dir = args.output.parent if args.output else input_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    csv_path = args.csv or infer_csv_path(input_path)
    if csv_path and not csv_path.exists():
        csv_path = None
        if input_path.suffix.lower() == ".log":
            print("경고: CSV 파일 없음. ground_truth, id 없이 진행합니다.")

    def run_once():
        if input_path.suffix.lower() == ".json":
            print(f"JSON 변환 중: {input_path}")
            out_path = load_json_and_convert(input_path, output_dir, data_type, timestamp)
            if out_path:
                print(f"저장됨: {out_path}")
            return out_path
        else:
            print(f"로그 파싱 중: {input_path}")
            blocks, run_boundaries, num_runs, completed_run_nums = parse_log_blocks(input_path)
            print(f"  추출된 블록: {len(blocks)}개, Run: {num_runs}회, 마무리된 Run: {sorted(completed_run_nums) or '없음'}")

            if csv_path:
                blocks = merge_with_csv(blocks, csv_path)
                print(f"  CSV 병합 완료: {csv_path}")

            blocks = [b for b in blocks if b.get("run_idx", 1) in completed_run_nums]
            if not blocks:
                print("  마무리된 run 없음. Excel 저장 건너뜀.")
                return None

            run_indices = sorted({b.get("run_idx", 1) for b in blocks})
            if not run_indices or (len(run_indices) == 1 and run_indices[0] == 1 and num_runs == 1):
                all_results = [sorted(blocks, key=lambda x: (x.get("test_index", 0), x.get("entity_name", "")))]
            else:
                all_results = []
                for r in run_indices:
                    run_blocks = [b for b in blocks if b.get("run_idx") == r]
                    run_blocks.sort(key=lambda x: (x.get("test_index", 0), x.get("entity_name", "")))
                    all_results.append(run_blocks)

            out_path = save_xlsx_from_log(all_results, output_dir, data_type, timestamp)
            print(f"저장됨: {out_path} (마무리된 Run {len(run_indices)}개)")
            return out_path

    if args.watch:
        import time
        print("Watch 모드: 10초마다 갱신 (Ctrl+C 종료)")
        while True:
            try:
                run_once()
                time.sleep(10)
            except KeyboardInterrupt:
                print("\n종료")
                break
    else:
        run_once()


if __name__ == "__main__":
    main()
