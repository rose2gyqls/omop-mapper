#!/usr/bin/env python3
"""Build the MapOMOP performance / ablation table from the eval/ data.

Reproduces the senior's Table-2 style summary directly from the artefacts that
live under ``eval/`` (no external base dir required):

    eval/_run20/*.json                      -> per-run stage1/2/3 candidate pools
    eval/consensus_evaluation/*.xlsx        -> consensus (S) per-run top-1 scores
    eval/omophub_usagi_baseline/*.xlsx      -> OMOPHub / USAGI single-best scores

Key facts established while inspecting the data (see the design notes below):

* Stage 2 and Stage 3 candidate *sets* are identical in every record (Stage 3
  only re-ranks). Therefore a faithful pool-presence Recall@K makes
  ``+ Stage 2 - Relationship traversal`` == ``+ Stage 3 - Full pipeline``.
  Stage 3's contribution shows up only in Recall@1.
* Human evaluators only scored the *unique top-1 outputs* across the 20 runs
  (the blind-test "terms" files contain <=5 concepts/entity, all of which were
  selected outputs). The deep retrieval pool was never scored, so the only
  concepts that can satisfy Recall@K are ones MapOMOP actually output.

Score scale (raw): 0 / 1 / 2 (Drug additionally uses 0.5).
    E (exact, normalized 1.0)      <-> raw == 2
    A (acceptable, normalized>=.5) <-> raw >= 1   (Drug 0.5 -> 0.25, excluded)

Usage:
    python eval/build_performance_table.py
    python eval/build_performance_table.py --bootstrap 1000 --seed 0
"""

from __future__ import annotations

import argparse
import json
import random
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl


def entity_alias(name: Optional[str]) -> Optional[str]:
    """Strip a trailing parenthetical suffix (matches build_consensus_files).

    Run/JSON entity names sometimes carry a suffix (e.g. ``Glucose(검사24시간가능)``)
    that the baseline blind-test files record without it (``Glucose``).
    """
    if name is None:
        return name
    s = str(name).strip()
    base = re.sub(r"\([^)]*\)$", "", s).strip()
    return base if base and base != s else s

EVAL_DIR = Path(__file__).resolve().parent

# --- thresholds -------------------------------------------------------------
THRESH = {"E": 2.0, "A": 1.0}  # raw-score threshold for exact / acceptable


# --- arm reconstruction -----------------------------------------------------
# relation_type values that belong to the "Maps-to standard concept" stage
# (everything else in stage2 comes from relationship traversal).
MAPS_TO_RELATIONS = {"original", "Maps to"}


def stage1_pool(record: dict, search_type: Optional[str]) -> set:
    out = set()
    for c in record.get("stage1_candidates", []):
        if search_type is None or c.get("search_type") == search_type:
            out.add(str(c.get("concept_id")))
    return out


def stage2_pool(record: dict, maps_to_only: bool) -> set:
    out = set()
    for c in record.get("stage2_candidates", []):
        if maps_to_only and c.get("relation_type") not in MAPS_TO_RELATIONS:
            continue
        out.add(str(c.get("concept_id")))
    return out


def stage3_pool(record: dict) -> set:
    return {str(c.get("concept_id")) for c in record.get("stage3_candidates", [])}


# Ordered list of pipeline-component arms (label -> pool extractor).
ARMS: List[Tuple[str, callable]] = [
    ("Stage 1 - lexical retrieval", lambda r: stage1_pool(r, "lexical")),
    ("Stage 1 - semantic retrieval", lambda r: stage1_pool(r, "semantic")),
    ("Stage 1 - hybrid retrieval", lambda r: stage1_pool(r, None)),
    ("+ Stage 2 - Maps-to standard concept", lambda r: stage2_pool(r, maps_to_only=True)),
    ("+ Stage 2 - Relationship traversal", lambda r: stage2_pool(r, maps_to_only=False)),
    ("+ Stage 3 - Full pipeline", stage3_pool),
]
FULL_ARM = "+ Stage 3 - Full pipeline"


# --- dataset configuration --------------------------------------------------
def _c(name: str) -> Path:
    return EVAL_DIR / "consensus_evaluation" / name


def _b(name: str) -> Path:
    return EVAL_DIR / "omophub_usagi_baseline" / name


def _j(name: str) -> Path:
    return EVAL_DIR / "_run20" / name


# Each unit: one (group, domain) block.
DATASETS = [
    # ---- RWD (SNUH) ----
    dict(group="RWD", domain="Condition", json=_j("mapping_snuh_con.json"),
         json_domain=None,
         consensus=(_c("1_SNUH_condition_consensus.xlsx"), "S_SNUH_Condition"),
         baseline=(_b("1_SNUH_condition_baseline.xlsx"), "SNUH_Condition")),
    dict(group="RWD", domain="Measurement", json=_j("mapping_snuh_meas.json"),
         json_domain=None,
         consensus=(_c("2_SNUH_measurement_drug_procedure_consensus.xlsx"), "S_SNUH_Measurement"),
         baseline=(_b("2_SNUH_measurement_drug_procedure_baseline.xlsx"), "SNUH_Measurement")),
    dict(group="RWD", domain="Drug", json=_j("mapping_snuh_drug.json"),
         json_domain=None,
         consensus=(_c("2_SNUH_measurement_drug_procedure_consensus.xlsx"), "S_SNUH_Drug"),
         baseline=(_b("2_SNUH_measurement_drug_procedure_baseline.xlsx"), "SNUH_Drug")),
    dict(group="RWD", domain="Procedure", json=_j("mapping_snuh_proc.json"),
         json_domain=None,
         consensus=(_c("2_SNUH_measurement_drug_procedure_consensus.xlsx"), "S_SNUH_Procedure"),
         baseline=(_b("2_SNUH_measurement_drug_procedure_baseline.xlsx"), "SNUH_Procedure")),
    # ---- Benchmark (SNOMED) ----
    dict(group="Benchmark", domain="Condition", json=_j("mapping_snomed_merged.json"),
         json_domain="Condition",
         consensus=(_c("3_SNOMED_consensus.xlsx"), "S_SNOMED_Condition"),
         baseline=(_b("3_SNOMED_baseline.xlsx"), "SNOMED_Condition")),
    dict(group="Benchmark", domain="Measurement", json=_j("mapping_snomed_merged.json"),
         json_domain="Measurement",
         consensus=(_c("3_SNOMED_consensus.xlsx"), "S_SNOMED_Measurement"),
         baseline=(_b("3_SNOMED_baseline.xlsx"), "SNOMED_Measurement")),
    dict(group="Benchmark", domain="Observation", json=_j("mapping_snomed_merged.json"),
         json_domain="Observation",
         consensus=(_c("3_SNOMED_consensus.xlsx"), "S_SNOMED_Observation"),
         baseline=(_b("3_SNOMED_baseline.xlsx"), "SNOMED_Observation")),
    dict(group="Benchmark", domain="Procedure", json=_j("mapping_snomed_merged.json"),
         json_domain="Procedure",
         consensus=(_c("3_SNOMED_consensus.xlsx"), "S_SNOMED_Procedure"),
         baseline=(_b("3_SNOMED_baseline.xlsx"), "SNOMED_Procedure")),
]


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------
def _xlsx_rows(path: Path, sheet: str) -> List[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    return rows


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


class Entity:
    """One clinical term with its arm pools, gold concepts and run-1 scores."""

    __slots__ = ("group", "domain", "key", "name", "pools", "gold_E", "gold_A",
                 "run_scores", "baseline", "omophub", "usagi")

    def __init__(self, group, domain, key):
        self.group = group
        self.domain = domain
        self.key = key
        self.name: Optional[str] = None
        self.pools: Dict[str, set] = {label: set() for label, _ in ARMS}
        self.gold_E: set = set()
        self.gold_A: set = set()
        self.run_scores: List[Optional[float]] = []
        # baseline mappability label + external single-best consensus scores
        self.baseline = None
        self.omophub: Optional[float] = None
        self.usagi: Optional[float] = None

    @property
    def evaluable(self) -> bool:
        return any(s is not None for s in self.run_scores)


def load_unit(cfg: dict) -> List[Entity]:
    """Build per-entity records for one (group, domain) unit."""
    runs = json.load(open(cfg["json"], encoding="utf-8"))["runs"]
    dom_filter = cfg["json_domain"]

    # 1) arm pools (union over runs) + per-(test_index, run) best concept id
    ents: Dict[int, Entity] = {}
    best_by_run: Dict[int, Dict[int, str]] = defaultdict(dict)
    for ri, run in enumerate(runs):
        for rec in run:
            if dom_filter and rec.get("input_domain") != dom_filter:
                continue
            ti = rec["test_index"]
            e = ents.get(ti)
            if e is None:
                e = Entity(cfg["group"], cfg["domain"], (cfg["domain"], ti))
                ents[ti] = e
            for label, fn in ARMS:
                e.pools[label] |= fn(rec)
            best_by_run[ti][ri] = str(rec["best_concept_id"])

    # 2) consensus S sheet -> per-run top-1 scores, joined to best concept ids.
    #    Some consensus files use a cumulative global row index (e.g. Drug starts
    #    at 111, Procedure at 1092) while each JSON restarts test_index at 1.
    #    Recover the offset by anchoring on entity names, then join on index.
    cpath, csheet = cfg["consensus"]
    crows = _xlsx_rows(cpath, csheet)
    chdr = crows[0]
    run_cols = [i for i, h in enumerate(chdr) if h and str(h).startswith("run")]

    cidx2scores: Dict[int, List[Optional[float]]] = {}
    cname2idx: Dict[str, int] = {}
    cname_dup: set = set()
    for row in crows[1:]:
        cidx = row[0]
        if cidx is None:
            continue
        cidx2scores[cidx] = [
            _to_float(row[col]) if col < len(row) else None for col in run_cols
        ]
        name = row[1]
        if name in cname2idx:
            cname_dup.add(name)
        else:
            cname2idx[name] = cidx

    # entity_name -> test_index from JSON (unique names only)
    jname2ti: Dict[str, int] = {}
    jname_dup: set = set()
    ti2name: Dict[int, str] = {}
    for run in runs:
        for rec in run:
            if dom_filter and rec.get("input_domain") != dom_filter:
                continue
            ti2name.setdefault(rec["test_index"], rec["entity_name"])
    for ti, name in ti2name.items():
        if name in jname2ti:
            jname_dup.add(name)
        else:
            jname2ti[name] = ti

    # offset = mode of (consensus_index - json_test_index) over shared unique names
    offsets: Dict[int, int] = defaultdict(int)
    for name, cidx in cname2idx.items():
        if name in cname_dup or name in jname_dup:
            continue
        ti = jname2ti.get(name)
        if ti is not None:
            offsets[cidx - ti] += 1
    offset = max(offsets, key=offsets.get) if offsets else 0

    for ti, e in ents.items():
        e.name = ti2name.get(ti)
        scores = cidx2scores.get(ti + offset)
        if scores is None:
            continue
        e.run_scores = scores
        for ci, sc in enumerate(scores):
            cid = best_by_run.get(ti, {}).get(ci)
            if cid is not None and sc is not None:
                if sc >= THRESH["E"]:
                    e.gold_E.add(cid)
                if sc >= THRESH["A"]:
                    e.gold_A.add(cid)

    # 3) baseline mappability label + external single-best scores (join by name)
    bpath, bsheet = cfg["baseline"]
    brows = _xlsx_rows(bpath, bsheet)
    bcol = {h: i for i, h in enumerate(brows[1]) if h is not None}
    # Index baseline rows by exact name and by alias (a few measurement terms
    # carry mismatched suffixes, e.g. baseline "Glucose(24)" vs run
    # "Glucose(검사24시간가능)"; both reduce to "Glucose").
    blabel: Dict[str, object] = {}
    bext: Dict[str, Tuple[Optional[float], Optional[float]]] = {}
    blabel_alias: Dict[str, object] = {}
    bext_alias: Dict[str, Tuple[Optional[float], Optional[float]]] = {}
    for row in brows[2:]:
        nm = row[bcol["entity_name"]]
        if nm is None:
            continue
        ext = (_to_float(row[bcol["omophub_score"]]),
               _to_float(row[bcol["usagi_score"]]))
        blabel[nm] = row[bcol["baseline"]]
        bext[nm] = ext
        blabel_alias[entity_alias(nm)] = row[bcol["baseline"]]
        bext_alias[entity_alias(nm)] = ext
    for e in ents.values():
        if e.name in blabel:
            e.baseline, (e.omophub, e.usagi) = blabel[e.name], bext[e.name]
        else:
            al = entity_alias(e.name)
            if al in blabel_alias:
                e.baseline = blabel_alias[al]
                e.omophub, e.usagi = bext_alias[al]

    return list(ents.values())


# ---------------------------------------------------------------------------
# Metrics
# ---------------------------------------------------------------------------
def _bootstrap_ci(num: Sequence[float], den: Sequence[float], B: int,
                  rng: random.Random) -> Tuple[float, float, float]:
    """Ratio sum(num)/sum(den) with a percentile bootstrap CI (entity units)."""
    n = len(num)
    tot_d = sum(den)
    point = sum(num) / tot_d if tot_d else float("nan")
    if n == 0 or B <= 0:
        return point, point, point
    idx = range(n)
    samples = []
    for _ in range(B):
        pick = [rng.randrange(n) for _ in idx]
        sd = sum(den[i] for i in pick)
        if sd:
            samples.append(sum(num[i] for i in pick) / sd)
    samples.sort()
    lo = samples[int(0.025 * len(samples))]
    hi = samples[int(0.975 * len(samples)) - 1]
    return point, lo, hi


def recall_k(entities: List[Entity], arm: str, level: str, B: int,
             rng: random.Random):
    """Pool-presence Recall@K over the (already filtered) entity set."""
    num, den = [], []
    for e in entities:
        gold = e.gold_E if level == "E" else e.gold_A
        num.append(1.0 if (e.pools[arm] & gold) else 0.0)
        den.append(1.0)
    return _bootstrap_ci(num, den, B, rng)


def recall1_full(entities: List[Entity], level: str, B: int,
                 rng: random.Random):
    """Run-weighted Recall@1 for the full pipeline (per-run top-1 score)."""
    thr = THRESH[level]
    num, den = [], []  # per-entity contributions for entity-level bootstrap
    for e in entities:
        cells = [s for s in e.run_scores if s is not None]
        if not cells:
            continue
        num.append(sum(1.0 for s in cells if s >= thr))
        den.append(float(len(cells)))
    return _bootstrap_ci(num, den, B, rng)


def recall1_external(entities: List[Entity], system: str, level: str, B: int,
                     rng: random.Random):
    """Recall@1 among covered entities for an external baseline + coverage."""
    thr = THRESH[level]
    num, den = [], []
    covered = 0
    for e in entities:
        sc = getattr(e, system)
        if sc is None:  # not in the human-evaluated set -> coverage miss
            continue
        covered += 1
        num.append(1.0 if sc >= thr else 0.0)
        den.append(1.0)
    point, lo, hi = _bootstrap_ci(num, den, B, rng)
    coverage = covered / len(entities) if entities else float("nan")
    return point, lo, hi, coverage


# ---------------------------------------------------------------------------
# Table assembly
# ---------------------------------------------------------------------------
def _fmt(point, lo, hi) -> str:
    if point != point:  # nan
        return "-"
    return f"{point:.3f} [{lo:.3f}-{hi:.3f}]"


def build_table(entities: List[Entity], B: int, seed: int) -> List[dict]:
    rng = random.Random(seed)
    n = len(entities)
    table: List[dict] = []

    # External baselines: Recall@1
    for sys_key, sys_name in [("usagi", "Usagi"), ("omophub", "OMOPHub")]:
        e_pt = recall1_external(entities, sys_key, "E", B, rng)
        a_pt = recall1_external(entities, sys_key, "A", B, rng)
        table.append({
            "section": "External baselines",
            "method": sys_name,
            "n": n,
            "coverage": f"{e_pt[3]:.3f}",
            "r1_E": _fmt(*e_pt[:3]),
            "r1_A": _fmt(*a_pt[:3]),
            "rk_E": "-",
            "rk_A": "-",
        })

    # MapOMOP pipeline components: Recall@K (+ Recall@1 for full pipeline)
    for arm, _ in ARMS:
        rkE = recall_k(entities, arm, "E", B, rng)
        rkA = recall_k(entities, arm, "A", B, rng)
        row = {
            "section": "MapOMOP pipeline components",
            "method": arm,
            "n": n,
            "coverage": "",
            "r1_E": "-",
            "r1_A": "-",
            "rk_E": _fmt(*rkE),
            "rk_A": _fmt(*rkA),
        }
        if arm == FULL_ARM:
            r1E = recall1_full(entities, "E", B, rng)
            r1A = recall1_full(entities, "A", B, rng)
            row["r1_E"] = _fmt(*r1E)
            row["r1_A"] = _fmt(*r1A)
        table.append(row)

    return table


COLUMNS = [
    ("method", "Method", 42),
    ("n", "n", 7),
    ("coverage", "Cov", 7),
    ("r1_E", "Recall@1^E", 22),
    ("r1_A", "Recall@1^A", 22),
    ("rk_E", "Recall@K^E", 22),
    ("rk_A", "Recall@K^A", 22),
]


def print_table(title: str, table: List[dict]) -> None:
    print(f"\n### {title}")
    header = "  ".join(f"{lbl:<{w}}" for _, lbl, w in COLUMNS)
    print(header)
    print("-" * len(header))
    section = None
    for row in table:
        if row["section"] != section:
            section = row["section"]
            print(f"[{section}]")
        print("  ".join(f"{str(row[k]):<{w}}" for k, _, w in COLUMNS))


def write_xlsx(path: Path, tables: Dict[str, List[dict]]) -> None:
    """Write every table stacked onto a single worksheet."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "performance"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    hdr_fill = PatternFill("solid", fgColor="366092")
    hdr_font = Font(bold=True, color="FFFFFF")
    sec_font = Font(bold=True, italic=True, color="1F4E78")
    sec_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    ncol = len(COLUMNS)

    r = 1
    for tname, table in tables.items():
        # block title (merged across all columns)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        tc = ws.cell(row=r, column=1, value=tname)
        tc.fill = title_fill
        tc.font = title_font
        tc.alignment = Alignment(horizontal="left", vertical="center")
        r += 1
        # column header
        for c, (_, lbl, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=r, column=c, value=lbl)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = thin
        r += 1
        section = None
        for row in table:
            if row["section"] != section:
                section = row["section"]
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
                sc = ws.cell(row=r, column=1, value=section)
                sc.font = sec_font
                sc.fill = sec_fill
                r += 1
            for c, (k, _, _) in enumerate(COLUMNS, 1):
                cell = ws.cell(row=r, column=c, value=row[k])
                cell.border = thin
                if c >= 2:
                    cell.alignment = center
            r += 1
        r += 1  # blank spacer row between blocks

    widths = [44, 7, 7, 24, 24, 24, 24]
    for c, wd in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = wd
    ws.freeze_panes = "A1"
    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--bootstrap", type=int, default=1000,
                    help="bootstrap resamples for 95%% CI (0 = disable)")
    ap.add_argument("--seed", type=int, default=0)
    ap.add_argument("--keep-baseline", default="1,5,9",
                    help="comma-separated baseline labels to KEEP (others, e.g. "
                         "the non-evaluable label 0, are dropped)")
    ap.add_argument("--out", type=Path, default=EVAL_DIR / "performance_table.xlsx",
                    help="output xlsx path (single sheet with all blocks)")
    args = ap.parse_args()

    keep = {int(x) for x in str(args.keep_baseline).split(",") if x.strip() != ""}

    def _label_int(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    # Load all units, keep only baseline-evaluable entities, and de-duplicate
    # clinical terms by name *within each baseline file* (the senior's universe).
    # A few terms appear under two domains of the same file (e.g. SNOMED "wt",
    # "exercise") and collapse; terms shared across the two SNUH files (e.g.
    # "Routine gynecologic examination") are distinct evaluations and kept.
    # -> RWD 2876, Benchmark 964, total 3840.
    by_domain_ents: Dict[str, List[Entity]] = defaultdict(list)
    by_group_ents: Dict[str, List[Entity]] = defaultdict(list)
    seen_name: set = set()

    for cfg in DATASETS:
        g = cfg["group"]
        dkey = f"{g}-{cfg['domain']}"
        file_id = str(cfg["baseline"][0])
        for e in load_unit(cfg):
            if _label_int(e.baseline) not in keep:
                continue
            by_domain_ents[dkey].append(e)
            fkey = (file_id, e.name)
            if fkey not in seen_name:
                seen_name.add(fkey)
                by_group_ents[g].append(e)

    # "All" = RWD + Benchmark (already de-duplicated within each group); groups
    # are summed (n = 2876 + 964 = 3840), no cross-group de-duplication.
    all_ents = by_group_ents["RWD"] + by_group_ents["Benchmark"]

    tables: Dict[str, List[dict]] = {}
    tables["All"] = build_table(all_ents, args.bootstrap, args.seed)
    for g in ("RWD", "Benchmark"):
        tables[g] = build_table(by_group_ents[g], args.bootstrap, args.seed)
    for dkey in sorted(by_domain_ents):
        tables[dkey] = build_table(by_domain_ents[dkey], args.bootstrap, args.seed)

    for name in ["All", "RWD", "Benchmark"]:
        print_table(name, tables[name])

    args.out.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(args.out, tables)
    print(f"\nSaved -> {args.out}")


if __name__ == "__main__":
    main()
