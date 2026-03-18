#!/usr/bin/env python3
"""
MapOMOP 정량평가 테이블 생성 (인간 평가자 기반)

20회 매핑 결과 각각을 인간 평가자 점수에 따라 변환하여 정확도 계산.

[기본 3단계: Condition, Procedure, Measurement 등]
- 2점 → 100%, 1점 → 50%, 0점 → 0%
- Acc_2만(%): 2점만 정답
- 가중평균(%): (1, 0.5, 0) 평균 * 100

[Drug 4단계: SNUH]
- 2, 1, 0.5, 0 점수체계
- Acc_2만(%): 2점만 정답 (2 vs 1+0.5+0)
- Acc_2+1(%): 2점 또는 1점 정답 (2+1 vs 0.5+0)
- 가중평균(%): 2→100%, 1→50%, 0.5→25%, 0→0%

지원 구조:
- 다중 데이터셋: SNUH condition, SNUH other(drug/procedure/measurement), SNOMED 1000
- 다중 인간 평가자: 평가자별 시트 자동 감지 (첫 시트 제외)
- 평가자별 비교, 도메인별 분석, 전체 통합 정확도

Usage:
    python eval/compute_quantitative_metrics.py
    python eval/compute_quantitative_metrics.py --config eval/eval_config.yaml
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Optional

_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_root))

import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False


def extract_concept_id(concept_str: str) -> Optional[str]:
    """concept_name(12345) -> 12345"""
    if not concept_str or pd.isna(concept_str):
        return None
    m = re.search(r"\((\d+)\)\s*$", str(concept_str).strip())
    return m.group(1) if m else None


DRUG_DOMAIN = "Drug"


def _parse_score(score_raw) -> Optional[float]:
    """점수 파싱. 2, 1, 0.5, 0 지원."""
    if pd.isna(score_raw):
        return None
    try:
        return float(score_raw)
    except (ValueError, TypeError):
        return None


def _normalize_concept_id(cid) -> Optional[str]:
    """best_concept_id 정규화 (str/int/None)."""
    if cid is None or (isinstance(cid, float) and pd.isna(cid)):
        return None
    s = str(cid).strip()
    if not s or s.lower() == "nan":
        return None
    if isinstance(cid, (int, float)) and not pd.isna(cid):
        return str(int(cid))
    return s


def build_evaluator_scores(
    result_df: pd.DataFrame,
    evaluator_df: pd.DataFrame,
) -> dict[int, dict[str, float]]:
    """
    결과 시트(unique concept) + 평가자 시트(score1~5) → (test_index) -> {concept_id: raw_score}
    """
    concept_cols = sorted([c for c in result_df.columns if re.match(r"^concept\d+$", c)], key=lambda x: int(re.search(r"\d+", x).group()))
    score_cols = sorted([c for c in evaluator_df.columns if re.match(r"^score\d+$", c)], key=lambda x: int(re.search(r"\d+", x).group()))

    entity_scores: dict[int, dict[str, float]] = {}
    merged = result_df.merge(
        evaluator_df[["index"] + score_cols],
        on="index",
        how="inner",
    )

    for _, row in merged.iterrows():
        idx = int(row["index"])
        entity_scores[idx] = {}
        for i, cc in enumerate(concept_cols):
            if i >= len(score_cols):
                break
            concept = row.get(cc)
            score_raw = row.get(score_cols[i])
            if pd.isna(concept) or str(concept).strip() == "":
                continue
            cid = extract_concept_id(str(concept))
            if not cid:
                continue
            s = _parse_score(score_raw)
            if s is not None and s in (0, 0.5, 1, 2):
                entity_scores[idx][cid] = s

    return entity_scores


def load_run_concepts(json_path: Path) -> tuple[list[list[dict]], int]:
    """JSON에서 runs 로드. Returns (all_results, num_runs)."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict) and "runs" in data:
        return data["runs"], len(data["runs"])
    if isinstance(data, list):
        return [data], 1
    raise ValueError("Unsupported JSON format")


def _compute_domain_metrics(
    raw_scores_list: list[float],
    domain: str,
    use_4pt: bool = False,
) -> dict[str, Any]:
    """도메인별 메트릭 계산."""
    valid = [s for s in raw_scores_list if s is not None and s == s]
    n = len(valid)
    if n == 0:
        return {"acc_2": 0.0, "acc_2_1": None, "weighted": 0.0, "n": 0}

    is_4pt = domain == DRUG_DOMAIN or use_4pt
    acc_2 = sum(1 for s in valid if s >= 2.0) / n * 100

    if is_4pt:
        acc_2_1 = sum(1 for s in valid if s >= 1.0) / n * 100
        weighted_sum = sum(
            100 if s >= 2 else (50 if s >= 1 else (25 if s >= 0.5 else 0))
            for s in valid
        )
        return {"acc_2": acc_2, "acc_2_1": acc_2_1, "weighted": weighted_sum / n, "n": n}
    else:
        norm = [1.0 if s >= 2 else (0.5 if s >= 1 else 0.0) for s in valid]
        weighted = sum(norm) / n * 100
        return {"acc_2": acc_2, "acc_2_1": None, "weighted": weighted, "n": n}


def compute_dataset_metrics(
    eval_path: Path,
    json_path: Path,
    dataset_name: str = "",
) -> dict[str, Any]:
    """
    단일 데이터셋에 대해 모든 평가자별 메트릭 계산.
    Returns: {
        "evaluators": { "평가자명": { "by_domain": {...}, "overall": {...}, "raw_scores": [...] } },
        "domains": [...],
        "num_runs": int,
    }
    """
    xl = pd.ExcelFile(eval_path)
    sheet_names = xl.sheet_names

    if len(sheet_names) < 2:
        return {"evaluators": {}, "domains": [], "num_runs": 0, "error": "평가자 시트 없음"}

    result_sheet = sheet_names[0]
    evaluator_sheets = sheet_names[1:]

    result_df = pd.read_excel(xl, sheet_name=result_sheet)
    all_results, num_runs = load_run_concepts(json_path)

    # by test_index: [run0_result, run1_result, ...]
    by_index: dict[int, list[dict]] = {}
    for run_results in all_results:
        for r in run_results:
            idx = r.get("test_index")
            if idx is None:
                continue
            if idx not in by_index:
                by_index[idx] = []
            by_index[idx].append(r)

    out: dict[str, Any] = {
        "evaluators": {},
        "domains": [],
        "num_runs": num_runs,
    }

    for eval_name in evaluator_sheets:
        eval_df = pd.read_excel(xl, sheet_name=eval_name)
        entity_scores = build_evaluator_scores(result_df, eval_df)

        domain_data: dict[str, dict] = {}
        raw_scores_all: list[float] = []
        run_scores_detail: list[dict] = []  # [{test_index, entity_name, domain, run_scores: [s1,s2,...]}]

        for test_index, rows in by_index.items():
            scores_map = entity_scores.get(test_index, {})
            if not scores_map:
                continue

            entity_scores_per_run: list[Optional[float]] = []
            r0 = rows[0]
            entity_name = str(r0.get("entity_name", ""))
            domain = str(r0.get("input_domain", "All")).strip()

            for run_idx in range(min(num_runs, len(rows))):
                r = rows[run_idx]
                dom = str(r.get("input_domain", "All")).strip()
                if domain != dom:
                    domain = dom
                if domain not in domain_data:
                    domain_data[domain] = {"raw_scores": [], "entity_ids": set()}

                cid = _normalize_concept_id(r.get("best_concept_id"))
                raw_score = scores_map.get(cid) if cid else None
                entity_scores_per_run.append(raw_score)

                if raw_score is not None:
                    domain_data[domain]["raw_scores"].append(raw_score)
                    domain_data[domain]["entity_ids"].add(test_index)
                    raw_scores_all.append(raw_score)

            run_scores_detail.append({
                "test_index": test_index,
                "entity_name": entity_name,
                "domain": domain,
                "run_scores": entity_scores_per_run,
            })

        domains = sorted(domain_data.keys())
        out["domains"] = domains
        has_drug = DRUG_DOMAIN in domain_data
        total_raw = raw_scores_all

        by_domain = {}
        for d in domains:
            raw_list = domain_data[d]["raw_scores"]
            m = _compute_domain_metrics(raw_list, d)
            by_domain[d] = m

        total_m = _compute_domain_metrics(total_raw, "All", use_4pt=has_drug)

        out["evaluators"][eval_name] = {
            "by_domain": by_domain,
            "overall": total_m,
            "raw_scores": raw_scores_all,
            "raw_scores_by_domain": {d: domain_data[d]["raw_scores"] for d in domains},
            "has_drug": has_drug,
            "run_scores_detail": run_scores_detail,
        }

    # 평가자별 상이 평가 concept 목록 (같은 concept에 대해 점수가 다른 경우)
    out["score_disagreements"] = _compute_score_disagreements(
        result_df, evaluator_sheets, xl
    )

    return out


def _compute_score_disagreements(
    result_df: pd.DataFrame,
    evaluator_sheets: list[str],
    xl: pd.ExcelFile,
) -> list[dict[str, Any]]:
    """
    concept1~5에 대해 평가자별 점수가 다른 항목 수집.
    Returns: [{entity_name, domain_id, concept, eval1_score, eval2_score, ...}]
    """
    concept_cols = sorted(
        [c for c in result_df.columns if re.match(r"^concept\d+$", c)],
        key=lambda x: int(re.search(r"\d+", x).group()),
    )
    score_cols = [f"score{i+1}" for i in range(len(concept_cols))]

    # 평가자 시트 한 번만 로드
    eval_dfs: dict[str, pd.DataFrame] = {}
    for eval_name in evaluator_sheets:
        eval_dfs[eval_name] = pd.read_excel(xl, sheet_name=eval_name).set_index("index", drop=False)

    disagreements: list[dict[str, Any]] = []

    for _, row in result_df.iterrows():
        entity_name = str(row.get("entity_name", ""))
        domain_id = str(row.get("domain_id", ""))
        idx = row.get("index")

        for i, cc in enumerate(concept_cols):
            concept = row.get(cc)
            if pd.isna(concept) or str(concept).strip() == "":
                continue
            concept_str = str(concept).strip()

            scores: dict[str, Optional[float]] = {}
            for eval_name in evaluator_sheets:
                eval_df = eval_dfs[eval_name]
                if idx not in eval_df.index:
                    continue
                eval_row = eval_df.loc[idx]
                sc = eval_row.get(score_cols[i]) if i < len(score_cols) else None
                s = _parse_score(sc)
                if s is not None:
                    scores[eval_name] = s

            if len(scores) >= 2 and len(set(v for v in scores.values() if v is not None)) > 1:
                rec: dict[str, Any] = {
                    "entity_name": entity_name,
                    "domain_id": domain_id,
                    "concept": concept_str,
                }
                for ev in evaluator_sheets:
                    rec[ev] = scores.get(ev, "")
                disagreements.append(rec)

    return disagreements


# 고정 템플릿: 데이터 소스별 도메인, 평가자 순서
TEMPLATE_DATA_DOMAINS: dict[str, list[str]] = {
    "SNUH": ["Condition", "Measurement", "Procedure", "Drug"],
    "SNOMED": ["Condition", "Measurement", "Procedure", "Observation"],
}
TEMPLATE_EVALUATORS = ["박혜진", "양지영", "안정은"]


def load_config(config_path: Optional[Path]) -> list[dict]:
    """설정 로드. YAML 또는 기본값."""
    if config_path and config_path.exists() and HAS_YAML:
        with open(config_path, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        return cfg.get("datasets", [])
    return [
        {
            "name": "SNUH_condition",
            "data_source": "SNUH",
            "eval_file": "eval/evaluation_blind_test_1_SNUH_condition_용어.xlsx",
            "json_file": "test_logs/SNUH/mapping_snuh_20260308_144131_con.json",
        },
    ]


def run_evaluation(
    config_path: Optional[Path] = None,
    output_path: Optional[Path] = None,
) -> dict[str, Any]:
    """
    전체 평가 실행: 모든 데이터셋, 모든 평가자에 대해 메트릭 계산.
    """
    root = Path(__file__).resolve().parent.parent
    datasets = load_config(config_path)

    all_results: dict[str, Any] = {}
    all_evaluators: set[str] = set()
    all_domains: set[str] = set()
    combined_raw: dict[str, list[float]] = {}  # evaluator -> all raw scores
    combined_by_domain: dict[str, list[float]] = {}  # domain -> all raw scores (모든 평가자)

    for ds in datasets:
        name = ds.get("name", "unknown")
        eval_file = root / ds["eval_file"]
        json_file = root / ds["json_file"]

        if not eval_file.exists():
            print(f"[건너뜀] {name}: eval 파일 없음 - {eval_file}")
            continue
        if not json_file.exists():
            print(f"[건너뜀] {name}: JSON 파일 없음 - {json_file}")
            continue

        data_source = ds.get("data_source", "SNUH" if "SNUH" in name else "SNOMED")
        res = compute_dataset_metrics(eval_file, json_path=json_file, dataset_name=name)
        res["data_source"] = data_source
        all_results[name] = res

        for ev_name, ev_data in res.get("evaluators", {}).items():
            all_evaluators.add(ev_name)
            for d in res.get("domains", []):
                all_domains.add(d)
            if ev_name not in combined_raw:
                combined_raw[ev_name] = []
            combined_raw[ev_name].extend(ev_data.get("raw_scores", []))
            for dom, scores in ev_data.get("raw_scores_by_domain", {}).items():
                if dom not in combined_by_domain:
                    combined_by_domain[dom] = []
                combined_by_domain[dom].extend(scores)

    # 통합 정확도 (모든 평가자 점수 합침)
    total_combined: list[float] = []
    for scores in combined_raw.values():
        total_combined.extend(scores)
    has_4pt = DRUG_DOMAIN in all_domains
    combined_metrics = _compute_domain_metrics(total_combined, "All", use_4pt=has_4pt)

    combined_by_domain_metrics = {
        dom: _compute_domain_metrics(scores, dom, use_4pt=(dom == DRUG_DOMAIN))
        for dom, scores in combined_by_domain.items()
    }

    # (data_source, domain) -> raw_scores for domain/total accuracy
    by_data_domain: dict[tuple[str, str], list[float]] = {}
    by_data_total: dict[str, list[float]] = {}
    cell_data: dict[tuple[str, str, str], dict] = {}  # (data_source, domain, evaluator) -> metrics

    for ds_name, ds_data in all_results.items():
        data_source = ds_data.get("data_source", "SNUH")
        if data_source not in by_data_total:
            by_data_total[data_source] = []
        for ev_name, ev_data in ds_data.get("evaluators", {}).items():
            for domain, dm in ev_data.get("by_domain", {}).items():
                key = (data_source, domain, ev_name)
                cell_data[key] = {
                    "acc_2": dm.get("acc_2", 0),
                    "acc_2_1": dm.get("acc_2_1"),
                    "weighted": dm.get("weighted", 0),
                }
                dk = (data_source, domain)
                if dk not in by_data_domain:
                    by_data_domain[dk] = []
                by_data_domain[dk].extend(ev_data.get("raw_scores_by_domain", {}).get(domain, []))
                by_data_total[data_source].extend(ev_data.get("raw_scores_by_domain", {}).get(domain, []))

    domain_accuracy: dict[tuple[str, str], dict] = {}
    for (ds, dom), scores in by_data_domain.items():
        domain_accuracy[(ds, dom)] = _compute_domain_metrics(scores, dom, use_4pt=(dom == DRUG_DOMAIN))
    total_accuracy: dict[str, dict] = {}
    for ds, scores in by_data_total.items():
        has_drug = any(dom == DRUG_DOMAIN for (d, dom) in by_data_domain if d == ds)
        total_accuracy[ds] = _compute_domain_metrics(scores, "All", use_4pt=has_drug)

    # 평가자별 상이 평가 목록 (데이터셋별)
    score_disagreements: list[dict[str, Any]] = []
    for ds_name, ds_data in all_results.items():
        for rec in ds_data.get("score_disagreements", []):
            rec_with_ds = {"dataset": ds_name, **rec}
            score_disagreements.append(rec_with_ds)

    summary = {
        "by_dataset": all_results,
        "evaluators": sorted(all_evaluators),
        "domains": sorted(all_domains),
        "combined_overall": combined_metrics,
        "combined_by_evaluator": {
            ev: _compute_domain_metrics(scores, "All", use_4pt=has_4pt)
            for ev, scores in combined_raw.items()
        },
        "combined_by_domain": combined_by_domain_metrics,
        "cell_data": cell_data,
        "domain_accuracy": domain_accuracy,
        "total_accuracy": total_accuracy,
        "score_disagreements": score_disagreements,
    }

    if output_path and HAS_OPENPYXL:
        _save_excel(summary, output_path)

    return summary


def _save_excel(summary: dict[str, Any], output_path: Path) -> None:
    """엑셀 저장: 시트1=고정 템플릿 종합평가, 시트2~=평가자별 run별 점수 상세."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    cell_data = summary.get("cell_data", {})
    domain_accuracy = summary.get("domain_accuracy", {})
    total_accuracy = summary.get("total_accuracy", {})

    # ----- 시트1: 고정 템플릿 (데이터 | 도메인 | 평가자 | Acc_2(%) | Acc_2+1(%, drug) | 가중평균(%) | 도메인 정확도 | 총 정확도) -----
    ws1 = wb.active
    ws1.title = "최종 종합 평가"

    headers = ["데이터", "도메인", "평가자", "Acc_2(%)", "Acc_2+1(%, drug)", "가중평균(%)", "도메인 정확도", "총 정확도"]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    row = 2
    merge_ranges: list[tuple[str, int, int]] = []  # (range_str, start_row, end_row)

    for data_source in ["SNUH", "SNOMED"]:
        domains = TEMPLATE_DATA_DOMAINS.get(data_source, [])
        data_start_row = row
        data_end_row = row + len(domains) * len(TEMPLATE_EVALUATORS) - 1

        for domain in domains:
            dom_start = row
            dom_acc = domain_accuracy.get((data_source, domain), {})
            dom_n = dom_acc.get("n", 0) if dom_acc else 0
            dom_acc_2 = round(dom_acc.get("acc_2", 0), 2) if dom_acc and dom_n > 0 else ""
            dom_acc_21 = round(dom_acc["acc_2_1"], 2) if dom_acc and dom_acc.get("acc_2_1") is not None and dom_n > 0 else ""
            dom_weighted = round(dom_acc.get("weighted", 0), 2) if dom_acc and dom_n > 0 else ""

            for evaluator in TEMPLATE_EVALUATORS:
                key = (data_source, domain, evaluator)
                m = cell_data.get(key, {})
                has_cell = key in cell_data and (m.get("acc_2") is not None or m.get("weighted") is not None)
                acc_2 = round(m.get("acc_2", 0), 2) if has_cell else ""
                acc_21 = round(m["acc_2_1"], 2) if has_cell and m.get("acc_2_1") is not None else ""
                weighted = round(m.get("weighted", 0), 2) if has_cell else ""

                total_acc = total_accuracy.get(data_source, {})
                total_n = total_acc.get("n", 0) if total_acc else 0
                total_acc_2 = round(total_acc.get("acc_2", 0), 2) if total_acc and total_n > 0 else ""
                total_acc_21 = round(total_acc["acc_2_1"], 2) if total_acc and total_acc.get("acc_2_1") is not None and total_n > 0 else ""
                total_weighted = round(total_acc.get("weighted", 0), 2) if total_acc and total_n > 0 else ""

                ws1.cell(row=row, column=1, value=data_source if row == data_start_row else "")
                ws1.cell(row=row, column=2, value=domain if row == dom_start else "")
                ws1.cell(row=row, column=3, value=evaluator)
                ws1.cell(row=row, column=4, value=acc_2 if acc_2 != "" else "")
                ws1.cell(row=row, column=5, value=acc_21 if acc_21 != "" else "")
                ws1.cell(row=row, column=6, value=weighted if weighted != "" else "")
                ws1.cell(row=row, column=7, value=(dom_acc_2 if dom_acc_2 != "" else "") if row == dom_start else "")
                ws1.cell(row=row, column=8, value=(total_acc_2 if total_acc_2 != "" else "") if row == data_start_row else "")

                row += 1

            # 도메인 열 병합 (3행)
            merge_ranges.append((f"B{dom_start}:B{row - 1}", dom_start, row - 1))
            # 도메인 정확도 열 병합 (3행)
            merge_ranges.append((f"G{dom_start}:G{row - 1}", dom_start, row - 1))

        # 데이터 열 병합
        merge_ranges.append((f"A{data_start_row}:A{data_end_row}", data_start_row, data_end_row))
        # 총 정확도 열 병합
        merge_ranges.append((f"H{data_start_row}:H{data_end_row}", data_start_row, data_end_row))

    for range_str, sr, er in merge_ranges:
        if sr < er:
            ws1.merge_cells(range_str)

    for col in range(1, 9):
        ws1.column_dimensions[get_column_letter(col)].width = 14

    # ----- 시트2: 평가자별 상이 평가 concept 목록 (논의용) -----
    disagreements = summary.get("score_disagreements", [])
    if disagreements:
        ws_disc = wb.create_sheet(title="평가상이_목록", index=1)
        # 평가자 컬럼: TEMPLATE 순서로, 실제 데이터에 있는 평가자만 포함
        evals_in_data = set()
        for rec in disagreements:
            for k in rec:
                if k in TEMPLATE_EVALUATORS:
                    evals_in_data.add(k)
        all_evals = [e for e in TEMPLATE_EVALUATORS if e in evals_in_data]
        headers_disc = ["dataset", "entity_name", "domain_id", "concept"] + all_evals
        for col, h in enumerate(headers_disc, 1):
            c = ws_disc.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        for r_idx, rec in enumerate(disagreements, 2):
            ws_disc.cell(row=r_idx, column=1, value=rec.get("dataset", ""))
            ws_disc.cell(row=r_idx, column=2, value=rec.get("entity_name", ""))
            ws_disc.cell(row=r_idx, column=3, value=rec.get("domain_id", ""))
            ws_disc.cell(row=r_idx, column=4, value=rec.get("concept", ""))
            for c_idx, ev in enumerate(all_evals, 5):
                val = rec.get(ev, "")
                ws_disc.cell(row=r_idx, column=c_idx, value=val)
        for col in range(1, len(headers_disc) + 1):
            ws_disc.column_dimensions[get_column_letter(col)].width = 14

    # ----- 시트3~: [평가자]_[데이터]_[도메인] run별 점수 상세 -----
    for ds_name, ds_data in summary.get("by_dataset", {}).items():
        data_source = ds_data.get("data_source", "SNUH").lower()
        for ev_name, ev_data in ds_data.get("evaluators", {}).items():
            detail = ev_data.get("run_scores_detail", [])
            if not detail:
                continue
            num_runs = len(detail[0]["run_scores"]) if detail else 0

            # 도메인별로 시트 분리
            by_domain: dict[str, list[dict]] = {}
            for d in detail:
                dom = d.get("domain", "All")
                if dom not in by_domain:
                    by_domain[dom] = []
                by_domain[dom].append(d)

            for domain, rows in by_domain.items():
                sheet_name = f"{ev_name}_{data_source}_{domain.lower()}"[:31]
                ws = wb.create_sheet(title=sheet_name)

                run_cols = [f"run{i+1}" for i in range(num_runs)]
                headers_detail = ["index", "entity_name", "domain"] + run_cols
                for col, h in enumerate(headers_detail, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.font = header_font
                    c.fill = header_fill

                for r_idx, d in enumerate(sorted(rows, key=lambda x: x["test_index"]), 2):
                    ws.cell(row=r_idx, column=1, value=d["test_index"])
                    ws.cell(row=r_idx, column=2, value=d.get("entity_name", ""))
                    ws.cell(row=r_idx, column=3, value=d.get("domain", ""))
                    for j, s in enumerate(d.get("run_scores", [])):
                        val = s if s is not None and s == s else ""
                        ws.cell(row=r_idx, column=4 + j, value=val)

                for col in range(1, len(headers_detail) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(output_path)
    print(f"저장: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="MapOMOP 정량평가 (인간 평가자 기반)")
    parser.add_argument(
        "--config", "-c",
        default=None,
        help="설정 YAML (기본: eval/eval_config.yaml)",
    )
    parser.add_argument(
        "--output", "-o",
        default=str(_root / "eval" / "quantitative_evaluation.xlsx"),
        help="출력 엑셀 경로",
    )
    args = parser.parse_args()

    config_path = Path(args.config) if args.config else _root / "eval" / "eval_config.yaml"
    output_path = Path(args.output)

    summary = run_evaluation(config_path=config_path, output_path=output_path)

    # 콘솔 출력
    print("\n=== 평가자별 정확도 비교 ===")
    for ev in summary["evaluators"]:
        m = summary["combined_by_evaluator"].get(ev, {})
        acc2 = m.get("acc_2", 0)
        wgt = m.get("weighted", 0)
        n = m.get("n", 0)
        print(f"  {ev}: Acc_2만={acc2:.2f}%, 가중평균={wgt:.2f}% (n={n})")

    print("\n=== 통합 정확도 (모든 평가자) ===")
    m = summary["combined_overall"]
    print(f"  Acc_2만={m.get('acc_2', 0):.2f}%, 가중평균={m.get('weighted', 0):.2f}% (n={m.get('n', 0)})")

    if summary.get("combined_by_domain"):
        print("\n=== 도메인별 통합 정확도 ===")
        for dom in sorted(summary["combined_by_domain"].keys()):
            dm = summary["combined_by_domain"][dom]
            extra = f", Acc_2+1={dm['acc_2_1']:.2f}%" if dm.get("acc_2_1") is not None else ""
            print(f"  {dom}: Acc_2만={dm.get('acc_2', 0):.2f}%, 가중평균={dm.get('weighted', 0):.2f}%{extra} (n={dm.get('n', 0)})")

    print("\n=== 데이터셋별 상세 ===")
    for ds_name, ds_data in summary["by_dataset"].items():
        print(f"\n[{ds_name}]")
        for ev_name, ev_data in ds_data.get("evaluators", {}).items():
            o = ev_data.get("overall", {})
            print(f"  {ev_name}: Acc_2만={o.get('acc_2', 0):.2f}%, 가중평균={o.get('weighted', 0):.2f}%")


if __name__ == "__main__":
    main()
