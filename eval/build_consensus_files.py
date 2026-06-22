#!/usr/bin/env python3
"""Generate consensus evaluation Excel files with S/A/Y/P run-20 sheets."""

import argparse
import json
import os
import re
from pathlib import Path
from typing import Optional

import openpyxl

# Working dir (_run20 etc. inputs, consensus_evaluation output) is resolved at runtime (see main)
OUT_DIR: Optional[Path] = None
RUN_HEADERS = ["index", "entity_name", "domain"] + [f"run{i}" for i in range(1, 21)]


def load_rows(path, sheet=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = sheet or wb.sheetnames[0]
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    return rows


def concept_key(name, concept_id):
    return f"{name}({concept_id})"


def normalize_concept(concept_str):
    if concept_str is None:
        return None
    s = str(concept_str).strip()
    m = re.search(r"\((\d+)\)\s*$", s)
    if m:
        name = s[: m.start()].strip().lower()
        return concept_key(name, m.group(1))
    return s.lower()


def entity_alias(name):
    """Map run-sheet entity names to blind-test entity names when they differ."""
    if name is None:
        return name
    s = str(name).strip()
    base = re.sub(r"\([^)]*\)$", "", s).strip()
    return base if base and base != s else s


def load_terms_pairs(terms_path, max_concept_col=10):
    rows = load_rows(terms_path)
    concept_start = next(
        i for i, h in enumerate(rows[0]) if h and "concept" in str(h).lower()
    )
    pairs = []
    for row in rows[1:]:
        if row[1] is None:
            continue
        for ci in range(concept_start, min(len(rows[0]), max_concept_col)):
            if ci < len(row) and row[ci] is not None:
                pairs.append(
                    {
                        "entity_name": row[1],
                        "domain_id": row[2],
                        "concept": normalize_concept(row[ci]),
                    }
                )
    return pairs


def _store_score(scores, entity, concept, value):
    scores[(entity, concept)] = value
    alias = entity_alias(entity)
    if alias != entity:
        scores[(alias, concept)] = value


def load_evaluator_concept_scores(scores_path, terms_path, sheet=None, max_concept_col=10):
    rows = load_rows(scores_path, sheet)
    headers = rows[0]

    if any(h and str(h).startswith("concept1") for h in headers):
        scores = {}
        for row in rows[1:]:
            entity = row[1]
            if entity is None:
                continue
            for i, h in enumerate(headers):
                if h and str(h).startswith("concept") and not str(h).startswith("concept_"):
                    si = i + 1
                    if si < len(row) and row[i] and row[si] is not None and row[si] != "N/A":
                        _store_score(scores, entity, normalize_concept(row[i]), float(row[si]))
        return scores

    term_rows = load_rows(terms_path)
    concept_start = next(
        i for i, h in enumerate(term_rows[0]) if h and "concept" in str(h).lower()
    )
    scores = {}
    for srow in rows[1:]:
        entity = srow[1]
        if entity is None:
            continue
        for trow in term_rows[1:]:
            if trow[1] != entity:
                continue
            for ci in range(concept_start, min(len(trow), max_concept_col)):
                if ci < len(trow) and trow[ci] is not None:
                    sc = srow[ci] if ci < len(srow) else None
                    if sc is not None and sc != "N/A":
                        _store_score(
                            scores, entity, normalize_concept(trow[ci]), float(sc)
                        )
            break
    return scores


def load_consensus_scores(consensus_paths, sheet_names):
    consensus = {}
    for path, sheet_name in zip(consensus_paths, sheet_names):
        rows = load_rows(path, sheet_name)
        headers = [h for h in rows[0]]
        idx = {str(h): i for i, h in enumerate(headers) if h is not None}
        for row in rows[1:]:
            entity = row[idx["entity_name"]]
            concept = row[idx["concept"]]
            final = row[idx["최종"]]
            if entity is None or concept is None or final is None:
                continue
            _store_score(consensus, entity, normalize_concept(concept), float(final))
    return consensus


def build_full_concept_scores(terms_path, evaluator_a_path, evaluator_a_sheet, consensus_paths, consensus_sheets, max_concept_col):
    pairs = load_terms_pairs(terms_path, max_concept_col)
    agreed = load_evaluator_concept_scores(
        evaluator_a_path, terms_path, evaluator_a_sheet, max_concept_col
    )
    disagreed = load_consensus_scores(consensus_paths, consensus_sheets)
    full = dict(agreed)
    full.update(disagreed)

    disagreed_entities = set()
    for (entity, _concept) in disagreed:
        disagreed_entities.add(entity)
        disagreed_entities.add(entity_alias(entity))

    missing = sum(
        1
        for p in pairs
        if (p["entity_name"], p["concept"]) not in full
        and (entity_alias(p["entity_name"]), p["concept"]) not in full
    )
    if missing:
        print(f"  Warning: {missing} concept pairs without score")
    return full, pairs, disagreed_entities


def load_json_runs(json_path):
    with open(json_path, encoding="utf-8") as fp:
        return json.load(fp)["runs"]


def lookup_score(score_map, entity, concept_name, concept_id):
    cid = str(concept_id)
    norm = normalize_concept(concept_key(concept_name, concept_id))
    entities = [entity]
    alias = entity_alias(entity)
    if alias not in entities:
        entities.append(alias)

    for ent in entities:
        if (ent, norm) in score_map:
            return score_map[(ent, norm)]
        for (e, c), sc in score_map.items():
            if e == ent and c and cid in c:
                return sc

    # Single-concept entity fallback
    for ent in entities:
        ent_scores = {c: sc for (e, c), sc in score_map.items() if e == ent}
        if len(ent_scores) == 1:
            return next(iter(ent_scores.values()))
    return None


def read_quant_run_sheet(path, sheet_name):
    rows = load_rows(path, sheet_name)
    run_idx = [i for i, h in enumerate(rows[0]) if h and str(h).startswith("run")]
    out = []
    for row in rows[1:]:
        if row[1] is None:
            continue
        out.append(
            {
                "index": row[0],
                "entity_name": row[1],
                "domain": row[2],
                "runs": [row[i] if i < len(row) else None for i in run_idx],
            }
        )
    return out


def build_consensus_run_sheet(
    runs, quant_rows, score_map, domain_filter=None, fallback_rows=None, disagreed_entities=None
):
    disagreed_entities = disagreed_entities or set()
    entity_order = [r["entity_name"] for r in quant_rows]
    meta = {r["entity_name"]: r for r in quant_rows}
    fallback = {r["entity_name"]: r["runs"] for r in (fallback_rows or [])}
    rows = []
    for entity in entity_order:
        m = meta[entity]
        run_scores = []
        for run_i, run in enumerate(runs):
            mapped = next(
                (
                    item
                    for item in run
                    if item["entity_name"] == entity
                    and (not domain_filter or item["input_domain"] == domain_filter)
                ),
                None,
            )
            if mapped is None:
                fb = fallback.get(entity, [])
                run_scores.append(fb[run_i] if run_i < len(fb) else None)
                continue
            sc = lookup_score(
                score_map,
                entity,
                mapped["best_concept_name"],
                mapped["best_concept_id"],
            )
            if sc is None and entity not in disagreed_entities and entity_alias(entity) not in disagreed_entities:
                fb = fallback.get(entity, [])
                sc = fb[run_i] if run_i < len(fb) else None
            run_scores.append(sc)
        rows.append(
            {
                "index": m["index"],
                "entity_name": entity,
                "domain": m["domain"],
                "runs": run_scores,
            }
        )
    return rows


def run_level_metrics(run_rows, is_drug=False):
    scores = [
        float(sc)
        for row in run_rows
        for sc in row["runs"]
        if sc is not None
    ]
    if not scores:
        return None
    n = len(scores)
    return {
        "acc2": sum(1 for s in scores if s == 2) / n * 100,
        "acc21": sum(1 for s in scores if s >= 1) / n * 100 if is_drug else None,
        "wavg": sum(s / 2 for s in scores) / n * 100,
    }


def concept_level_wavg(score_map, domain, pairs):
    scores = [
        score_map[(p["entity_name"], p["concept"])]
        for p in pairs
        if p["domain_id"] == domain and (p["entity_name"], p["concept"]) in score_map
    ]
    return sum(s / 2 for s in scores) / len(scores) * 100 if scores else None


def write_run_sheet(ws, rows):
    ws.append(RUN_HEADERS)
    for row in rows:
        ws.append([row["index"], row["entity_name"], row["domain"], *row["runs"]])


def write_overall_sheet(ws, dataset_label, domains_info, total_accuracy):
    ws.append(
        [
            "Dataset",
            "Domain",
            "Evaluator",
            "Acc_2(%)",
            "Acc_2+1(%, drug)",
            "Weighted Avg (%)",
            "Domain Accuracy",
            "Total Accuracy",
        ]
    )
    first_row = True
    for info in domains_info:
        domain = info["domain"]
        for j, ev in enumerate(info["evaluators"]):
            ws.append(
                [
                    dataset_label if first_row else None,
                    domain if j == 0 else None,
                    ev["label"],
                    round(ev["acc2"], 2),
                    round(ev["acc21"], 2)
                    if ev["acc21"] is not None
                    else ("-" if domain != "Drug" else None),
                    round(ev["wavg"], 2),
                    round(info["domain_accuracy"], 2) if j == 0 else None,
                    round(total_accuracy, 2) if first_row and j == 0 else None,
                ]
            )
            first_row = False


def generate_file(config):
    print(f"Generating {config['filename']}...")
    concept_scores, pairs, disagreed_entities = build_full_concept_scores(
        config["terms"],
        config["evaluator_a"]["path"],
        config["evaluator_a"]["sheet"],
        config["consensus_paths"],
        config["consensus_sheets"],
        config["max_concept_col"],
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    domains_info = []

    for domain in config["domains"]:
        runs = load_json_runs(config["json_map"][domain])
        is_drug = domain == "Drug"
        qsheet_a = f"A_{config['sheet_prefix']}_{domain.lower()}"
        quant_a = read_quant_run_sheet(config["quant_path"], qsheet_a)
        s_rows = build_consensus_run_sheet(
            runs,
            quant_a,
            concept_scores,
            domain_filter=domain if config["filter_json_domain"] else None,
            fallback_rows=quant_a,
            disagreed_entities=disagreed_entities,
        )

        evaluators = []
        for ev_key, ev_label in [("S", "S"), ("P", "P"), ("A", "A"), ("Y", "Y")]:
            sheet_name = f"{ev_key}_{config['dataset_label']}_{domain}"
            ws = wb.create_sheet(sheet_name)
            if ev_key == "S":
                rows = s_rows
            else:
                qsheet = f"{ev_key}_{config['sheet_prefix']}_{domain.lower()}"
                rows = read_quant_run_sheet(config["quant_path"], qsheet)
            write_run_sheet(ws, rows)
            metrics = run_level_metrics(rows, is_drug)
            evaluators.append(
                {
                    "label": ev_label,
                    "acc2": metrics["acc2"],
                    "acc21": metrics["acc21"],
                    "wavg": metrics["wavg"],
                }
            )

        domains_info.append(
            {
                "domain": domain,
                "domain_accuracy": concept_level_wavg(concept_scores, domain, pairs),
                "evaluators": evaluators,
            }
        )

    total_acc = sum(d["domain_accuracy"] for d in domains_info) / len(domains_info)
    overall_ws = wb.create_sheet("Overall", 0)
    write_overall_sheet(overall_ws, config["dataset_label"], domains_info, total_acc)

    out_path = OUT_DIR / config["filename"]
    wb.save(out_path)
    print(f"  Saved: {out_path} ({len(domains_info)} domains, total_acc={total_acc:.2f})")


def resolve_base() -> Path:
    """Resolve the working directory from --base or the OMOP_EVAL_BASE env var."""
    parser = argparse.ArgumentParser(
        description="Generate consensus evaluation Excel files (S/A/Y/P run-20 sheets)"
    )
    parser.add_argument(
        "--base",
        default=os.environ.get("OMOP_EVAL_BASE"),
        help="Working directory (holds _run20 etc. inputs and the consensus_evaluation output). "
             "Falls back to the OMOP_EVAL_BASE environment variable.",
    )
    args = parser.parse_args()
    if not args.base:
        parser.error(
            "No working directory: pass --base <path> or set the OMOP_EVAL_BASE environment variable"
        )
    return Path(args.base).expanduser()


def main():
    global OUT_DIR
    base = resolve_base()
    OUT_DIR = base / "consensus_evaluation"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    configs = [
        {
            "filename": "1_SNUH_condition_consensus.xlsx",
            "dataset_label": "SNUH",
            "domains": ["Condition"],
            "json_map": {"Condition": base / "_run20/mapping_snuh_con.json"},
            "quant_path": base / "_evaluation_blind_test_1_SNUH_condition/quantitative_evaluation.xlsx",
            "terms": base / "_original_blind_test_file/evaluation_blind_test_1_SNUH_condition_용어.xlsx",
            "consensus_paths": [base / "_human-test-full/1. snuh_condition_quantitative_evaluation_최종.xlsx"],
            "consensus_sheets": ["평가상이_목록"],
            "evaluator_a": {
                "path": base / "_evaluation_blind_test_1_SNUH_condition/evaluation_blind_test_1_SNUH_condition_점수_안정은_20260317.xlsx",
                "sheet": "평가",
            },
            "sheet_prefix": "snuh",
            "filter_json_domain": False,
            "max_concept_col": 8,
        },
        {
            "filename": "2_SNUH_measurement_drug_procedure_consensus.xlsx",
            "dataset_label": "SNUH",
            "domains": ["Measurement", "Drug", "Procedure"],
            "json_map": {
                "Measurement": base / "_run20/mapping_snuh_meas.json",
                "Drug": base / "_run20/mapping_snuh_drug.json",
                "Procedure": base / "_run20/mapping_snuh_proc.json",
            },
            "quant_path": base / "_evaluation_blind_test_2_SNUH_drug_meas_proc/quantitative_evaluation.xlsx",
            "terms": base / "_original_blind_test_file/evaluation_blind_test_2_SNUH_drug_meas_proc_용어.xlsx",
            "consensus_paths": [base / "_human-test-full/2. snuh_meas_proc_drug_quantitative_evaluation_최종.xlsx"],
            "consensus_sheets": ["평가상이_목록"],
            "evaluator_a": {
                "path": base / "_evaluation_blind_test_2_SNUH_drug_meas_proc/evaluation_blind_test_2_SNUH_drug_meas_proc_점수_안정은_20260323.xlsx",
                "sheet": "평가",
            },
            "sheet_prefix": "snuh",
            "filter_json_domain": False,
            "max_concept_col": 8,
        },
        {
            "filename": "3_SNOMED_consensus.xlsx",
            "dataset_label": "SNOMED",
            "domains": ["Condition", "Measurement", "Observation", "Procedure"],
            "json_map": {d: base / "_run20/mapping_snomed_merged.json" for d in ["Condition", "Measurement", "Observation", "Procedure"]},
            "quant_path": base / "_evaluation_blind_test_3_SNOMED/quantitative_evaluation.xlsx",
            "terms": base / "_original_blind_test_file/evaluation_blind_test_3_SNOMED_용어.xlsx",
            "consensus_paths": [
                base / "_human-test-full/3. SNOMED_P_filtering.xlsx",
                base / "_human-test-full/3_SNOMED_filtering_회의후.xlsx",
            ],
            "consensus_sheets": ["평가상이_PY불일치_YA일치", "평가상이_필터후"],
            "evaluator_a": {
                "path": base / "_evaluation_blind_test_3_SNOMED/evaluation_blind_test_3_SNOMED_점수_안정은_20260401.xlsx",
                "sheet": "Sheet1",
            },
            "sheet_prefix": "snomed",
            "filter_json_domain": True,
            "max_concept_col": 10,
        },
    ]
    for cfg in configs:
        generate_file(cfg)


if __name__ == "__main__":
    main()
