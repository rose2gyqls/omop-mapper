#!/usr/bin/env python3
"""Build MapOMOP vs OMOPHub vs USAGI baseline comparison files."""

import json
import re
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_consensus_files import (
    BASE,
    build_full_concept_scores,
    concept_key,
    entity_alias,
    load_evaluator_concept_scores,
    load_json_runs,
    lookup_score,
    normalize_concept,
)

RAW_DIR = BASE / "_human-test-raw"
OUT_DIR = BASE / "omophub_usagi_baseline"

HEADERS = [
    "source_id",
    "domain_id",
    "entity_name",
    "gt_concept_id",
    "gt_concept_name",
    "mapomop_concept_id",
    "mapomop_concept_name",
    "mapomop_domain",
    "mapomop_run_count",
    "mapomop_score",
    "omophub_concept_id",
    "omophub_concept_name",
    "omophub_domain",
    "omophub_match_score",
    "omophub_score",
    "usagi_concept_id",
    "usagi_concept_name",
    "usagi_domain",
    "usagi_match_score",
    "usagi_score",
    "is_same_omophub_usagi",
    "is_same_3models",
    "baseline",
]

# (group label, column count, header fill, header font, data fill)
COLUMN_GROUPS = [
    ("Input Entity", 5, "D9D9D9", "333333", "F7F7F7"),
    ("MapOMOP", 5, "2F75B5", "FFFFFF", "DEEAF6"),
    ("OMOPHub", 5, "548235", "FFFFFF", "E2EFDA"),
    ("USAGI", 5, "C65911", "FFFFFF", "FCE4D6"),
    ("Compare", 2, "7030A0", "FFFFFF", "E4DFEC"),
    ("baseline", 1, "BF8F00", "FFFFFF", "FFF2CC"),
]

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

COL_WIDTHS = {
    "source_id": 10,
    "domain_id": 12,
    "entity_name": 36,
    "gt_concept_id": 14,
    "gt_concept_name": 32,
    "mapomop_concept_id": 14,
    "mapomop_concept_name": 32,
    "mapomop_domain": 12,
    "mapomop_run_count": 14,
    "mapomop_score": 12,
    "omophub_concept_id": 14,
    "omophub_concept_name": 32,
    "omophub_domain": 12,
    "omophub_match_score": 14,
    "omophub_score": 12,
    "usagi_concept_id": 14,
    "usagi_concept_name": 32,
    "usagi_domain": 12,
    "usagi_match_score": 14,
    "usagi_score": 12,
    "is_same_omophub_usagi": 18,
    "is_same_3models": 14,
    "baseline": 10,
}


def norm_entity(name):
    if name is None:
        return None
    return str(name).strip()


def parse_score(val):
    if val is None or val == "null" or val == "N/A":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def parse_concept_id(val):
    if val is None or val == "null":
        return None
    try:
        return str(int(float(val)))
    except (TypeError, ValueError):
        return None


def load_baseline_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["comparison_baseline"].iter_rows(values_only=True))
    wb.close()
    data = []
    for row in rows[2:]:
        if row[2] is None:
            continue
        data.append(
            {
                "source_id": row[0],
                "domain_id": row[1],
                "entity_name": norm_entity(row[2]),
                "gt_concept_id": row[3],
                "gt_concept_name": row[4],
                "omophub_concept_id": parse_concept_id(row[5]),
                "omophub_concept_name": row[6],
                "omophub_domain": row[7],
                "omophub_match_score": row[8],
                "omophub_score": parse_score(row[9]),
                "usagi_concept_id": parse_concept_id(row[10]),
                "usagi_concept_name": row[11],
                "usagi_domain": row[12],
                "usagi_match_score": row[13],
                "usagi_score": parse_score(row[14]),
                "is_same_omophub_usagi": row[15],
                "baseline": row[16] if len(row) > 16 else None,
            }
        )
    return data


def build_entity_run_index(runs, domain_filter=None):
    index = {}
    for run in runs:
        for item in run:
            if domain_filter and item["input_domain"] != domain_filter:
                continue
            ent = norm_entity(item["entity_name"])
            index.setdefault(ent, []).append(item)
    return index


def mode_concept(run_items):
    if not run_items:
        return None
    counts = Counter()
    meta = {}
    for item in run_items:
        cid = str(item["best_concept_id"])
        counts[cid] += 1
        meta[cid] = item
    cid, cnt = counts.most_common(1)[0]
    item = meta[cid]
    return {
        "concept_id": cid,
        "concept_name": item["best_concept_name"],
        "domain": item["best_result_domain"] or item["input_domain"],
        "run_count": cnt,
    }


def compare_three(mapomop_id, omophub_id, usagi_id):
    ids = [x for x in [mapomop_id, omophub_id, usagi_id] if x is not None]
    if len(ids) < 2:
        return None
    return "Same" if len(set(ids)) == 1 else "Different"


def build_run_entity_lookup(runs, domain_filter=None):
    lookup = {}
    for run in runs:
        for item in run:
            if domain_filter and item["input_domain"] != domain_filter:
                continue
            lookup[norm_entity(item["entity_name"])] = norm_entity(item["entity_name"])
            alias = entity_alias(item["entity_name"])
            if alias:
                lookup.setdefault(norm_entity(alias), norm_entity(item["entity_name"]))
    return lookup


def resolve_run_entity(entity_name, run_lookup):
    key = norm_entity(entity_name)
    if key in run_lookup:
        return run_lookup[key]
    alias = entity_alias(key)
    return run_lookup.get(alias, key)


def merge_domain(
    baseline_rows,
    runs,
    consensus_scores,
    domain_filter=None,
    run_lookup=None,
):
    entity_index = build_entity_run_index(runs, domain_filter)
    run_lookup = run_lookup or {}
    merged = []
    for row in baseline_rows:
        ent = row["entity_name"]
        run_ent = resolve_run_entity(ent, run_lookup) if run_lookup else ent
        run_items = entity_index.get(run_ent) or entity_index.get(ent) or entity_index.get(entity_alias(ent))
        mode = mode_concept(run_items)
        mapomop_score = None
        if mode:
            mapomop_score = lookup_score(
                consensus_scores,
                run_ent,
                mode["concept_name"],
                mode["concept_id"],
            )
            if mapomop_score is None:
                mapomop_score = lookup_score(
                    consensus_scores, ent, mode["concept_name"], mode["concept_id"]
                )
        mapomop_id = mode["concept_id"] if mode else None
        merged.append(
            {
                **row,
                "mapomop_concept_id": mapomop_id,
                "mapomop_concept_name": mode["concept_name"] if mode else None,
                "mapomop_domain": mode["domain"] if mode else None,
                "mapomop_run_count": mode["run_count"] if mode else None,
                "mapomop_score": mapomop_score,
                "is_same_3models": compare_three(
                    mapomop_id, row["omophub_concept_id"], row["usagi_concept_id"]
                ),
            }
        )
    return merged


def write_sheet(ws, rows):
    col_styles = []
    col_idx = 1
    for _label, count, header_fill, header_font, data_fill in COLUMN_GROUPS:
        for _ in range(count):
            col_styles.append(
                {
                    "header_fill": PatternFill("solid", fgColor=header_fill),
                    "header_font": Font(bold=True, color=header_font, size=10),
                    "data_fill": PatternFill("solid", fgColor=data_fill),
                    "group_header_fill": PatternFill("solid", fgColor=header_fill),
                    "group_header_font": Font(bold=True, color=header_font, size=11),
                }
            )
            col_idx += 1

    # Row 1: group labels (merged)
    col_start = 1
    for label, count, header_fill, header_font, _data_fill in COLUMN_GROUPS:
        col_end = col_start + count - 1
        ws.merge_cells(
            start_row=1, start_column=col_start, end_row=1, end_column=col_end
        )
        cell = ws.cell(row=1, column=col_start, value=label)
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(bold=True, color=header_font, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        col_start = col_end + 1

    # Row 2: column headers
    for i, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=i, value=header)
        style = col_styles[i - 1]
        cell.fill = style["header_fill"]
        cell.font = style["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        width = COL_WIDTHS.get(header, 14)
        ws.column_dimensions[get_column_letter(i)].width = width

    # Data rows
    same_fill = PatternFill("solid", fgColor="C6EFCE")
    diff_fill = PatternFill("solid", fgColor="FFC7CE")
    for row_offset, r in enumerate(rows, start=3):
        values = [
            r["source_id"],
            r["domain_id"],
            r["entity_name"],
            r["gt_concept_id"],
            r["gt_concept_name"],
            r["mapomop_concept_id"],
            r["mapomop_concept_name"],
            r["mapomop_domain"],
            r["mapomop_run_count"],
            r["mapomop_score"],
            r["omophub_concept_id"],
            r["omophub_concept_name"],
            r["omophub_domain"],
            r["omophub_match_score"],
            r["omophub_score"],
            r["usagi_concept_id"],
            r["usagi_concept_name"],
            r["usagi_domain"],
            r["usagi_match_score"],
            r["usagi_score"],
            r["is_same_omophub_usagi"],
            r["is_same_3models"],
            r["baseline"],
        ]
        for i, value in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=i, value=value)
            style = col_styles[i - 1]
            header_name = HEADERS[i - 1]
            if header_name in ("is_same_omophub_usagi", "is_same_3models"):
                if value == "Same":
                    cell.fill = same_fill
                elif value == "Different":
                    cell.fill = diff_fill
                else:
                    cell.fill = style["data_fill"]
            else:
                cell.fill = style["data_fill"]
            cell.border = THIN_BORDER
            if header_name == "entity_name":
                cell.alignment = Alignment(vertical="center", wrap_text=False)
            elif header_name.endswith("_name"):
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "F3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{max(2, len(rows) + 2)}"


def generate_file(filename, domains_config, consensus_scores_fn):
    print(f"Generating {filename}...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for domain_cfg in domains_config:
        baseline = load_baseline_rows(domain_cfg["raw_path"])
        runs = load_json_runs(domain_cfg["json_path"])
        consensus_scores = consensus_scores_fn(domain_cfg)
        run_lookup = build_run_entity_lookup(
            runs, domain_filter=domain_cfg.get("domain_filter")
        )
        merged = merge_domain(
            baseline,
            runs,
            consensus_scores,
            domain_filter=domain_cfg.get("domain_filter"),
            run_lookup=run_lookup,
        )
        ws = wb.create_sheet(domain_cfg["sheet_name"])
        write_sheet(ws, merged)
        print(f"  {domain_cfg['sheet_name']}: {len(merged)} rows")

    out = OUT_DIR / filename
    wb.save(out)
    print(f"  Saved: {out}")


def main():
    OUT_DIR.mkdir(exist_ok=True)

    snuh_condition_scores = build_full_concept_scores(
        BASE / "_original_blind_test_file/evaluation_blind_test_1_SNUH_condition_용어.xlsx",
        BASE / "_evaluation_blind_test_1_SNUH_condition/evaluation_blind_test_1_SNUH_condition_점수_안정은_20260317.xlsx",
        "평가",
        [BASE / "_human-test-full/1. snuh_condition_quantitative_evaluation_최종.xlsx"],
        ["평가상이_목록"],
        8,
    )[0]

    snuh_other_scores = build_full_concept_scores(
        BASE / "_original_blind_test_file/evaluation_blind_test_2_SNUH_drug_meas_proc_용어.xlsx",
        BASE / "_evaluation_blind_test_2_SNUH_drug_meas_proc/evaluation_blind_test_2_SNUH_drug_meas_proc_점수_안정은_20260323.xlsx",
        "평가",
        [BASE / "_human-test-full/2. snuh_meas_proc_drug_quantitative_evaluation_최종.xlsx"],
        ["평가상이_목록"],
        8,
    )[0]

    snomed_scores = build_full_concept_scores(
        BASE / "_original_blind_test_file/evaluation_blind_test_3_SNOMED_용어.xlsx",
        BASE / "_evaluation_blind_test_3_SNOMED/evaluation_blind_test_3_SNOMED_점수_안정은_20260401.xlsx",
        "Sheet1",
        [
            BASE / "_human-test-full/3. SNOMED_P_filtering.xlsx",
            BASE / "_human-test-full/3_SNOMED_filtering_회의후.xlsx",
        ],
        ["평가상이_PY불일치_YA일치", "평가상이_필터후"],
        10,
    )[0]

    generate_file(
        "1_SNUH_condition_baseline.xlsx",
        [
            {
                "sheet_name": "SNUH_Condition",
                "raw_path": RAW_DIR / "SNUH_condition_평가자P_comparison_baseline.xlsx",
                "json_path": BASE / "_run20/mapping_snuh_con.json",
            }
        ],
        lambda _: snuh_condition_scores,
    )

    generate_file(
        "2_SNUH_measurement_drug_procedure_baseline.xlsx",
        [
            {
                "sheet_name": "SNUH_Measurement",
                "raw_path": RAW_DIR / "SNUH_measurement_평가자P_comparison_baseline.xlsx",
                "json_path": BASE / "_run20/mapping_snuh_meas.json",
            },
            {
                "sheet_name": "SNUH_Drug",
                "raw_path": RAW_DIR / "SNUH_drug_평가자P_comparison_baseline.xlsx",
                "json_path": BASE / "_run20/mapping_snuh_drug.json",
            },
            {
                "sheet_name": "SNUH_Procedure",
                "raw_path": RAW_DIR / "SNUH_procedure_평가자P_comparison_baseline.xlsx",
                "json_path": BASE / "_run20/mapping_snuh_proc.json",
            },
        ],
        lambda _: snuh_other_scores,
    )

    snomed_json = BASE / "_run20/mapping_snomed_merged.json"
    generate_file(
        "3_SNOMED_baseline.xlsx",
        [
            {
                "sheet_name": "SNOMED_Condition",
                "raw_path": RAW_DIR / "SNOMED_condition_평가자P_comparison_baseline.xlsx",
                "json_path": snomed_json,
                "domain_filter": "Condition",
            },
            {
                "sheet_name": "SNOMED_Measurement",
                "raw_path": RAW_DIR / "SNOMED_measurement_평가자P_comparison_baseline.xlsx",
                "json_path": snomed_json,
                "domain_filter": "Measurement",
            },
            {
                "sheet_name": "SNOMED_Observation",
                "raw_path": RAW_DIR / "SNOMED_observation_평가자P_comparison_baseline.xlsx",
                "json_path": snomed_json,
                "domain_filter": "Observation",
            },
            {
                "sheet_name": "SNOMED_Procedure",
                "raw_path": RAW_DIR / "SNOMED_procedure_평가자P_comparison_baseline.xlsx",
                "json_path": snomed_json,
                "domain_filter": "Procedure",
            },
        ],
        lambda _: snomed_scores,
    )


if __name__ == "__main__":
    main()
