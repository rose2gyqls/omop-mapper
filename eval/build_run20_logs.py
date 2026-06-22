#!/usr/bin/env python3
"""Generate run-20 mapping log Excel files from _run20 JSON (same format as mapping_common.save_xlsx_repeat)."""

import argparse
import json
import os
import sys
from copy import deepcopy
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Project root (omop-mapper), derived from __file__ (parent of eval/)
MAPPER_ROOT = Path(__file__).resolve().parent.parent

# Working dir (_run20 input JSON, run20_logs output) is resolved at runtime (see main)
RUN20_DIR: Optional[Path] = None
OUT_DIR: Optional[Path] = None

# mapping_common lives in the scripts/ folder
sys.path.insert(0, str(MAPPER_ROOT / "scripts"))
from mapping_common import (  # noqa: E402
    SUMMARY_BASE_HEADERS,
    XLSX_HEADERS,
    _sort_stage1_by_score,
    _sort_stage2_by_score,
    _sort_stage3_by_score,
)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

DETAIL_WIDTHS = {
    "A": 10,
    "B": 18,
    "C": 40,
    "D": 15,
    "E": 20,
    "F": 45,
    "G": 10,
    "H": 12,
    "I": 18,
    "J": 15,
    "K": 45,
    "L": 12,
    "M": 70,
    "N": 70,
    "O": 95,
}

DOMAIN_FILES = [
    ("SNUH_Condition_run20_log.xlsx", "mapping_snuh_con.json", None, "SNUH"),
    ("SNUH_Measurement_run20_log.xlsx", "mapping_snuh_meas.json", None, "SNUH"),
    ("SNUH_Drug_run20_log.xlsx", "mapping_snuh_drug.json", None, "SNUH"),
    ("SNUH_Procedure_run20_log.xlsx", "mapping_snuh_proc.json", None, "SNUH"),
    ("SNOMED_Condition_run20_log.xlsx", "mapping_snomed_merged.json", "Condition", "SNOMED"),
    ("SNOMED_Measurement_run20_log.xlsx", "mapping_snomed_merged.json", "Measurement", "SNOMED"),
    ("SNOMED_Observation_run20_log.xlsx", "mapping_snomed_merged.json", "Observation", "SNOMED"),
    ("SNOMED_Procedure_run20_log.xlsx", "mapping_snomed_merged.json", "Procedure", "SNOMED"),
]


def _format_candidates_for_cell(candidates: List[dict], stage_type: str) -> str:
    """mapping_common._format_candidates_for_cell + stage3 llm_reasoning."""
    if not candidates:
        return "No candidates"

    if stage_type == "stage1":
        sorted_candidates = _sort_stage1_by_score(candidates)
    elif stage_type == "stage2":
        sorted_candidates = _sort_stage2_by_score(candidates)
    else:
        sorted_candidates = _sort_stage3_by_score(candidates)

    lines = []
    max_show = 15 if stage_type in ("stage1", "stage2") else 10

    for i, c in enumerate(sorted_candidates[:max_show], 1):
        st = c.get("search_type", "unknown")
        name = c.get("concept_name", "N/A")
        cid = c.get("concept_id", "N/A")

        if stage_type == "stage1":
            es_score = float(c.get("elasticsearch_score") or c.get("_score") or 0)
            line = f"{i}. [{st}] {name} (ID: {cid})\n"
            line += (
                f"   ES score: {es_score:.4f}, Standard: {c.get('standard_concept', 'N/A')}, "
                f"Domain: {c.get('domain_id', 'N/A')}"
            )
        elif stage_type == "stage2":
            is_std = "✓" if c.get("is_original_standard", True) else "→"
            line = f"{i}. [{st}] {is_std} {name} (ID: {cid})\n"
            line += f"   Standard: {c.get('standard_concept', 'N/A')}, Domain: {c.get('domain_id', 'N/A')}"
            if not c.get("is_original_standard", True):
                ons = c.get("original_non_standard", {})
                if ons:
                    line += (
                        f"\n   Original Non-std: {ons.get('concept_name', 'N/A')} "
                        f"(ID: {ons.get('concept_id', 'N/A')})"
                    )
        else:
            fin = float(c.get("final_score") or 0)
            sem = c.get("semantic_similarity")
            line = f"{i}. [{st}] {name} (ID: {cid})\n"
            line += f"   Final: {fin:.1f}"
            if sem is not None:
                line += f", Semantic: {sem:.4f}"
            line += f", Standard: {c.get('standard_concept', 'N/A')}, Domain: {c.get('domain_id', 'N/A')}"
            reasoning = c.get("llm_reasoning")
            if reasoning:
                line += f"\n   LLM reasoning: {reasoning}"

        lines.append(line)

    return "\n\n".join(lines)


def load_snuh_gt_lookup() -> Dict[Tuple[str, str], str]:
    """Build a (entity_name, domain) -> concept_name lookup from the SNUH baseline CSV."""
    csv_path = MAPPER_ROOT / "data/snuh-baseline-mapping-data.csv"
    if not csv_path.exists():
        return {}
    df = pd.read_csv(csv_path)
    lookup = {}
    for _, row in df.iterrows():
        entity = str(row.get("source_value", "")).strip()
        domain = str(row.get("domain_id", "")).strip()
        name = row.get("concept_name")
        if entity and domain and pd.notna(name):
            lookup[(entity, domain)] = str(name).strip()
    return lookup


def enrich_result_gt_name(result: dict, gt_lookup: Dict[Tuple[str, str], str]) -> dict:
    if result.get("ground_truth_concept_name"):
        return result
    if not result.get("ground_truth_concept_id"):
        return result
    key = (str(result.get("entity_name", "")).strip(), str(result.get("input_domain", "")).strip())
    gt_name = gt_lookup.get(key)
    if not gt_name:
        return result
    enriched = deepcopy(result)
    enriched["ground_truth_concept_name"] = gt_name
    return enriched


def load_runs(
    json_path: Path,
    domain_filter: Optional[str] = None,
    gt_lookup: Optional[Dict[Tuple[str, str], str]] = None,
) -> List[List[dict]]:
    with open(json_path, encoding="utf-8") as fp:
        data = json.load(fp)
    runs = data["runs"]
    if domain_filter:
        runs = [
            [item for item in run if item.get("input_domain") == domain_filter]
            for run in runs
        ]
    if gt_lookup:
        runs = [
            [enrich_result_gt_name(item, gt_lookup) for item in run]
            for run in runs
        ]
    return runs


def _record_id(result: dict) -> str:
    return result.get("id", result.get("row_id", result.get("snuh_id", result.get("note_id", "N/A"))))


def _write_detail_row(ws, row_idx: int, result: dict) -> None:
    ws.cell(row=row_idx, column=1, value=result.get("test_index", ""))
    ws.cell(row=row_idx, column=2, value=_record_id(result))
    ws.cell(row=row_idx, column=3, value=result.get("entity_name", ""))
    ws.cell(row=row_idx, column=4, value=result.get("input_domain", "All"))
    ws.cell(row=row_idx, column=5, value=result.get("ground_truth_concept_id", ""))
    ws.cell(row=row_idx, column=6, value=result.get("ground_truth_concept_name", ""))
    ws.cell(row=row_idx, column=7, value="Success" if result.get("success") else "Fail")

    correct_cell = ws.cell(
        row=row_idx, column=8, value="Correct" if result.get("mapping_correct") else "Incorrect"
    )
    if result.get("mapping_correct"):
        correct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        correct_cell.font = Font(color="006100")
    else:
        correct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        correct_cell.font = Font(color="9C0006")

    ws.cell(row=row_idx, column=9, value=result.get("best_result_domain", "N/A"))
    ws.cell(row=row_idx, column=10, value=result.get("best_concept_id", "N/A"))
    ws.cell(row=row_idx, column=11, value=result.get("best_concept_name", "N/A"))
    ws.cell(row=row_idx, column=12, value=result.get("best_score", 0.0))

    ws.cell(row=row_idx, column=13, value=_format_candidates_for_cell(result.get("stage1_candidates", []), "stage1"))
    ws.cell(row=row_idx, column=14, value=_format_candidates_for_cell(result.get("stage2_candidates", []), "stage2"))
    ws.cell(row=row_idx, column=15, value=_format_candidates_for_cell(result.get("stage3_candidates", []), "stage3"))

    for col in range(13, 16):
        ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")


def write_repeat_workbook(all_results: List[List[dict]]) -> openpyxl.Workbook:
    """Single-domain workbook: Summary + per-run detail sheets (save_xlsx_repeat format)."""
    num_runs = len(all_results)
    wb = openpyxl.Workbook()

    summary_headers = SUMMARY_BASE_HEADERS + [f"Mapped Concept {i}" for i in range(1, num_runs + 1)]
    ws_summary = wb.active
    ws_summary.title = "Summary"
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    by_index: dict = {}
    for run_results in all_results:
        for result in run_results:
            idx = result.get("test_index")
            by_index.setdefault(idx, []).append(result)

    for row_idx, test_index in enumerate(sorted(by_index.keys(), key=lambda x: (x is None, x)), 2):
        rows = by_index[test_index]
        first = rows[0]
        ws_summary.cell(row=row_idx, column=1, value=test_index)
        ws_summary.cell(row=row_idx, column=2, value=_record_id(first))
        ws_summary.cell(row=row_idx, column=3, value=first.get("entity_name", ""))
        ws_summary.cell(row=row_idx, column=4, value=first.get("input_domain", "All"))
        ws_summary.cell(row=row_idx, column=5, value=first.get("ground_truth_concept_id", ""))
        ws_summary.cell(row=row_idx, column=6, value=first.get("ground_truth_concept_name", ""))

        concept_ids = [str(r.get("best_concept_id") or "") for r in rows]
        all_same = len(set(concept_ids)) == 1 if concept_ids else False
        ws_summary.cell(row=row_idx, column=7, value="Y" if all_same else "N")

        correct_count = sum(1 for r in rows if r.get("mapping_correct"))
        ws_summary.cell(row=row_idx, column=8, value=correct_count)

        for i, result in enumerate(rows[:num_runs]):
            cid = result.get("best_concept_id") or "N/A"
            cname = result.get("best_concept_name") or "N/A"
            ws_summary.cell(row=row_idx, column=9 + i, value=f"{cname}({cid})")

    for col_letter, width in {"A": 10, "B": 18, "C": 40, "D": 15, "E": 20, "F": 45, "G": 10, "H": 8}.items():
        ws_summary.column_dimensions[col_letter].width = width
    for i in range(num_runs):
        ws_summary.column_dimensions[get_column_letter(9 + i)].width = 50

    for run_idx, run_results in enumerate(all_results, 1):
        ws = wb.create_sheet(title=f"Mapping {run_idx}")
        for col, header in enumerate(XLSX_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

        for row_idx, result in enumerate(run_results, 2):
            _write_detail_row(ws, row_idx, result)

        for letter, width in DETAIL_WIDTHS.items():
            ws.column_dimensions[letter].width = width
        for row_num in range(2, len(run_results) + 2):
            ws.row_dimensions[row_num].height = 180

    return wb


def generate_file(
    filename: str,
    json_name: str,
    domain_filter: Optional[str],
    dataset: str,
    gt_lookup: Optional[Dict[Tuple[str, str], str]],
) -> Path:
    print(f"Generating {filename}...")
    json_path = RUN20_DIR / json_name
    runs = load_runs(json_path, domain_filter, gt_lookup if dataset == "SNUH" else None)
    print(f"  runs={len(runs)}, entities/run={len(runs[0]) if runs else 0}")

    wb = write_repeat_workbook(runs)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / filename
    wb.save(out_path)
    print(f"  Saved: {out_path} ({len(wb.sheetnames)} sheets)")
    return out_path


def resolve_base() -> Path:
    """Resolve the working directory from --base or the OMOP_EVAL_BASE env var."""
    parser = argparse.ArgumentParser(
        description="Generate run-20 mapping log Excel files (_run20 JSON -> run20_logs xlsx)"
    )
    parser.add_argument(
        "--base",
        default=os.environ.get("OMOP_EVAL_BASE"),
        help="Working directory (holds the _run20 input JSON and run20_logs output). "
             "Falls back to the OMOP_EVAL_BASE environment variable.",
    )
    args = parser.parse_args()
    if not args.base:
        parser.error(
            "No working directory: pass --base <path> or set the OMOP_EVAL_BASE environment variable"
        )
    return Path(args.base).expanduser()


def main():
    global RUN20_DIR, OUT_DIR
    base = resolve_base()
    RUN20_DIR = base / "_run20"
    OUT_DIR = base / "run20_logs"

    gt_lookup = load_snuh_gt_lookup()
    print(f"SNUH GT name lookup loaded: {len(gt_lookup)} entries")

    for filename, json_name, domain_filter, dataset in DOMAIN_FILES:
        generate_file(filename, json_name, domain_filter, dataset, gt_lookup)

    # Remove the previous 3-file version
    for old in [
        "1_SNUH_condition_run20_log.xlsx",
        "2_SNUH_measurement_drug_procedure_run20_log.xlsx",
        "3_SNOMED_run20_log.xlsx",
    ]:
        old_path = OUT_DIR / old
        if old_path.exists():
            old_path.unlink()
            print(f"Removed old file: {old_path.name}")


if __name__ == "__main__":
    main()
