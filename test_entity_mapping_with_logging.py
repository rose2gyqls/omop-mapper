import pandas as pd
import logging
import os
from datetime import datetime
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.append('/home/work/skku/hyo/omop-mapper/src')

from omop_mapper.entity_mapping_api import EntityMappingAPI, EntityInput, DomainID
from omop_mapper.elasticsearch_client import ElasticsearchClient

class EntityMappingTester:
    def __init__(self, log_dir: str = "test_logs"):
        """테스터 초기화"""
        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(exist_ok=True)
        
        self.setup_logging()
        
        self.es_client = ElasticsearchClient()
        self.es_client.concept_index = "concept-small"
        self.es_client.concept_synonym_index = "concept-synonym"
        
        self.api = EntityMappingAPI(es_client=self.es_client)
        
        self.domain_mapping = {
            'Condition': DomainID.CONDITION,
            'Procedure': DomainID.PROCEDURE,
            'Drug': DomainID.DRUG,
            'Observation': DomainID.OBSERVATION,
            'Measurement': DomainID.MEASUREMENT,
            'Period': DomainID.PERIOD,
            'Provider': DomainID.PROVIDER,
            'Device': DomainID.DEVICE,
        }
    
    def setup_logging(self):
        """로깅 설정"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        log_file = self.log_dir / f"entity_mapping_test_{timestamp}.log"
        
        self.logger = logging.getLogger('entity_mapping_test')
        self.logger.setLevel(logging.INFO)
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        api_logger = logging.getLogger('omop_mapper.entity_mapping_api')
        api_logger.setLevel(logging.INFO)
        api_logger.addHandler(file_handler)
        
        stage1_logger = logging.getLogger('omop_mapper.mapping_stages.stage1_candidate_retrieval')
        stage1_logger.setLevel(logging.INFO)
        stage1_logger.addHandler(file_handler)
        
        stage2_logger = logging.getLogger('omop_mapper.mapping_stages.stage2_standard_collection')
        stage2_logger.setLevel(logging.INFO)
        stage2_logger.addHandler(file_handler)
        
        stage3_logger = logging.getLogger('omop_mapper.mapping_stages.stage3_hybrid_scoring')
        stage3_logger.setLevel(logging.INFO)
        stage3_logger.addHandler(file_handler)
        
        self.logger.info(f"로그 파일: {log_file}")
    
    def load_test_data_from_list(self, entity_list: list) -> pd.DataFrame:
        """리스트에서 테스트 데이터 생성"""
        self.logger.info(f"테스트 데이터 생성: {len(entity_list)}개 엔티티")
        
        test_data = []
        for i, entity_name in enumerate(entity_list):
            test_data.append({
                'entity_plain_name': entity_name,
                'sheet': 'manual'  # 수동 입력 표시
            })
            self.logger.info(f"  {i+1}. {entity_name}")
        
        df = pd.DataFrame(test_data)
        self.logger.info(f"전체 테스트 데이터: {len(df)}개 엔티티")
        
        return df
    
    def create_entity_input(self, row) -> EntityInput:
        """DataFrame 행에서 EntityInput 생성"""
        entity_name = str(row['entity_plain_name']).strip()
        
        # 도메인은 기본적으로 None (모든 도메인 검색)
        # 사용자가 entity_domain을 지정한 경우에만 사용
        domain_id = None
        if 'entity_domain' in row and pd.notna(row['entity_domain']):
            domain_str = str(row['entity_domain']).strip()
            if domain_str and domain_str in self.domain_mapping:
                domain_id = self.domain_mapping[domain_str]
        
        return EntityInput(
            entity_name=entity_name,
            domain_id=domain_id,
            vocabulary_id=None
        )
    
    def test_single_entity(self, entity_input: EntityInput, test_index: int, sheet: str) -> dict:
        """단일 엔티티 테스트 (도메인별 결과)"""
        self.logger.info("=" * 100)
        self.logger.info(f"🧪 테스트 #{test_index} (시트 {sheet}): {entity_input.entity_name}")
        self.logger.info("=" * 100)
        
        try:
            # 매핑 수행 (도메인별 결과 반환)
            results = self.api.map_entity(entity_input)
            
            # 단계별 상세 정보 로깅 (마지막 도메인)
            stage1_candidates = []
            stage2_candidates = []
            stage3_candidates = []
            
            if hasattr(self.api, '_last_stage1_candidates') and self.api._last_stage1_candidates:
                stage1_candidates = self.api._last_stage1_candidates
            
            if hasattr(self.api, '_last_stage2_candidates') and self.api._last_stage2_candidates:
                stage2_candidates = self.api._last_stage2_candidates
                self.logger.info("📊 Stage 2 후보군 상세 정보:")
                for i, candidate in enumerate(stage2_candidates, 1):
                    search_type = candidate.get('search_type', 'unknown')
                    is_std = "✓" if candidate['is_original_standard'] else "→"
                    self.logger.info(f"   {i}. [{search_type}] {is_std} {candidate['concept_name']} (ID: {candidate['concept_id']})")
                    self.logger.info(f"      - Domain: {candidate.get('domain_id', 'N/A')}, Vocabulary: {candidate['vocabulary_id']}")
            
            if hasattr(self.api, '_last_rerank_candidates') and self.api._last_rerank_candidates:
                stage3_candidates = self.api._last_rerank_candidates
            
            # 도메인별 결과 정리
            domain_results = []
            if results:
                self.logger.info("\n" + "=" * 100)
                self.logger.info("📊 도메인별 매핑 결과 요약")
                self.logger.info("=" * 100)
                
                for idx, result in enumerate(results, 1):
                    domain_info = {
                        'domain_id': result.domain_id,
                        'mapped_concept_id': result.mapped_concept_id,
                        'mapped_concept_name': result.mapped_concept_name,
                        'mapping_score': result.mapping_score,
                        'mapping_confidence': result.mapping_confidence,
                        'mapping_method': result.mapping_method,
                        'vocabulary_id': result.vocabulary_id
                    }
                    domain_results.append(domain_info)
                    
                    self.logger.info(f"\n{idx}. [{result.domain_id}] 매핑 성공!")
                    self.logger.info(f"   개념: {result.mapped_concept_name} (ID: {result.mapped_concept_id})")
                    self.logger.info(f"   점수: {result.mapping_score:.4f} | 신뢰도: {result.mapping_confidence}")
                    self.logger.info(f"   방법: {result.mapping_method} | Vocabulary: {result.vocabulary_id}")
            
            # 결과 정리 (최고 점수 도메인 선택)
            best_result = max(results, key=lambda x: x.mapping_score) if results else None
            
            # 도메인별 Stage 경로 정보 추출
            domain_stage_paths = {}
            best_search_domain = None
            if hasattr(self.api, '_all_domain_stage_results') and self.api._all_domain_stage_results:
                domain_stage_paths = self.api._all_domain_stage_results
                
                # Best result의 검색 도메인 찾기
                if best_result:
                    for search_domain, stage_info in domain_stage_paths.items():
                        if stage_info.get('result_domain') == best_result.domain_id:
                            # 가장 높은 점수를 가진 검색 도메인 찾기
                            # (같은 결과 도메인이 여러 검색 도메인에서 나올 수 있음)
                            for domain_result in domain_results:
                                if domain_result['domain_id'] == best_result.domain_id and \
                                   domain_result['mapped_concept_id'] == best_result.mapped_concept_id and \
                                   domain_result['mapping_score'] == best_result.mapping_score:
                                    # 이 결과를 낳은 검색 도메인 찾기
                                    for sd, si in domain_stage_paths.items():
                                        if si.get('result_domain') == best_result.domain_id:
                                            best_search_domain = sd
                                            break
                                    break
                            if best_search_domain:
                                break
            
            test_result = {
                'test_index': test_index,
                'sheet': sheet,
                'entity_name': entity_input.entity_name,
                'success': results is not None and len(results) > 0,
                'domain_count': len(results) if results else 0,
                'domain_results': domain_results,
                'domain_stage_paths': domain_stage_paths,
                'best_search_domain': best_search_domain,
                'best_result_domain': best_result.domain_id if best_result else None,
                'best_concept_id': best_result.mapped_concept_id if best_result else None,
                'best_concept_name': best_result.mapped_concept_name if best_result else None,
                'best_score': best_result.mapping_score if best_result else 0.0,
                'best_confidence': best_result.mapping_confidence if best_result else None,
                'stage1_candidates': stage1_candidates,
                'stage2_candidates': stage2_candidates,
                'stage3_candidates': stage3_candidates
            }
            
            if not results:
                self.logger.info(f"❌ 모든 도메인에서 매핑 실패")
            else:
                self.logger.info(f"\n" + "=" * 100)
                self.logger.info(f"📊 최종 요약: {len(results)}개 도메인에서 매핑 성공")
                self.logger.info("=" * 100)
                self.logger.info(f"🏆 최고 점수: [{best_result.domain_id}] {best_result.mapped_concept_name} ({best_result.mapping_score:.4f})")
                
                # 도메인별 Stage 경로 출력
                if hasattr(self.api, '_all_domain_stage_results') and self.api._all_domain_stage_results:
                    self.logger.info(f"\n📈 도메인별 Stage 경로:")
                    for domain_name, stage_info in self.api._all_domain_stage_results.items():
                        self.logger.info(f"  [{domain_name}] Stage1: {stage_info.get('stage1_count', 0)}개 → "
                                       f"Stage2: {stage_info.get('stage2_count', 0)}개 → "
                                       f"Stage3: {stage_info.get('stage3_count', 0)}개")
                self.logger.info("=" * 100)
                
            return test_result
            
        except Exception as e:
            self.logger.error(f"❌ 테스트 오류: {str(e)}")
            return {
                'test_index': test_index,
                'sheet': sheet,
                'entity_name': entity_input.entity_name,
                'success': False,
                'domain_count': 0,
                'domain_results': [],
                'best_search_domain': None,
                'best_result_domain': None,
                'best_concept_id': None,
                'best_concept_name': None,
                'best_score': 0.0,
                'best_confidence': None,
                'error': str(e),
                'stage1_candidates': [],
                'stage2_candidates': [],
                'stage3_candidates': []
            }
    
    def run_test_with_entities(self, entity_list: list, max_entities: int = None):
        """엔티티 리스트로 테스트 실행"""
        self.logger.info("🚀 Entity Mapping API 테스트 시작")
        self.logger.info(f"테스트 엔티티 리스트: {len(entity_list)}개")
        
        # 데이터 생성
        test_data = self.load_test_data_from_list(entity_list)
        
        if max_entities:
            test_data = test_data.head(max_entities)
            self.logger.info(f"테스트 제한: 최대 {max_entities}개 엔티티")
        
        # 테스트 결과 저장
        test_results = []
        successful_tests = 0
        
        for idx, row in test_data.iterrows():
            try:
                entity_input = self.create_entity_input(row)
                result = self.test_single_entity(entity_input, idx + 1, row['sheet'])
                test_results.append(result)
                
                if result['success']:
                    successful_tests += 1
                    
            except Exception as e:
                self.logger.error(f"테스트 #{idx + 1} 처리 오류: {str(e)}")
                continue
        
        # 결과 요약
        total_tests = len(test_results)
        success_rate = (successful_tests / total_tests * 100) if total_tests > 0 else 0
        
        self.logger.info("=" * 100)
        self.logger.info("📊 테스트 결과 요약")
        self.logger.info("=" * 100)
        self.logger.info(f"총 테스트: {total_tests}개")
        self.logger.info(f"성공: {successful_tests}개")
        self.logger.info(f"실패: {total_tests - successful_tests}개")
        self.logger.info(f"성공률: {success_rate:.2f}%")
        
        # 엔티티별 요약
        for i, result in enumerate(test_results, 1):
            status = "✅ 성공" if result['success'] else "❌ 실패"
            self.logger.info(f"  {i}. {result['entity_name']}: {status}")
            if result['success']:
                search_domain = result.get('best_search_domain', 'N/A')
                result_domain = result.get('best_result_domain', 'N/A')
                if search_domain == result_domain:
                    domain_info = f"[{result_domain}]"
                else:
                    domain_info = f"[{search_domain} → {result_domain}]"
                self.logger.info(f"     -> {domain_info} {result.get('best_concept_name', 'N/A')} (점수: {result.get('best_score', 0.0):.4f})")
        
        # 결과를 CSV와 XLSX로 저장
        self.save_results_to_csv(test_results)
        self.save_results_to_xlsx(test_results)
        
        return test_results
    
    def save_results_to_csv(self, test_results: list):
        """테스트 결과를 CSV 파일로 저장 (도메인별 결과 포함)"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_file = self.log_dir / f"test_results_{timestamp}.csv"
        
        # CSV용 데이터 정리 (도메인별 결과 평탄화)
        csv_results = []
        for result in test_results:
            # 기본 정보
            base_info = {
                'test_index': result['test_index'],
                'entity_name': result['entity_name'],
                'success': result['success'],
                'domain_count': result.get('domain_count', 0),
                'best_search_domain': result.get('best_search_domain', 'N/A'),
                'best_result_domain': result.get('best_result_domain', 'N/A'),
                'best_concept_id': result.get('best_concept_id', 'N/A'),
                'best_concept_name': result.get('best_concept_name', 'N/A'),
                'best_score': result.get('best_score', 0.0),
                'best_confidence': result.get('best_confidence', 'N/A')
            }
            csv_results.append(base_info)
        
        df_results = pd.DataFrame(csv_results)
        df_results.to_csv(csv_file, index=False, encoding='utf-8')
        
        self.logger.info(f"📄 테스트 결과 CSV 저장: {csv_file}")
    
    def save_results_to_xlsx(self, test_results: list):
        """테스트 결과를 XLSX 파일로 저장 (stage1, stage3 후보군을 열로 분리)"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_file = self.log_dir / f"test_results_detailed_{timestamp}.xlsx"
        
        # 엑셀 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detailed Results"
        
        # 통합 상세 시트 생성
        self._create_integrated_detail_sheet(ws, test_results)
        
        # 파일 저장
        wb.save(xlsx_file)
        self.logger.info(f"📊 테스트 결과 XLSX 저장: {xlsx_file}")
    
    def _create_integrated_detail_sheet(self, ws, test_results):
        """통합 상세 시트 생성 (모든 엔티티를 하나의 시트에, 도메인별 결과 포함)"""
        
        # 헤더 설정
        headers = [
            "Test Index", "Entity Name", "Success", "Domain Count",
            "Best Search Domain", "Best Result Domain", "Best Concept ID", "Best Concept Name", 
            "Best Score", "Best Confidence",
            "All Domains", "Domain Stage Paths", "Stage1 Candidates", "Stage2 Candidates", "Stage3 Candidates"
        ]
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row, result in enumerate(test_results, 2):
            ws.cell(row=row, column=1, value=result['test_index'])
            ws.cell(row=row, column=2, value=result['entity_name'])
            ws.cell(row=row, column=3, value="성공" if result['success'] else "실패")
            ws.cell(row=row, column=4, value=result.get('domain_count', 0))
            ws.cell(row=row, column=5, value=result.get('best_search_domain', 'N/A'))
            ws.cell(row=row, column=6, value=result.get('best_result_domain', 'N/A'))
            ws.cell(row=row, column=7, value=result.get('best_concept_id', 'N/A'))
            ws.cell(row=row, column=8, value=result.get('best_concept_name', 'N/A'))
            ws.cell(row=row, column=9, value=result.get('best_score', 0.0))
            ws.cell(row=row, column=10, value=result.get('best_confidence', 'N/A'))
            
            # 모든 도메인 결과를 문자열로 변환
            domain_results_text = self._format_domain_results(result.get('domain_results', []))
            ws.cell(row=row, column=11, value=domain_results_text)
            
            # 도메인별 Stage 경로 정보
            stage_paths_text = self._format_stage_paths(result.get('domain_stage_paths', {}))
            ws.cell(row=row, column=12, value=stage_paths_text)
            
            # Stage1 후보군 정보를 문자열로 변환
            stage1_text = self._format_candidates_for_cell(result.get('stage1_candidates', []), 'stage1')
            ws.cell(row=row, column=13, value=stage1_text)
            
            # Stage2 후보군 정보를 문자열로 변환
            stage2_text = self._format_candidates_for_cell(result.get('stage2_candidates', []), 'stage2')
            ws.cell(row=row, column=14, value=stage2_text)
            
            # Stage3 후보군 정보를 문자열로 변환
            stage3_text = self._format_candidates_for_cell(result.get('stage3_candidates', []), 'stage3')
            ws.cell(row=row, column=15, value=stage3_text)
            
            # 셀 스타일 설정 (텍스트 줄바꿈 허용)
            for col in range(11, 16):  # All Domains, Stage Paths, Stage1, Stage2, Stage3 열
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 열 너비 설정
        column_widths = {
            'A': 10,  # Test Index
            'B': 35,  # Entity Name
            'C': 10,  # Success
            'D': 12,  # Domain Count
            'E': 15,  # Best Search Domain
            'F': 15,  # Best Result Domain
            'G': 15,  # Best Concept ID
            'H': 45,  # Best Concept Name
            'I': 12,  # Best Score
            'J': 15,  # Best Confidence
            'K': 50,  # All Domains
            'L': 45,  # Domain Stage Paths
            'M': 70,  # Stage1 Candidates
            'N': 70,  # Stage2 Candidates
            'O': 85   # Stage3 Candidates
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 행 높이 자동 조정 (후보군 정보가 많은 경우)
        for row_num in range(2, len(test_results) + 2):
            ws.row_dimensions[row_num].height = 150  # 충분한 높이 설정
    
    def _format_domain_results(self, domain_results):
        """도메인별 결과를 엑셀 셀용 텍스트로 포맷팅"""
        if not domain_results:
            return "도메인 결과 없음"
        
        lines = []
        for i, domain in enumerate(domain_results, 1):
            line = f"{i}. [{domain.get('domain_id', 'N/A')}] {domain.get('mapped_concept_name', 'N/A')}\n"
            line += f"   ID: {domain.get('mapped_concept_id', 'N/A')}, "
            line += f"Score: {domain.get('mapping_score', 0):.4f}, "
            line += f"Conf: {domain.get('mapping_confidence', 'N/A')}\n"
            line += f"   Vocab: {domain.get('vocabulary_id', 'N/A')}"
            lines.append(line)
        
        return "\n\n".join(lines)
    
    def _format_stage_paths(self, stage_paths):
        """도메인별 Stage 경로를 엑셀 셀용 텍스트로 포맷팅"""
        if not stage_paths:
            return "경로 정보 없음"
        
        lines = []
        for domain_name, stage_info in sorted(stage_paths.items()):
            search_domain = stage_info.get('search_domain', domain_name)
            result_domain = stage_info.get('result_domain', 'N/A')
            
            # 검색 도메인과 결과 도메인이 다른 경우 표시
            if search_domain != result_domain:
                line = f"[{search_domain} → {result_domain}]\n"
            else:
                line = f"[{search_domain}]\n"
            
            line += f"  Stage1: {stage_info.get('stage1_count', 0)}개\n"
            line += f"  Stage2: {stage_info.get('stage2_count', 0)}개\n"
            line += f"  Stage3: {stage_info.get('stage3_count', 0)}개"
            lines.append(line)
        
        return "\n\n".join(lines)
    
    def _format_candidates_for_cell(self, candidates, stage_type):
        """후보군 정보를 엑셀 셀용 텍스트로 포맷팅"""
        if not candidates:
            return "후보 없음"
        
        lines = []
        max_candidates = 15 if stage_type == 'stage1' else (15 if stage_type == 'stage2' else 10)  # Stage1, Stage2는 15개, Stage3는 10개 표시
        
        for i, candidate in enumerate(candidates[:max_candidates], 1):
            if stage_type == 'stage1':
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   ES점수: {candidate.get('elasticsearch_score', 0):.4f}, "
                line += f"Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            elif stage_type == 'stage2':
                search_type = candidate.get('search_type', 'unknown')
                is_std = "✓" if candidate.get('is_original_standard', True) else "→"
                line = f"{i}. [{search_type}] {is_std} {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
                if not candidate.get('is_original_standard', True):
                    original_non_std = candidate.get('original_non_standard', {})
                    if original_non_std:
                        line += f"\n   원본 Non-std: {original_non_std.get('concept_name', 'N/A')} (ID: {original_non_std.get('concept_id', 'N/A')})"
            else:  # stage3
                search_type = candidate.get('search_type', 'unknown')
                line = f"{i}. [{search_type}] {candidate.get('concept_name', 'N/A')} (ID: {candidate.get('concept_id', 'N/A')})\n"
                line += f"   텍스트: {candidate.get('text_similarity', 0):.4f}, "
                line += f"의미적: {candidate.get('semantic_similarity', 0):.4f}, "
                line += f"최종: {candidate.get('final_score', 0):.4f}\n"
                line += f"   Standard: {candidate.get('standard_concept', 'N/A')}, "
                line += f"Domain: {candidate.get('domain_id', 'N/A')}"
            
            lines.append(line)
        
        return "\n\n".join(lines)

def main():
    """메인 함수"""
    tester = EntityMappingTester()
    
    # 테스트할 엔티티 리스트
    test_entities = test_entities = [
        'ST-segment elevation myocardial infarction',
        'acute coronary syndromes',
        'acute myocardial infarction',
        'adrenal incidentaloma',
        'adrenal vein sampling',
        'aldosterone-producing adenoma',
        'aldosterone-to-renin ratio',
        'angiotensin receptor blocker',
        'angiotensin-converting enzyme inhibitor',
        'atherosclerotic cardiovascular disease',
        'atrial fibrillation',
        'blood pressure',
        'cardiac rehabilitation',
        'cardiac troponin',
        'cardiovascular',
        'cardiovascular disease',
        'chronic coronary disease',
        'computed tomography',
        'coronary artery bypass grafting',
        'coronary artery disease',
        'diastolic blood pressure',
        'direct oral anticoagulant',
        'electrocardiogram',
        'fractional flow reserve',
        'glucagon-like peptide-1',
        'glucocorticoid-remediable aldosteronism',
        'heart failure',
        'high-sensitivity cardiac troponin',
        'hypertension',
        'idiopathic hyperaldosteronism',
        'implantable cardioverter-defibrillator',
        'intra-aortic balloon pump',
        'intravascular ultrasound',
        'left ventricular ejection fraction',
        'left ventricular hypertrophy',
        'low-density lipoprotein',
        'low-density lipoprotein cholesterol',
        'major adverse cardiovascular event',
        'mechanical circulatory support',
        'mineralocorticoid receptor antagonist',
        'multivessel disease',
        'non–ST-segment elevation ACS',
        'non–ST-segment elevation myocardial infarction',
        'optical coherence tomography',
        'percutaneous coronary intervention',
        'plasma aldosterone concentration',
        'plasma renin activity',
        'primary aldosteronism',
        'proprotein convertase subtilisin/kexin type 9',
        'primary percutaneous coronary intervention',
        'proton pump inhibitor',
        'randomized controlled trial',
        'return of spontaneous circulation',
        'sodium-glucose cotransporter-2',
        'systolic blood pressure',
        'unfractionated heparin',
        'venoarterial extracorporeal membrane oxygenation',
        'myocardial ischemia',
        'breast cancer',
        'type 2 diabetes',
        # --- 추가 (여기서부터 확장) ---
        'chronic kidney disease',
        'glomerular filtration rate',
        'end-stage renal disease',
        'atrial flutter',
        'ventricular tachycardia',
        'ventricular fibrillation',
        'sudden cardiac death',
        'ischemic stroke',
        'hemorrhagic stroke',
        'pulmonary embolism',
        'deep vein thrombosis',
        'chronic obstructive pulmonary disease',
        'obstructive sleep apnea',
        'acute respiratory distress syndrome',
        'body mass index',
        'fasting plasma glucose',
        'oral glucose tolerance test',
        'glycated hemoglobin',
        'insulin resistance',
        'metabolic syndrome',
        'nonalcoholic fatty liver disease',
        'nonalcoholic steatohepatitis',
        'hepatocellular carcinoma',
        'prostate cancer',
        'colorectal cancer',
    ]
    
    results = tester.run_test_with_entities(test_entities)
    
    print(f"\n✅ 테스트 완료! 로그는 {tester.log_dir} 디렉토리에 저장되었습니다.")

if __name__ == "__main__":
    main()